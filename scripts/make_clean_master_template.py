"""
Build a clean master template from the original 2026-3_Commission B2B.xlsx.

Strategy:
  - Keep B2B Summary fully (only clear hardcoded values like B2B Executive Account = 63316.22
    and replace 'JANUARY 2026' placeholder text).
  - Keep ALL salesperson sheets with their original layouts and styles.
  - Clear data values in every salesperson sheet's detail blocks, preserving formulas.
  - Keep Table + R_LP + R_SO + R_INV + R_SH reference sheets intact.

The builder will then load this template, detect each salesperson sheet's block
layout, fill in current data, and recompute subtotal/summary formulas.
"""

from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SOURCE = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B.xlsx"
)
OUTPUT = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"


SALESPERSON_SHEETS = [
    "Paul", "Jose", "Michael", "Jim", "Weston",
    "Brett", "Leslie", "Carmen", "Garrett", "Company Acct",
]


DATA_VALUE_COLS = [
    4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17,
    19, 20, 21, 22,
    24, 25, 26,
    27, 28,
    31,
    34,
]


def find_subtotal_rows(ws: Worksheet) -> list[int]:
    """Find all rows where col K = 'Subtotal'."""
    rows = []
    for r in range(20, min(ws.max_row + 1, 250)):
        v = ws.cell(r, 11).value
        if v and str(v).lower().startswith("subtotal"):
            rows.append(r)
    return rows


def find_header_rows(ws: Worksheet) -> list[int]:
    """Find all rows where col D = 'Order Date' (data block headers)."""
    rows = []
    for r in range(20, min(ws.max_row + 1, 250)):
        v = ws.cell(r, 4).value
        if v and str(v).strip().lower() == "order date":
            rows.append(r)
    return rows


def clear_data_block(ws: Worksheet, header_row: int, subtotal_row: int) -> None:
    """Clear data values between header (exclusive) and subtotal (exclusive)."""
    for r in range(header_row + 1, subtotal_row):
        for col_idx in DATA_VALUE_COLS:
            cell = ws.cell(r, col_idx)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                continue
            cell.value = None


def clean_salesperson_sheet(ws: Worksheet) -> dict[str, list[int]]:
    """Clean a salesperson sheet: clear data, set placeholders. Return found anchors."""
    # Title placeholders
    ws.cell(2, 4, "[SALESPERSON NAME]")
    ws.cell(3, 4, "COMMISSION REPORT - [MONTH] [YEAR]")
    # Section I label (col C, row 20) — may exist in some sheets
    if ws.cell(20, 3).value:
        ws.cell(20, 3, "SALES ORDERS INVOICED IN [MONTH] [YEAR]")

    headers = find_header_rows(ws)
    subtotals = find_subtotal_rows(ws)

    # Pair each subtotal with the nearest preceding header
    block_pairs: list[tuple[int, int]] = []
    used_headers = set()
    for sub in subtotals:
        candidates = [h for h in headers if h < sub and h not in used_headers]
        if not candidates:
            continue
        header = max(candidates)
        block_pairs.append((header, sub))
        used_headers.add(header)

    for header, sub in block_pairs:
        clear_data_block(ws, header, sub)

    return {"headers": headers, "subtotals": subtotals}


def clean_b2b_summary(ws: Worksheet) -> None:
    """Clear hardcoded values and replace placeholders in B2B Summary."""
    # I18: 'JANUARY 2026 ORDERS - SHIPPED AND INVOICED' -> placeholder
    if ws.cell(18, 9).value:
        ws.cell(18, 9, "[MONTH] [YEAR] ORDERS - SHIPPED AND INVOICED")
    # F38, G38: hardcoded B2B Executive Account values
    ws.cell(38, 6, None)
    ws.cell(38, 7, None)
    # P34, R34: hardcoded zero values
    ws.cell(34, 16, None)
    ws.cell(34, 18, None)


def main() -> None:
    print(f"Loading source: {SOURCE.name}")
    wb = load_workbook(SOURCE)

    # Clean B2B Summary
    if "B2B Summary" in wb.sheetnames:
        print("Cleaning B2B Summary...")
        clean_b2b_summary(wb["B2B Summary"])

    # Clean each salesperson sheet
    for name in SALESPERSON_SHEETS:
        if name not in wb.sheetnames:
            print(f"  WARN: sheet {name!r} not found, skipping")
            continue
        ws = wb[name]
        info = clean_salesperson_sheet(ws)
        print(f"  Cleaned {name}: {len(info['headers'])} headers, {len(info['subtotals'])} subtotals")

    # Fix Jose!AE22 hardcoded 1899.99 -> VLOOKUP formula
    if "Jose" in wb.sheetnames:
        ws = wb["Jose"]
        if isinstance(ws.cell(22, 31).value, (int, float)):
            ws.cell(22, 31, "=IFERROR(VLOOKUP(K22,R_LP!$A:$G,7,FALSE),0)")
            print("  Fixed Jose!AE22: hardcoded -> VLOOKUP")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    wb.close()
    print(f"\nOK: master template saved to {OUTPUT}")
    print(f"  Size: {OUTPUT.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
