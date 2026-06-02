"""
Quick smoke test for src/commission/workbook_builder.

Reads real Brett March 2026 data from the historical workbook,
feeds it into the builder, and writes a generated workbook for comparison.
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from src.commission.workbook_builder import (
    Block,
    DetailRow,
    SalespersonData,
    build_workbook,
)


HISTORICAL_BRETT = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B_BRETT BERN.xlsx"
)

OUTPUT = BASE_DIR / "data" / "output" / "_test_generated_brett_march_2026.xlsx"


def read_block_rows(ws, header_row: int, subtotal_row: int) -> list[DetailRow]:
    """Read data rows from `header_row+1` to `subtotal_row-1` of the Jennifer-style sheet."""
    rows: list[DetailRow] = []
    for r in range(header_row + 1, subtotal_row):
        sku = ws.cell(r, 11).value  # K
        item_total = ws.cell(r, 13).value  # M
        if not sku and not item_total:
            continue
        rows.append(
            DetailRow(
                order_date=ws.cell(r, 4).value,
                sales_order_number=ws.cell(r, 5).value or "",
                invoice_date=ws.cell(r, 6).value,
                invoice_number=ws.cell(r, 7).value or "",
                invoice_status=ws.cell(r, 8).value or "",
                customer_name=ws.cell(r, 9).value or "",
                estimate_number=ws.cell(r, 10).value or "",
                sku=sku or "",
                quantity=float(ws.cell(r, 12).value or 0),
                item_total=float(ws.cell(r, 13).value or 0) if isinstance(ws.cell(r, 13).value, (int, float)) else 0,
                account=ws.cell(r, 14).value or "",
                account_code=str(ws.cell(r, 15).value or ""),
                payment_terms=ws.cell(r, 16).value or "",
                delivery_method=ws.cell(r, 17).value or "",
                shipment_date=ws.cell(r, 19).value,
                shipment_status=ws.cell(r, 20).value or "",
                ar_status=ws.cell(r, 21).value or "",
                payment_date=ws.cell(r, 22).value,
                shipping_method=ws.cell(r, 24).value or "",
                reason=ws.cell(r, 25).value or "",
                include_in_commission=ws.cell(r, 26).value or "No",
                shipping_income=float(ws.cell(r, 27).value or 0) if isinstance(ws.cell(r, 27).value, (int, float)) else 0,
                shipping_expenses=float(ws.cell(r, 28).value or 0) if isinstance(ws.cell(r, 28).value, (int, float)) else 0,
                discount_rate=ws.cell(r, 34).value if isinstance(ws.cell(r, 34).value, (int, float)) else None,
            )
        )
    return rows


def load_brett_data() -> SalespersonData:
    wb = load_workbook(HISTORICAL_BRETT, data_only=False)
    ws = wb["Brett"]

    # Block I.1: header R23, subtotal R45 -> rows 24..43
    i1_rows = read_block_rows(ws, 23, 45)
    # Block I.2 (Shipping Income): header R48, subtotal R53
    i2_rows = read_block_rows(ws, 48, 53)
    # Block I.3 (Other Charges): header R56, subtotal R59
    i3_rows = read_block_rows(ws, 56, 59)
    # Block II.1 (Prior Orders): header R64, subtotal R66
    ii1_rows = read_block_rows(ws, 64, 66)
    # Block II.2 (Prior Shipping): header R69, subtotal R71
    ii2_rows = read_block_rows(ws, 69, 71)

    wb.close()

    return SalespersonData(
        name="Brett",
        full_name="BRETT BERN",
        rate_type="non_salaried",
        month_name="March",
        year=2026,
        current_commissionable=Block(
            label="ORDERS COMMISSIONABLE - SHIPPED AND INVOICED",
            rows=i1_rows,
        ),
        current_shipping=Block(label="SHIPPING INCOME", rows=i2_rows),
        current_other=Block(label="OTHER CHARGES", rows=i3_rows),
        prior_commissionable=Block(label="ORDERS COMMISSIONABLE", rows=ii1_rows),
        prior_shipping=Block(label="SHIPPING CHARGE", rows=ii2_rows),
    )


def load_reference_data() -> tuple[dict, list[tuple[float, float, float]]]:
    """Read R_LP, R_SO, R_INV, R_SH and the Table tier from the source workbook."""
    wb = load_workbook(HISTORICAL_BRETT, data_only=True)

    reference = {}
    for sheet, header_row in [("R_LP", 2), ("R_SO", 2), ("R_INV", 2), ("R_SH", 2)]:
        ws = wb[sheet]
        headers = [
            ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)
            if ws.cell(header_row, c).value
        ]
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            if any(v not in (None, "") for v in row):
                rows.append(row)
        reference[sheet] = {"headers": headers, "rows": rows}

    # Read Table tier
    ws_table = wb["Table"]
    tiers: list[tuple[float, float, float]] = []
    # Find header row by looking for "Discount" in col B
    header_row = None
    for r in range(1, 10):
        v = ws_table.cell(r, 2).value
        if v and "discount" in str(v).lower():
            header_row = r
            break
    if header_row:
        for r in range(header_row + 1, ws_table.max_row + 1):
            disc = ws_table.cell(r, 2).value
            sal = ws_table.cell(r, 3).value
            nonsal = ws_table.cell(r, 4).value
            if isinstance(disc, (int, float)):
                tiers.append((float(disc), float(sal or 0), float(nonsal or 0)))

    wb.close()
    return reference, tiers


def main() -> None:
    print(f"Reading source: {HISTORICAL_BRETT.name}")
    brett = load_brett_data()
    print(f"  I.1 commissionable rows: {len(brett.current_commissionable.rows)}")
    print(f"  I.2 shipping rows:       {len(brett.current_shipping.rows)}")
    print(f"  I.3 other rows:          {len(brett.current_other.rows)}")
    print(f"  II.1 prior comm rows:    {len(brett.prior_commissionable.rows)}")
    print(f"  II.2 prior ship rows:    {len(brett.prior_shipping.rows)}")

    print("\nReading reference data...")
    reference, tiers = load_reference_data()
    print(f"  R_LP rows: {len(reference['R_LP']['rows'])}")
    print(f"  R_SO rows: {len(reference['R_SO']['rows'])}")
    print(f"  R_INV rows: {len(reference['R_INV']['rows'])}")
    print(f"  R_SH rows: {len(reference['R_SH']['rows'])}")
    print(f"  Tier table rows: {len(tiers)}")

    print(f"\nBuilding workbook -> {OUTPUT}")
    result = build_workbook(
        output_path=OUTPUT,
        salespeople=[brett],
        reference=reference,
        tier_table=tiers,
        include_b2b_summary=True,
    )
    print(f"OK: wrote {result}")
    print(f"  Size: {result.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
