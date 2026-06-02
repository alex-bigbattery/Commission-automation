"""
Test the v2 (template-based) builder with real Brett March 2026 data.
"""

from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from src.commission.workbook_builder_v2 import (
    Block,
    DetailRow,
    SalespersonData,
    build_salesperson_workbook,
)


HISTORICAL_BRETT = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B_BRETT BERN.xlsx"
)
TEMPLATE = BASE_DIR / "data" / "templates" / "salesperson_template_clean.xlsx"
OUTPUT = BASE_DIR / "data" / "output" / "_test_v2_brett_march_2026.xlsx"


def read_block_rows(ws, header_row: int, subtotal_row: int) -> list[DetailRow]:
    rows: list[DetailRow] = []
    for r in range(header_row + 1, subtotal_row):
        sku = ws.cell(r, 11).value
        item_total = ws.cell(r, 13).value
        if not sku and not item_total:
            continue
        # Safe number conversion
        def num(c):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(v) if v not in (None, "") else 0.0
            except (ValueError, TypeError):
                return 0.0
        rows.append(
            DetailRow(
                order_date=ws.cell(r, 4).value,
                sales_order_number=ws.cell(r, 5).value or "",
                invoice_date=ws.cell(r, 6).value,
                invoice_number=ws.cell(r, 7).value or "",
                invoice_status=ws.cell(r, 8).value or "",
                customer_name=ws.cell(r, 9).value or "",
                estimate_number=ws.cell(r, 10).value or "",
                sku=sku or "",
                quantity=num(12),
                item_total=num(13),
                account=ws.cell(r, 14).value or "",
                account_code=str(ws.cell(r, 15).value or ""),
                payment_terms=ws.cell(r, 16).value or "",
                delivery_method=ws.cell(r, 17).value or "",
                shipment_date=ws.cell(r, 19).value,
                shipment_status=ws.cell(r, 20).value or "",
                ar_status=ws.cell(r, 21).value or "",
                payment_date=ws.cell(r, 22).value,
                shipping_method=ws.cell(r, 24).value or "",
                reason=ws.cell(r, 25).value or "",
                include_in_commission=ws.cell(r, 26).value or "No",
                shipping_income=num(27),
                shipping_expenses=num(28),
                discount_rate=ws.cell(r, 34).value if isinstance(ws.cell(r, 34).value, (int, float)) else None,
            )
        )
    return rows


def main() -> None:
    print(f"Loading source data: {HISTORICAL_BRETT.name}")
    wb = load_workbook(HISTORICAL_BRETT, data_only=False)
    ws = wb["Brett"]

    i1 = read_block_rows(ws, 23, 45)
    i2 = read_block_rows(ws, 48, 53)
    i3 = read_block_rows(ws, 56, 59)
    ii1 = read_block_rows(ws, 64, 66)
    ii2 = read_block_rows(ws, 69, 71)

    wb.close()

    brett = SalespersonData(
        name="Brett",
        full_name="BRETT BERN",
        rate_type="non_salaried",
        month_name="March",
        year=2026,
        current_commissionable=Block("ORDERS COMMISSIONABLE - SHIPPED AND INVOICED", i1),
        current_shipping=Block("SHIPPING INCOME", i2),
        current_other=Block("OTHER CHARGES", i3),
        prior_commissionable=Block("ORDERS COMMISSIONABLE", ii1),
        prior_shipping=Block("SHIPPING CHARGE", ii2),
    )

    print(f"  I.1 commissionable: {len(i1)} rows")
    print(f"  I.2 shipping:       {len(i2)} rows")
    print(f"  I.3 other:          {len(i3)} rows")
    print(f"  II.1 prior comm:    {len(ii1)} rows")
    print(f"  II.2 prior ship:    {len(ii2)} rows")

    print(f"\nBuilding from template: {TEMPLATE.name}")
    result = build_salesperson_workbook(
        template_path=TEMPLATE,
        output_path=OUTPUT,
        data=brett,
    )
    print(f"OK: wrote {result}")
    print(f"  Size: {result.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
