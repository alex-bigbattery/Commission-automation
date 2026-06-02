"""
Build master_template_clean v2 with STANDARDIZED layout per salesperson.

Strategy:
  - Use the clean Brett template (5-block layout) as the canonical salesperson layout.
  - Build a new master workbook with:
      * B2B Summary from the original (cleaned)
      * 10 salesperson sheets, each = a copy of clean Brett sheet, renamed
      * Reference sheets (Table, R_LP, R_SO, R_INV, R_SH) from clean Brett

  The B2B Summary formulas all point to fixed positions (F7, F11, F14, etc.) of
  each salesperson sheet, so they keep working regardless of detail layout.
"""

from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ORIGINAL_MASTER = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B.xlsx"
)
BRETT_CLEAN = BASE_DIR / "data" / "templates" / "salesperson_template_clean.xlsx"
OUTPUT = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"


# In order of B2B Summary listing (salaried first, then non-salaried, then Company Acct)
SALESPERSON_SHEETS = [
    "Paul",
    "Jose",
    "Michael",
    "Jim",
    "Weston",
    "Brett",
    "Leslie",
    "Carmen",
    "Garrett",
    "Company Acct",
]

NON_SALARIED = {"Brett", "Leslie", "Carmen", "Garrett"}


def clean_b2b_summary(ws: Worksheet) -> None:
    """Clear hardcoded values and replace placeholders in B2B Summary."""
    if ws.cell(18, 9).value:
        ws.cell(18, 9, "[MONTH] [YEAR] ORDERS - SHIPPED AND INVOICED")
    # F38, G38: hardcoded B2B Executive Account values
    ws.cell(38, 6, None)
    ws.cell(38, 7, None)
    # P34, R34: hardcoded zero values
    ws.cell(34, 16, None)
    ws.cell(34, 18, None)


def copy_sheet_to_workbook(source_ws: Worksheet, target_wb, new_title: str) -> Worksheet:
    """Copy a worksheet (with styles, dimensions, formulas) into target_wb."""
    target_ws = target_wb.create_sheet(new_title)

    # Copy column dimensions (widths, hidden state)
    for col_letter, dim in source_ws.column_dimensions.items():
        target_dim = target_ws.column_dimensions[col_letter]
        target_dim.width = dim.width
        target_dim.hidden = dim.hidden
        target_dim.outlineLevel = dim.outlineLevel

    # Copy row dimensions
    for row_num, dim in source_ws.row_dimensions.items():
        target_dim = target_ws.row_dimensions[row_num]
        target_dim.height = dim.height
        target_dim.hidden = dim.hidden
        target_dim.outlineLevel = dim.outlineLevel

    # Copy merged cells
    for mr in list(source_ws.merged_cells.ranges):
        target_ws.merge_cells(str(mr))

    # Copy each cell with value + style
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    # Copy sheet print/view settings (sheet_view is read-only; copy attributes instead)
    try:
        target_ws.sheet_properties.tabColor = source_ws.sheet_properties.tabColor
    except Exception:
        pass
    try:
        target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
        target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth
    except Exception:
        pass

    return target_ws


def configure_salesperson_sheet(ws: Worksheet, name: str) -> None:
    """Set salesperson name and rate type label on a copied sheet."""
    full_name = name.upper()
    ws.cell(2, 4, full_name)
    ws.cell(3, 4, "COMMISSION REPORT - [MONTH] [YEAR]")
    rate_label = "Non-Salary Commission Rate" if name in NON_SALARIED else "Salary Commission Rate"
    ws.cell(4, 4, rate_label)


def main() -> None:
    print(f"Loading clean Brett: {BRETT_CLEAN.name}")
    if not BRETT_CLEAN.exists():
        raise FileNotFoundError(f"Run make_clean_salesperson_template.py first to create {BRETT_CLEAN}")
    wb_brett = load_workbook(BRETT_CLEAN)

    print(f"Loading original master to extract B2B Summary: {ORIGINAL_MASTER.name}")
    wb_master_orig = load_workbook(ORIGINAL_MASTER)

    # Use the Brett workbook as base — it already has Brett + Table + R_LP + R_SO + R_INV + R_SH
    # We need to:
    #   1. Add B2B Summary (cleaned, from original master)
    #   2. Add 9 more salesperson sheets (copies of Brett, renamed)

    # Step 1: Copy B2B Summary from original master into wb_brett
    print("Copying B2B Summary from original (cleaned)...")
    src_summary = wb_master_orig["B2B Summary"]
    clean_b2b_summary(src_summary)
    new_summary = copy_sheet_to_workbook(src_summary, wb_brett, "B2B Summary")
    # Move B2B Summary to first position
    wb_brett.move_sheet(new_summary, offset=-(len(wb_brett.sheetnames) - 1))

    # Step 2: For each salesperson != Brett, create a copy of Brett sheet
    brett_sheet = wb_brett["Brett"]
    configure_salesperson_sheet(brett_sheet, "Brett")

    for name in SALESPERSON_SHEETS:
        if name == "Brett":
            continue
        print(f"  Creating sheet '{name}' as copy of Brett...")
        # openpyxl's copy_worksheet copies within the same workbook with styles
        new_sheet = wb_brett.copy_worksheet(brett_sheet)
        new_sheet.title = name
        configure_salesperson_sheet(new_sheet, name)

    # Step 3: Close source master
    wb_master_orig.close()

    # Step 4: Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb_brett.save(OUTPUT)
    wb_brett.close()
    print(f"\nOK: master template saved to {OUTPUT}")
    print(f"  Size: {OUTPUT.stat().st_size:,} bytes")
    # Inspect resulting sheets
    wb_check = load_workbook(OUTPUT)
    print(f"  Sheets: {wb_check.sheetnames}")
    wb_check.close()


if __name__ == "__main__":
    main()
