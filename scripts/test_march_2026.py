"""
Validation test: run the existing Manual-Adjustments workflow on March 2026 and
compare against Jennifer's completed historical workbook.

Read-only on SQLite + historical files. Writes only the generated workbook to
data/output. Does NOT call Zoho, modify historical workbooks, or change logic.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from src.commission.sqlite_to_workbook import generate_commission_workbook
from src.db.adjustments import list_adjustments

YEAR, MONTH = 2026, 3
TPL = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"
OUT = BASE_DIR / "data" / "output" / "commission_b2b_march_2026.xlsx"
HIST = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B.xlsx"
)
INPUT_SHIP = BASE_DIR / "data" / "input" / "Shipments March 2026.xlsx"

NAME_TO_SHEET = {
    "Paul Perlman": "Paul", "Jose Ayala": "Jose", "Michael Ayala": "Michael",
    "Jim Sutton": "Jim", "Weston Fields": "Weston", "Brett Bern": "Brett",
    "Leslie Neipert": "Leslie", "Carmen Daetz": "Carmen", "Garrett Lockhart": "Garrett",
    "B2B Company Account": "Company Acct", "Company Account": "Company Acct",
}


def hist_totals(path: Path) -> dict[str, float]:
    wb = load_workbook(path, data_only=True)
    ws = wb["B2B Summary"]
    out: dict[str, float] = {}
    for r in range(20, 45):
        name = ws.cell(r, 5).value or ws.cell(r, 4).value
        if not name:
            continue
        name = str(name).strip()
        if name.lower() in ("salary commission", "non-salary commission", "b2b sales rep",
                            "total b2b", "b2b executive account"):
            continue
        cur = ws.cell(r, 13).value or 0
        pri = ws.cell(r, 19).value or 0
        try:
            tot = float(cur) + float(pri)
        except (TypeError, ValueError):
            continue
        if abs(tot) > 0.005:
            key = NAME_TO_SHEET.get(name, name)
            out[key] = round(out.get(key, 0) + tot, 2)
    wb.close()
    return out


def main() -> None:
    print("=" * 70)
    print(f"  MARCH {YEAR} WORKFLOW VALIDATION (generated vs Jennifer historical)")
    print("=" * 70)

    pre_adj = list_adjustments(YEAR, MONTH)
    print(f"\nStored adjustments for March (should be 0 for a clean test): {len(pre_adj)}")

    res = generate_commission_workbook(YEAR, MONTH, template_path=TPL, output_path=OUT)
    k = res.kpis
    gen = {s: round(v, 2) for s, v in res.totals_by_sheet.items() if abs(v) > 0.005}
    gen_total = round(sum(res.totals_by_sheet.values()), 2)

    hist = hist_totals(HIST)
    hist_total = round(sum(hist.values()), 2)

    # ----- per-salesperson comparison -----
    print("\n--- Per-salesperson commission (generated vs historical) ---")
    print(f"  {'Salesperson':<16}{'Generated':>12}{'Historical':>12}{'Diff':>12}")
    diffs = []
    for key in sorted(set(gen) | set(hist)):
        g = gen.get(key, 0.0)
        h = hist.get(key, 0.0)
        d = round(g - h, 2)
        diffs.append((key, g, h, d))
        flag = "" if abs(d) < 0.01 else "  <-- diff"
        print(f"  {key:<16}{g:>12,.2f}{h:>12,.2f}{d:>12,.2f}{flag}")
    print(f"  {'TOTAL':<16}{gen_total:>12,.2f}{hist_total:>12,.2f}{gen_total-hist_total:>12,.2f}")

    # ----- required report metrics (item 9) -----
    commissionable = [a for a in res.audit_rows if a["block"] == "commissionable" and not a["pending"]]
    missing_ship = 0 if k.get("shipment_data_present") else len(
        [a for a in res.audit_rows if a["block"] == "commissionable"]
    )

    print("\n--- COMPARISON REPORT ---")
    print(f"  Generated total commission (rep+company): ${gen_total:,.2f}")
    print(f"  Historical Jennifer total:                ${hist_total:,.2f}")
    print(f"  Difference:                               ${gen_total-hist_total:,.2f}  ({(gen_total-hist_total)/hist_total*100:+.1f}%)")
    print(f"  Pending lines (need assignment):          {k['pending_lines']}")
    print(f"  Missing shipment fields (lines):          {missing_ship}  (SQLite shipments synced: {k.get('shipment_data_present')})")
    print(f"  Exceptions flagged:                       {k['exceptions_count']}")
    print(f"  Draft status:                             {k['is_draft']}")

    print("\n--- Top reasons for differences (largest abs per-rep gaps) ---")
    for key, g, h, d in sorted(diffs, key=lambda x: -abs(x[3]))[:6]:
        if abs(d) < 0.01:
            continue
        if h == 0 and g > 0:
            reason = "we include lines Jennifer excluded (curation)"
        elif g == 0 and h > 0:
            reason = "Jennifer credited a rep we routed elsewhere (reassignment)"
        elif d > 0:
            reason = "higher: extra orders or lower computed discount vs hand-entered"
        else:
            reason = "lower: Jennifer manually added/kept lines we excluded"
        print(f"  {key:<16} diff ${d:>10,.2f}  — {reason}")

    # ----- item 7/8: local shipment export availability -----
    print("\n--- Shipment data sources (item 7/8) ---")
    print(f"  SQLite shipments for March: {'present' if k.get('shipment_data_present') else 'EMPTY (0 rows)'}")
    if INPUT_SHIP.exists():
        try:
            sw = load_workbook(INPUT_SHIP, read_only=True)
            ws0 = sw[sw.sheetnames[0]]
            rc = ws0.max_row
            sw.close()
        except Exception:
            rc = "?"
        print(f"  Local export EXISTS: {INPUT_SHIP.name}  (~{rc} rows)")
        print("  -> a supplemental local-shipment source could fill Shipment Date/Status/Carrier for validation.")
    else:
        print(f"  Local export not found: {INPUT_SHIP.name}")

    print(f"\nGenerated workbook: {OUT}")
    print("=" * 70)


if __name__ == "__main__":
    main()
