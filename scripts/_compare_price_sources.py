"""Read-only: compare March/May R_LP vs April DB snapshot."""
from __future__ import annotations

import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

from dotenv import load_dotenv

load_dotenv(REPO / ".env")

import openpyxl  # noqa: E402

from src.db.connection import get_connection  # noqa: E402


def parse_rlp(path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["R_LP"]
    rows: dict[str, float] = {}
    for r in range(1, ws.max_row + 1):
        sku_raw = ws.cell(r, 1).value
        price_raw = ws.cell(r, 7).value
        if not sku_raw:
            continue
        sku = str(sku_raw).strip().upper()
        if not sku or sku in ("SKU", "ITEM"):
            continue
        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        if sku not in rows:
            rows[sku] = price
    wb.close()
    return rows


def main() -> None:
    downloads = Path.home() / "Downloads"
    march = parse_rlp(downloads / "2026-3_Commission B2B.xlsx")
    may = parse_rlp(downloads / "commission_b2b_may_2026.xlsx")
    march_alt = parse_rlp(downloads / "commission_b2b_march_2026.xlsx")

    print("March primary SKUs:", len(march))
    print("March alt SKUs    :", len(march_alt))
    print("May SKUs          :", len(may))
    common_mm = set(march) & set(may)
    diff_mm = [s for s in common_mm if march[s] != may[s]]
    print("March vs May common:", len(common_mm), "price diffs:", len(diff_mm))
    if diff_mm[:10]:
        print("  sample diffs:", [(s, march[s], may[s]) for s in diff_mm[:10]])

    conn = get_connection()
    apr_rows = conn.execute(
        "SELECT sku, map_price FROM price_history "
        "WHERE snapshot_month='2026-04' AND source='accountant_fvprice_2026_04'"
    ).fetchall()
    conn.close()
    apr = {str(r["sku"]).upper(): float(r["map_price"]) for r in apr_rows}
    print("April DB (FV_PRICE) SKUs:", len(apr))

    overlap = set(march) & set(apr)
    diffs = [s for s in overlap if march[s] != apr[s]]
    print("March R_LP vs April DB overlap:", len(overlap))
    print("Price diffs on overlap:", len(diffs))
    if diffs:
        print("  first 20 diffs:", [(s, march[s], apr[s]) for s in sorted(diffs)[:20]])

    only_apr = set(apr) - set(march)
    only_march = set(march) - set(apr)
    print("Only in April FV_PRICE (not March R_LP):", len(only_apr))
    if only_apr:
        print("  sample:", sorted(only_apr)[:15])
    print("Only in March R_LP (not April FV_PRICE):", len(only_march))

    print("PWS015:", "march=", march.get("PWS015"), "may=", may.get("PWS015"), "apr=", apr.get("PWS015"))


if __name__ == "__main__":
    main()
