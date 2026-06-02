"""
Build a clean salesperson template from the original Brett March 2026 workbook.

Strategy: keep ALL styling, column widths, hidden columns, fonts, borders, fills.
Only clear the actual data values, preserving formulas that compute per-row
(AE, AF, AG, AJ, AK, AL) and subtotal formulas with their anchor positions.

Output: data/templates/salesperson_template_clean.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

SOURCE = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B_BRETT BERN.xlsx"
)
OUTPUT = BASE_DIR / "data" / "templates" / "salesperson_template_clean.xlsx"


# Columns that hold data values (cleared per row)
DATA_VALUE_COLS = [
    4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17,  # D..Q
    19, 20, 21, 22,        # S, T, U, V
    24, 25, 26,             # X, Y, Z
    27, 28,                 # AA, AB (shipping)
    31,                     # AE (Map Price) — sometimes hardcoded, sometimes VLOOKUP
    34,                     # AH (Discount Rate) — manual value
]

# Columns that have per-row formulas we want to KEEP
FORMULA_COLS_TO_KEEP = {29, 31, 32, 33, 36, 37, 38}  # AC, AE, AF, AG, AJ, AK, AL


def clear_data_in_block(ws, header_row: int, subtotal_row: int) -> None:
    """Clear data values in rows between header and subtotal, preserve formulas."""
    for r in range(header_row + 1, subtotal_row):
        for col in DATA_VALUE_COLS:
            cell = ws.cell(r, col)
            v = cell.value
            # If it's a formula, keep it
            if isinstance(v, str) and v.startswith("="):
                continue
            cell.value = None


def main() -> None:
    print(f"Loading source: {SOURCE}")
    wb = load_workbook(SOURCE)

    ws = wb["Brett"]

    # Clear title placeholders
    ws.cell(2, 4, "[SALESPERSON NAME]")
    ws.cell(3, 4, "COMMISSION REPORT - [MONTH] [YEAR]")

    # Clear section label hardcoded month
    ws.cell(20, 3, "SALES ORDERS INVOICED IN [MONTH] [YEAR]")

    # Block I.1: header R23, subtotal R45 -> clear rows 24..44
    print("Clearing Block I.1 (R24-R44)")
    clear_data_in_block(ws, 23, 45)

    # Block I.2 Shipping Income: header R48, subtotal R53 -> clear rows 49..52
    print("Clearing Block I.2 (R49-R52)")
    clear_data_in_block(ws, 48, 53)

    # Block I.3 Other Charges: header R56, subtotal R59 -> clear rows 57..58
    print("Clearing Block I.3 (R57-R58)")
    clear_data_in_block(ws, 56, 59)

    # Block II.1 Prior Orders: header R64, subtotal R66 -> clear row 65
    print("Clearing Block II.1 (R65)")
    clear_data_in_block(ws, 64, 66)

    # Block II.2 Prior Shipping: header R69, subtotal R71 -> clear row 70
    print("Clearing Block II.2 (R70)")
    clear_data_in_block(ws, 69, 71)

    # Clear manual N11 (Credits/Advances)
    ws.cell(11, 14, 0)

    # Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    wb.close()
    print(f"OK: clean template saved to {OUTPUT}")
    print(f"  Size: {OUTPUT.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
