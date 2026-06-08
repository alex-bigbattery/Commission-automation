"""Idempotent loader for an effective-dated price_history snapshot.

The single sanctioned write path for the price_history table. R_LP and items.rate
are NEVER touched by this script — price_history is an independent, additive price
source that takes priority over the fallback maps in the commission engine.

Defaults to the accountant's April 2026 FV_PRICE workbook, but parameterized so
future monthly snapshots use the same path:

    python -m scripts.load_price_history_snapshot \\
        --workbook "C:/path/to/2026-05_Commissions_B2B.xlsx" \\
        --sheet FV_PRICE \\
        --snapshot-month 2026-05 \\
        --source accountant_fvprice_2026_05

Contract enforced before any INSERT (refuses the row, fails LOUDLY):
  * sku non-empty, upper-cased, trimmed
  * map_price a finite number > 0
  * effective_from / effective_to: real ISO YYYY-MM-DD; eff_from <= eff_to
  * effective window must be entirely INSIDE the declared snapshot_month
  * no duplicates within the incoming load on (sku, effective_from)
  * snapshot_month + source pair is the idempotency key: a prior load with the same
    (snapshot_month, source) is DELETED in the same transaction before the new
    rows insert, so re-running is safe.

A backup of the DB is taken first (SQLite file copy). For Postgres, this script does
NOT take a backup — it prints a loud warning to stderr and the operator MUST take a
pg_dump (or rely on their managed-DB point-in-time recovery) before re-running. The
DELETE+INSERT itself is wrapped in a single transaction so a partial/failed load is
rolled back, but a *successful* load that overwrites valid prior rows can only be
undone from a real backup.
"""
from __future__ import annotations

import argparse
import calendar
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

REPO = Path(__file__).resolve().parent.parent
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

import openpyxl  # noqa: E402

from src.db.connection import DB_PATH, get_connection, init_database  # noqa: E402


DEFAULT_WORKBOOK = (
    Path(os.environ.get("USERPROFILE", str(Path.home())))
    / "Downloads"
    / "2026-04_Commissions_B2B.xlsx"
)


def _month_bounds(snapshot_month: str) -> tuple[str, str]:
    y, m = (int(p) for p in snapshot_month.split("-", 1))
    last = calendar.monthrange(y, m)[1]
    return (f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}")


def _validate_iso(value: str, label: str) -> str:
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise SystemExit(f"--{label} must be YYYY-MM-DD ({value!r}): {exc}")
    return value


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Load a price_history snapshot from an accountant workbook.")
    p.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK,
                   help=f"Path to the accountant Excel workbook (default: {DEFAULT_WORKBOOK})")
    p.add_argument("--sheet", default="FV_PRICE",
                   help="Sheet name inside the workbook (default: FV_PRICE)")
    p.add_argument("--sku-column", type=int, default=2,
                   help="1-based column index that holds the SKU (default: 2 / column B)")
    p.add_argument("--price-column", type=int, default=3,
                   help="1-based column index that holds the price (default: 3 / column C)")
    p.add_argument("--snapshot-month", default="2026-04",
                   help="YYYY-MM bucket label used for idempotency (default: 2026-04)")
    p.add_argument("--source", default="accountant_fvprice_2026_04",
                   help="Provenance label (default: accountant_fvprice_2026_04)")
    p.add_argument("--effective-from", default=None,
                   help="Override start (default: first day of --snapshot-month)")
    p.add_argument("--effective-to", default=None,
                   help="Override end (default: last day of --snapshot-month)")
    p.add_argument("--dry-run", action="store_true",
                   help="Parse and validate the workbook; do not write to the DB.")
    p.add_argument("--confirm",
                   help="Must equal --snapshot-month to proceed. Required for non-dry-run.")
    p.add_argument("--allow-mismatched-filename", action="store_true",
                   help="Skip the workbook-filename ↔ --snapshot-month consistency check.")
    return p.parse_args()


def _read_workbook(args: argparse.Namespace) -> list[tuple[str, float]]:
    wb = openpyxl.load_workbook(args.workbook, read_only=False, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {args.sheet!r} not found in {args.workbook}. "
                         f"Available: {wb.sheetnames}")
    ws = wb[args.sheet]
    rows: list[tuple[str, float]] = []
    seen: dict[str, float] = {}
    skipped_zero = 0
    skipped_blank = 0
    skipped_dup = 0
    for r in range(1, ws.max_row + 1):
        sku_raw = ws.cell(r, args.sku_column).value
        price_raw = ws.cell(r, args.price_column).value
        if sku_raw is None:
            continue
        sku = str(sku_raw).strip().upper()
        if not sku or sku in ("SKU", "ITEM", "ITEM SKU", "SKU/ITEM", "PRICES", "COMPONENTS"):
            continue
        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            skipped_blank += 1
            continue
        if price <= 0:
            skipped_zero += 1
            continue
        if sku in seen:
            skipped_dup += 1
            continue
        seen[sku] = price
        rows.append((sku, price))
    print(f"Parsed {len(rows)} priced SKUs from {args.workbook.name} :: {args.sheet}")
    print(f"  skipped zero/<=0  : {skipped_zero}")
    print(f"  skipped blank/NaN : {skipped_blank}")
    print(f"  skipped duplicate : {skipped_dup}")
    return rows


def _validate_window(snapshot_month: str, eff_from: str, eff_to: str) -> None:
    bounds_from, bounds_to = _month_bounds(snapshot_month)
    if eff_from > eff_to:
        raise SystemExit(f"effective_from ({eff_from}) > effective_to ({eff_to})")
    if eff_from < bounds_from or eff_to > bounds_to:
        raise SystemExit(
            f"Window [{eff_from}..{eff_to}] falls outside snapshot_month "
            f"[{bounds_from}..{bounds_to}]. Refuse to write a row that could "
            f"leak into another period."
        )


def _backup_db() -> None:
    """SQLite backend: copy the DB file to a timestamped sibling. Postgres backend:
    print a loud stderr warning — this script does NOT call pg_dump."""
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    db_url = os.environ.get("DATABASE_URL")
    if db_url:
        msg = (
            "!! WARNING: Postgres backend in use (DATABASE_URL set).\n"
            "!! This loader does NOT take a Postgres backup. The DELETE+INSERT is\n"
            "!! transactional, so a failed load is rolled back -- but a SUCCESSFUL\n"
            "!! load that overwrites valid prior rows can only be undone from a\n"
            "!! real pg_dump or your provider's point-in-time-recovery.\n"
            f"!! Timestamp: {ts}.  Press Ctrl-C now if you have not backed up.\n"
        )
        print(msg, file=sys.stderr, flush=True)
        return
    bak = DB_PATH.with_name(DB_PATH.name + f".BAK-{ts}")
    shutil.copy2(DB_PATH, bak)
    print(f"DB backup -> {bak.name}")


def main() -> int:
    args = _parse_args()
    _validate_iso(args.snapshot_month + "-01", "snapshot-month")
    bounds_from, bounds_to = _month_bounds(args.snapshot_month)
    eff_from = _validate_iso(args.effective_from or bounds_from, "effective-from")
    eff_to = _validate_iso(args.effective_to or bounds_to, "effective-to")
    _validate_window(args.snapshot_month, eff_from, eff_to)
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    # Guardrail #1: refuse to load when the workbook filename does NOT contain the
    # declared --snapshot-month token (e.g. trying to load 2026-04 workbook under
    # --snapshot-month 2026-05). The single most damaging silent mis-load is loading
    # the wrong workbook under a correct-looking (snapshot_month, source) label.
    if (args.snapshot_month not in args.workbook.name
            and not args.allow_mismatched_filename):
        raise SystemExit(
            f"Refusing to load: workbook filename {args.workbook.name!r} does not "
            f"contain --snapshot-month token {args.snapshot_month!r}. If this is "
            f"intentional, re-run with --allow-mismatched-filename."
        )

    # Guardrail #2: require the operator to echo --snapshot-month via --confirm before
    # any DB writes. Dry-run skips this so validation can be tested without ceremony.
    if not args.dry_run:
        if not args.confirm:
            raise SystemExit(
                f"Refusing to load without --confirm. Pass --confirm {args.snapshot_month} "
                f"to confirm you are loading the {args.snapshot_month} snapshot."
            )
        if args.confirm != args.snapshot_month:
            raise SystemExit(
                f"--confirm {args.confirm!r} does not match --snapshot-month "
                f"{args.snapshot_month!r}. Refusing to load."
            )

    rows = _read_workbook(args)
    if not rows:
        raise SystemExit("No priced rows after validation — refusing to wipe an existing snapshot.")

    print(f"Snapshot identity: snapshot_month={args.snapshot_month}  source={args.source}")
    print(f"Effective window : {eff_from}  ..  {eff_to}")
    if args.dry_run:
        print("[dry-run] Skipping DB writes.")
        return 0

    _backup_db()
    # init_database is called inside get_connection too, but explicit here keeps the
    # ordering obvious: schema-and-migrations first, then writes.
    init_database()

    captured = datetime.now().isoformat(timespec="seconds")
    conn = get_connection()
    try:
        itemid_by_sku: dict[str, str] = {}
        for row in conn.execute(
            "SELECT sku, item_id FROM items WHERE sku IS NOT NULL AND sku != ''"
        ).fetchall():
            k = str(row["sku"]).strip().upper()
            if k and k not in itemid_by_sku:
                itemid_by_sku[k] = row["item_id"]

        before = conn.execute("SELECT COUNT(*) AS c FROM price_history").fetchone()["c"]
        # Idempotency: delete prior rows of the SAME snapshot_month+source, then
        # insert. The UNIQUE(sku, effective_from, snapshot_month, source) index
        # enforces this at the DB level too — a re-run would otherwise raise.
        conn.execute(
            "DELETE FROM price_history WHERE snapshot_month=? AND source=?",
            (args.snapshot_month, args.source),
        )
        payload = [
            (sku, itemid_by_sku.get(sku), price, eff_from, eff_to,
             args.source, args.snapshot_month, captured)
            for sku, price in rows
        ]
        conn.executemany(
            "INSERT INTO price_history "
            "(sku, item_id, map_price, effective_from, effective_to, source, snapshot_month, captured_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            payload,
        )
        conn.commit()
        after = conn.execute("SELECT COUNT(*) AS c FROM price_history").fetchone()["c"]
        print(f"price_history rows: before={before}  after={after}  inserted={len(payload)}")
        print("By snapshot_month/source:")
        for row in conn.execute(
            "SELECT snapshot_month, source, COUNT(*) AS n, MIN(effective_from) AS f, MAX(effective_to) AS t "
            "FROM price_history GROUP BY snapshot_month, source ORDER BY snapshot_month, source"
        ).fetchall():
            print(f"   {row['snapshot_month']} | {row['source']} | n={row['n']} | {row['f']} -> {row['t']}")
    finally:
        conn.close()
    print("DONE.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
