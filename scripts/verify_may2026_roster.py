"""Verify roster assignment for May 2026 (SO-04119 / INV-05953). No Zoho calls."""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

from openpyxl import load_workbook

from src.commission.sqlite_to_workbook import (
    build_salespeople_from_sqlite,
    load_map_from_template,
    load_tiers_from_template,
)

YEAR, MONTH = 2026, 5
SO = "SO-04119"
INV = "INV-05953"
TEMPLATE = ROOT / "data" / "templates" / "master_template_clean.xlsx"
OUTPUT = ROOT / "data" / "output" / "commission_b2b_may_2026.xlsx"


def cf_sales_team(raw: dict) -> str | None:
    for item in raw.get("custom_fields") or []:
        if not isinstance(item, dict):
            continue
        label = (item.get("label") or item.get("api_name") or "").lower()
        if "sales team" in label or "sales_team" in label:
            return item.get("value") or item.get("value_formatted")
    cf = raw.get("custom_field_hash") or {}
    if isinstance(cf, dict):
        for k, v in cf.items():
            if "sales" in str(k).lower() and "team" in str(k).lower():
                return v
    return None


def print_so_row():
    from src.db.connection import get_connection, init_database

    init_database()
    conn = get_connection()
    row = conn.execute(
        """
        SELECT salesorder_number, salesperson_name, customer_name, raw_json
        FROM sales_orders WHERE salesorder_number = ?
        """,
        (SO,),
    ).fetchone()
    conn.close()
    if not row:
        print(f"ERROR: no sales_orders row for {SO}")
        return
    raw = json.loads(row["raw_json"]) if row["raw_json"] else {}
    print("=== 1) sales_orders (SO-04119) ===")
    print("salesorder_number:", row["salesorder_number"])
    print("salesperson_name:", row["salesperson_name"])
    print("salesperson_id:", raw.get("salesperson_id") or raw.get("salesperson"))
    print("cf_sales_team / CF.Sales Team:", cf_sales_team(raw))
    print("customer_name:", row["customer_name"])
    print("raw_json.salesperson_name:", raw.get("salesperson_name"))
    print("raw_json.salesperson:", raw.get("salesperson"))
    print()


def print_api_row():
    tiers = load_tiers_from_template(TEMPLATE)
    rlp = load_map_from_template(TEMPLATE)
    result = build_salespeople_from_sqlite(YEAR, MONTH, tiers=tiers, rlp_map=rlp, apply_adjustments=True)
    matches = [
        r
        for r in result.audit_rows
        if SO in str(r.get("sales_order") or "")
        or INV in str(r.get("invoice") or "")
    ]
    print(f"=== 3) Engine audit rows for {SO} / {INV} ({len(matches)} match) ===")
    keys = [
        "line_uid",
        "sales_order",
        "invoice",
        "sku",
        "original_zoho_salesperson",
        "final_commission_assignment",
        "accounting_category",
        "issue_found",
        "suggested_action",
        "system_salesperson",
        "salesperson",
        "pending",
        "flags",
    ]
    for r in matches:
        print(json.dumps({k: r.get(k) for k in keys}, indent=2, default=str))
        print()
    if not matches:
        print("WARNING: no audit row found for this SO/invoice")
    return matches


def audit_sheet_headers(path: Path) -> list[str]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Adjustments Audit" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Adjustments Audit"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    wb.close()
    return [str(h) for h in headers if h]


def regenerate():
    from src.commission.sqlite_to_workbook import generate_commission_workbook

    print("=== Regenerating May 2026 workbook ===")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    generate_commission_workbook(YEAR, MONTH, template_path=TEMPLATE, output_path=OUTPUT)
    print("Wrote:", OUTPUT)
    print()


def main():
    print_so_row()
    matches = print_api_row()
    if "--regenerate" in sys.argv:
        regenerate()
    headers = audit_sheet_headers(OUTPUT)
    print("=== Adjustments Audit headers ===")
    if not headers:
        print(f"No workbook or sheet at {OUTPUT} (pass --regenerate to build)")
    else:
        for h in headers:
            print(" -", h)
        required = [
            "Original Zoho Salesperson",
            "Final Commission Assignment",
            "Accounting Category",
            "Issue Found",
            "Suggested Action",
        ]
        missing = [h for h in required if h not in headers]
        if missing:
            print("MISSING columns:", missing)
        else:
            print("OK: all required columns present")
    if matches and OUTPUT.exists():
        wb = load_workbook(OUTPUT, read_only=True, data_only=True)
        ws = wb["Adjustments Audit"]
        hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col_so = hdr.get("Sales Order")
        col_inv = hdr.get("Invoice")
        for r in range(2, ws.max_row + 1):
            so_val = ws.cell(r, col_so).value if col_so else None
            inv_val = ws.cell(r, col_inv).value if col_inv else None
            if SO in str(so_val or "") or INV in str(inv_val or ""):
                print("=== Audit sheet row ===")
                for name, c in hdr.items():
                    if name in required + ("Sales Order", "Invoice", "Pending"):
                        print(f"  {name}: {ws.cell(r, c).value}")
        wb.close()


if __name__ == "__main__":
    main()
