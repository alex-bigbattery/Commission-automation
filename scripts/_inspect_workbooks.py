"""Inspect price tabs in candidate workbooks."""
import sys
from pathlib import Path
import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

files = [
    Path.home() / "Downloads" / "2026-3_Commission B2B.xlsx",
    Path.home() / "Downloads" / "2026-4_Commission B2B.xlsx",
    Path.home() / "Downloads" / "commission_b2b_march_2026.xlsx",
    Path.home() / "Downloads" / "commission_b2b_april_2026.xlsx",
    Path.home() / "Downloads" / "commission_b2b_may_2026.xlsx",
    Path.home() / "Downloads" / "2026-04_Commissions_B2B.xlsx",
]

def count_rlp(ws):
    n = 0
    sample = []
    for r in range(3, min(ws.max_row + 1, 5000)):
        sku = ws.cell(r, 1).value
        rate = ws.cell(r, 7).value
        if not sku:
            continue
        sku = str(sku).strip().upper()
        if not sku:
            continue
        try:
            p = float(rate)
        except (TypeError, ValueError):
            continue
        if p <= 0:
            continue
        n += 1
        if len(sample) < 5:
            sample.append((sku, p))
    return n, sample

for f in files:
    if not f.exists():
        print("MISSING", f.name)
        continue
    wb = openpyxl.load_workbook(f, data_only=True)
    print("===", f.name, "===")
    print("sheets:", ", ".join(wb.sheetnames[:20]))
    if "FV_PRICE" in wb.sheetnames:
        ws = wb["FV_PRICE"]
        print("  FV_PRICE max_row", ws.max_row)
        for r in range(1, 6):
            print("   ", [ws.cell(r, c).value for c in range(1, 5)])
    if "R_LP" in wb.sheetnames:
        n, sample = count_rlp(wb["R_LP"])
        print(f"  R_LP priced SKUs: {n}")
        print(f"  R_LP sample: {sample}")
        pws = [x for x in sample if x[0] == "PWS015"]
        ws = wb["R_LP"]
        for r in range(3, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip().upper() == "PWS015":
                print(f"  PWS015 row{r}: rate={ws.cell(r, 7).value}")
                break
    wb.close()
