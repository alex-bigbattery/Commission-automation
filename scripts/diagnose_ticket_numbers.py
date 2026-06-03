"""
Ticket-number discovery (READ-ONLY diagnostic).

Accounting rule (June 2026 governance form):
    "If Ticket numbers exist on an order then it is usually noncommissionable
     even if it received payment."

We do NOT know yet HOW a ticket number actually appears in our Zoho data, so
this script inspects the existing SQLite/Supabase data (never calls Zoho, never
modifies anything) to find every place a ticket indicator shows up:
    - structured columns: reference_number, sku, item_name, items.description
    - raw_json custom_fields  (label + value)   <-- most likely clean source
    - raw_json notes / terms
    - raw_json line_items[].description / .name

It also enumerates EVERY distinct custom-field label across sales orders and
invoices — that single list usually answers "is there a dedicated Ticket Number
field?" outright.

Output:
    - a console summary
    - data/output/ticket_number_diagnostic.xlsx (Summary + per-source sheets)

Run (only when NO Zoho sync is in progress):
    python scripts/diagnose_ticket_numbers.py                # scan everything
    python scripts/diagnose_ticket_numbers.py --year 2026 --month 5
    python scripts/diagnose_ticket_numbers.py --year 2026 --month 3

Nothing here changes commission outputs. The engine's TICKET_NUMBER flag stays
review-only until this diagnostic confirms the real detection rule.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

# --- make `src` importable when run as `python scripts/diagnose_ticket_numbers.py`
REPO = Path(__file__).resolve().parent.parent
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from src.db.connection import get_connection, database_label  # noqa: E402

OUT_XLSX = REPO / "data" / "output" / "ticket_number_diagnostic.xlsx"

# Two confidence tiers. STRONG terms almost certainly mean a real ticket/case;
# WEAK terms appear in normal product names/descriptions and MUST NOT drive any
# automatic exclusion — they are reported separately as false-positive risks.
STRONG_TERMS = ["ticket", "tkt", "rma", "case #", "case number", "case no"]
WEAK_TERMS = ["case", "service", "support", "warranty", "repair", "replacement", "return"]

# A tighter pattern for "Ticket #1234 / Ticket No 1234 / TKT-1234".
TICKET_NUM_RE = re.compile(r"(ticket|tkt|case|rma)[\s#:.\-]*\d{2,}", re.IGNORECASE)


def parse_json(raw) -> dict:
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}


def find_terms(text: str, terms: list[str]) -> list[str]:
    if not text:
        return []
    low = text.lower()
    return [t for t in terms if t in low]


def snippet(text: str, term: str, width: int = 60) -> str:
    """Return a short context window around the first occurrence of `term`."""
    if not text:
        return ""
    low = text.lower()
    idx = low.find(term.lower())
    if idx < 0:
        # fall back to the regex hit
        m = TICKET_NUM_RE.search(text)
        idx = m.start() if m else 0
    start = max(0, idx - width)
    end = min(len(text), idx + len(term) + width)
    return ("…" if start > 0 else "") + text[start:end].replace("\n", " ").strip() + ("…" if end < len(text) else "")


def classify(text: str) -> tuple[str, list[str], list[str]]:
    """Return (confidence, strong_hits, weak_hits) for a blob of text."""
    strong = find_terms(text, STRONG_TERMS)
    weak = [w for w in find_terms(text, WEAK_TERMS) if w not in strong]
    if TICKET_NUM_RE.search(text or ""):
        confidence = "HIGH (ticket+number pattern)"
    elif strong:
        confidence = "MEDIUM (strong keyword)"
    elif weak:
        confidence = "LOW (weak keyword — false-positive risk)"
    else:
        confidence = ""
    return confidence, strong, weak


# ---------------------------------------------------------------------------
# Scanners — each returns a list of dict rows for the Excel report.
# ---------------------------------------------------------------------------

def scan_custom_field_labels(conn) -> tuple[Counter, list[dict]]:
    """Enumerate every distinct custom-field label across SOs + invoices, and
    return any whose label/value mentions a ticket indicator."""
    labels = Counter()
    label_hits: list[dict] = []
    for table, idcol, numcol in (
        ("sales_orders", "salesorder_id", "salesorder_number"),
        ("invoices", "invoice_id", "invoice_number"),
    ):
        rows = conn.execute(f"SELECT {idcol}, {numcol}, raw_json FROM {table}").fetchall()
        for r in rows:
            obj = parse_json(r["raw_json"])
            for f in obj.get("custom_fields") or []:
                label = str(f.get("label") or f.get("api_name") or "").strip()
                value = str(f.get("value") or "").strip()
                if not label:
                    continue
                labels[label] += 1
                blob = f"{label} = {value}"
                conf, strong, weak = classify(blob)
                if conf:
                    label_hits.append({
                        "source": table,
                        "doc_number": r[numcol] or "",
                        "custom_field_label": label,
                        "value": value,
                        "confidence": conf,
                        "matched_terms": ", ".join(strong + weak),
                    })
    return labels, label_hits


def scan_documents(conn, table: str, numcol: str, fields: list[str], year, month) -> list[dict]:
    """Scan structured + raw_json (notes/reference/terms) of SOs or invoices."""
    cols = ", ".join(["raw_json", numcol] + fields)
    rows = conn.execute(f"SELECT {cols} FROM {table}").fetchall()
    hits: list[dict] = []
    for r in rows:
        obj = parse_json(r["raw_json"])
        # structured columns + common raw_json free-text fields
        sources = {f: (r[f] or "") for f in fields}
        sources["raw.notes"] = str(obj.get("notes") or "")
        sources["raw.terms"] = str(obj.get("terms") or "")
        sources["raw.reference_number"] = str(obj.get("reference_number") or "")
        sources["raw.subject"] = str(obj.get("subject") or "")
        for src_name, text in sources.items():
            conf, strong, weak = classify(text)
            if conf:
                hits.append({
                    "source_field": f"{table}.{src_name}",
                    "doc_number": r[numcol] or "",
                    "customer_name": (r["customer_name"] if "customer_name" in r.keys() else "") or obj.get("customer_name", ""),
                    "salesperson_name": (r["salesperson_name"] if "salesperson_name" in r.keys() else "") or "",
                    "confidence": conf,
                    "matched_terms": ", ".join(strong + weak),
                    "snippet": snippet(text, (strong + weak + ["ticket"])[0]),
                })
    return hits


def scan_lines(conn, table: str, year, month) -> list[dict]:
    """Scan line-item sku / item_name / raw_json description."""
    rows = conn.execute(
        f"SELECT {table}.sku, {table}.item_name, {table}.raw_json, "
        f"{table}.line_item_id FROM {table}"
    ).fetchall()
    hits: list[dict] = []
    for r in rows:
        obj = parse_json(r["raw_json"])
        sources = {
            "sku": r["sku"] or "",
            "item_name": r["item_name"] or "",
            "raw.description": str(obj.get("description") or ""),
            "raw.name": str(obj.get("name") or ""),
        }
        for src_name, text in sources.items():
            conf, strong, weak = classify(text)
            if conf:
                hits.append({
                    "source_field": f"{table}.{src_name}",
                    "line_item_id": r["line_item_id"] or "",
                    "sku": r["sku"] or "",
                    "item_name": r["item_name"] or "",
                    "confidence": conf,
                    "matched_terms": ", ".join(strong + weak),
                    "snippet": snippet(text, (strong + weak + ["ticket"])[0]),
                })
    return hits


def scan_items(conn) -> list[dict]:
    """Scan the items catalog (sku / name / description)."""
    rows = conn.execute("SELECT sku, name, description, raw_json FROM items").fetchall()
    hits: list[dict] = []
    for r in rows:
        sources = {
            "sku": r["sku"] or "",
            "name": r["name"] or "",
            "description": r["description"] or "",
        }
        for src_name, text in sources.items():
            conf, strong, weak = classify(text)
            if conf:
                hits.append({
                    "source_field": f"items.{src_name}",
                    "sku": r["sku"] or "",
                    "name": r["name"] or "",
                    "confidence": conf,
                    "matched_terms": ", ".join(strong + weak),
                    "snippet": snippet(text, (strong + weak + ["ticket"])[0]),
                })
    return hits


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(sheets: dict[str, list[dict]], labels: Counter) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:  # openpyxl missing
        print(f"  (skipped Excel — openpyxl not available: {exc})")
        return False

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0E3B66")
    ws.append(["Ticket-Number Diagnostic"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Source", "Matches"])
    for c in ws[3]:
        c.font = head
        c.fill = fill
    for name, rows in sheets.items():
        ws.append([name, len(rows)])
    ws.append([])
    ws.append(["Distinct custom-field labels found:", len(labels)])
    ws.append(["(full list on the 'Custom Field Labels' sheet)"])

    # Custom field labels sheet — the key discovery output
    ws2 = wb.create_sheet("Custom Field Labels")
    ws2.append(["Custom Field Label", "Occurrences", "Looks ticket-related?"])
    for c in ws2[1]:
        c.font = head
        c.fill = fill
    for label, count in labels.most_common():
        low = label.lower()
        flag = "YES" if any(t in low for t in ("ticket", "tkt", "case", "rma")) else ""
        ws2.append([label, count, flag])

    # One sheet per match source
    for name, rows in sheets.items():
        wsx = wb.create_sheet(name[:31])
        if not rows:
            wsx.append(["(no matches)"])
            continue
        cols = list(rows[0].keys())
        wsx.append(cols)
        for c in wsx[1]:
            c.font = head
            c.fill = fill
        for row in rows:
            wsx.append([str(row.get(c, "")) for c in cols])

    try:
        wb.save(OUT_XLSX)
        return True
    except Exception as exc:
        print(f"  (could not write {OUT_XLSX}: {exc} — is it open in Excel?)")
        return False


# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Discover how ticket numbers appear in Zoho data (read-only).")
    ap.add_argument("--year", type=int, default=None)
    ap.add_argument("--month", type=int, default=None)
    args = ap.parse_args()

    print("=" * 72)
    print(" TICKET-NUMBER DIAGNOSTIC  (read-only — does not modify anything)")
    print("=" * 72)
    print(f" Database: {database_label()}")
    if args.year and args.month:
        print(f" NOTE: --year/--month given, but scan covers all rows for completeness.")
    print()

    with get_connection() as conn:
        labels, label_hits = scan_custom_field_labels(conn)
        so_hits = scan_documents(conn, "sales_orders", "salesorder_number",
                                 ["reference_number", "customer_name", "salesperson_name", "status"],
                                 args.year, args.month)
        inv_hits = scan_documents(conn, "invoices", "invoice_number",
                                  ["reference_number", "customer_name", "salesperson_name", "status"],
                                  args.year, args.month)
        so_line_hits = scan_lines(conn, "sales_order_lines", args.year, args.month)
        inv_line_hits = scan_lines(conn, "invoice_lines", args.year, args.month)
        item_hits = scan_items(conn)

    sheets = {
        "Custom Field Matches": label_hits,
        "Sales Order Matches": so_hits,
        "Invoice Matches": inv_hits,
        "SO Line Matches": so_line_hits,
        "Invoice Line Matches": inv_line_hits,
        "Item Catalog Matches": item_hits,
    }

    # ---- console report ----
    print(" MATCH COUNTS BY SOURCE")
    print(" " + "-" * 50)
    for name, rows in sheets.items():
        high = sum(1 for r in rows if str(r.get("confidence", "")).startswith("HIGH"))
        med = sum(1 for r in rows if str(r.get("confidence", "")).startswith("MEDIUM"))
        low = sum(1 for r in rows if str(r.get("confidence", "")).startswith("LOW"))
        print(f"  {name:<24} {len(rows):>5}   (HIGH {high} · MED {med} · LOW {low})")
    print()

    # dedicated ticket custom field?
    ticket_labels = [(l, c) for l, c in labels.most_common()
                     if any(t in l.lower() for t in ("ticket", "tkt", "case", "rma"))]
    print(" CUSTOM-FIELD LABEL SCAN")
    print(" " + "-" * 50)
    print(f"  distinct custom-field labels: {len(labels)}")
    if ticket_labels:
        print("  >>> POSSIBLE DEDICATED TICKET FIELD(S):")
        for l, c in ticket_labels:
            print(f"        • {l!r}  ({c} occurrences)")
        print("  >>> If one of these is consistent, USE IT as the detection rule.")
    else:
        print("  no custom-field label mentions ticket/case/rma.")
        print("  (ticket numbers likely live in notes / reference / SKU instead.)")
    print()

    # false-positive caution
    weak_only = sum(1 for rows in sheets.values() for r in rows
                    if str(r.get("confidence", "")).startswith("LOW"))
    print(" FALSE-POSITIVE RISK")
    print(" " + "-" * 50)
    print(f"  {weak_only} matches are LOW confidence (words like 'support', 'warranty',")
    print(f"  'service' that also appear in normal product names). Do NOT auto-exclude")
    print(f"  on these — review only.")
    print()

    # ---- Excel ----
    print(" WRITING EXCEL")
    print(" " + "-" * 50)
    if write_excel(sheets, labels):
        print(f"  wrote {OUT_XLSX}")
    print()

    # ---- recommendation ----
    print(" RECOMMENDED NEXT STEP")
    print(" " + "-" * 50)
    if ticket_labels:
        print("  A dedicated custom field exists — switch detection to read that field")
        print("  exactly, instead of guessing from SKU/name. Keep it review-only until")
        print("  Accounting confirms the field is the authoritative source.")
    else:
        high_total = sum(1 for rows in sheets.values() for r in rows
                         if str(r.get("confidence", "")).startswith("HIGH"))
        if high_total:
            print(f"  {high_total} HIGH-confidence 'ticket+number' matches found in free-text")
            print("  (notes/reference/SKU). A regex like (ticket|tkt|case|rma)[\\s#:.-]*\\d+")
            print("  is a reasonable detector — keep it review-only, verify against the xlsx.")
        else:
            print("  No strong ticket signal in the data. Ticket numbers may not be stored")
            print("  in Zoho consistently — keep TICKET_NUMBER as a manual review flag only.")
    print("=" * 72)


if __name__ == "__main__":
    main()
