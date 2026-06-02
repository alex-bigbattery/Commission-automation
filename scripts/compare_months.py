"""
Reconcile generated commission totals vs historical B2B Summary, per salesperson,
for the requested months. Read-only against SQLite + the historical workbooks.
"""

from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from src.commission.sqlite_to_workbook import (
    build_salespeople_from_sqlite,
    load_map_from_template,
    load_tiers_from_template,
)

HIST_ROOT = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001\Commissions\2026"
)
TPL = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"

MONTH_DIR = {
    1: "2026-1_January_Completed/2026-1_Commission B2B.xlsx",
    2: "2026-2_February_Completed/2026-2_Commission B2B.xlsx",
    3: "2026-3_March_Completed/2026-3_Commission B2B.xlsx",
}


def hist_totals(path: Path) -> dict[str, float]:
    """Per-salesperson total commission (current + prior) from B2B Summary."""
    wb = load_workbook(path, data_only=True)
    ws = wb["B2B Summary"]
    out: dict[str, float] = {}
    for r in range(20, 45):
        name = ws.cell(r, 5).value or ws.cell(r, 4).value
        if not name:
            continue
        name = str(name).strip()
        if name.lower() in ("salary commission", "non-salary commission", "b2b sales rep", "total b2b", "b2b executive account"):
            continue
        cur = ws.cell(r, 13).value or 0
        pri = ws.cell(r, 19).value or 0
        try:
            tot = float(cur) + float(pri)
        except (TypeError, ValueError):
            continue
        if abs(tot) > 0.005:
            out[name] = round(tot, 2)
    wb.close()
    return out


# Map historical display names -> our sheet keys
NAME_TO_SHEET = {
    "Paul Perlman": "Paul", "Jose Ayala": "Jose", "Michael Ayala": "Michael",
    "Jim Sutton": "Jim", "Weston Fields": "Weston", "Brett Bern": "Brett",
    "Leslie Neipert": "Leslie", "Carmen Daetz": "Carmen", "Garrett Lockhart": "Garrett",
    "B2B Company Account": "Company Acct", "Company Account": "Company Acct",
}


def main() -> None:
    tiers = load_tiers_from_template(TPL)
    rlp = load_map_from_template(TPL)
    grand_h = grand_o = 0.0
    for m in (1, 2, 3):
        hist_path = HIST_ROOT / MONTH_DIR[m]
        h = hist_totals(hist_path)
        res = build_salespeople_from_sqlite(2026, m, tiers=tiers, rlp_map=rlp)
        ours = {k: round(v, 2) for k, v in res.totals_by_sheet.items() if abs(v) > 0.005}

        # Normalize hist keys to sheet keys where possible
        h_sheet: dict[str, float] = {}
        for name, val in h.items():
            h_sheet[NAME_TO_SHEET.get(name, name)] = h_sheet.get(NAME_TO_SHEET.get(name, name), 0) + val

        keys = sorted(set(h_sheet) | set(ours))
        print("=" * 64)
        print(f"  2026-{m:02d}   (historical: {hist_path.name})")
        print("=" * 64)
        print(f"  {'Salesperson':<18}{'Ours':>12}{'Historical':>13}{'Diff':>12}")
        mh = mo = 0.0
        for k in keys:
            o = ours.get(k, 0.0)
            hv = h_sheet.get(k, 0.0)
            mh += hv
            mo += o
            flag = "" if abs(o - hv) < 0.01 else "  <-- diff"
            print(f"  {k:<18}{o:>12,.2f}{hv:>13,.2f}{o-hv:>12,.2f}{flag}")
        print(f"  {'TOTAL':<18}{mo:>12,.2f}{mh:>13,.2f}{mo-mh:>12,.2f}")
        print(f"  exceptions: {res.kpis.get('exceptions_count')}  commissionable_lines: {res.kpis.get('commissionable_lines')}")
        print()
        grand_h += mh
        grand_o += mo
    print("=" * 64)
    print(f"  GRAND TOTAL  ours={grand_o:,.2f}  hist={grand_h:,.2f}  diff={grand_o-grand_h:,.2f}")
    print("=" * 64)


if __name__ == "__main__":
    main()
