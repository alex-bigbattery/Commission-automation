"""
MISSING_MAP / possible-ticket diagnostic (READ-ONLY).

Finds commission-relevant lines that the engine CANNOT price because MAP is
missing (SKU not in R_LP / catalog) — these may be tickets or custom lines.
Does NOT call Zoho, does NOT modify historical workbooks, does NOT change
formulas, does NOT auto-exclude anything.

A line is reported when ALL hold:
  - revenue (item_total) > 0
  - it is a 'product' line with NO MAP, OR an 'other_charge' (blank-SKU custom line)
  - it is NOT a shipping charge, credit-card fee, or discount
  - ($0 kit components are already excluded by the revenue>0 filter)

Scope: Jan–May 2026, all sales teams (labelled).

Output:
  - console summary
  - data/output/missing_map_possible_ticket_report.xlsx

Run (no Zoho sync in progress):
  python scripts/missing_map_report.py
"""
from __future__ import annotations

import calendar
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from src.db.connection import get_connection
from src.commission.sqlite_to_workbook import (
    _load_invoice_meta_map, _load_item_map, _load_payment_dates,
    load_map_from_template, load_tiers_from_template,
    commission_rate, rate_type_for,
)
from src.commission.line_classification import classify_line_type
from src.commission.roster import (
    resolve_roster_sheet, classify_special_person, is_known_inactive,
    is_b2c_coupon_rep, ROUTING_UNASSIGNED,
)

TEMPLATE = REPO / "data" / "templates" / "master_template_clean.xlsx"
OUT_XLSX = REPO / "data" / "output" / "missing_map_possible_ticket_report.xlsx"
YEAR = 2026
MONTHS = [1, 2, 3, 4, 5]


def money(x) -> str:
    return f"${float(x or 0):,.2f}"


def team_bucket(team: str) -> str:
    t = (team or "").strip()
    tl = t.lower()
    if not t:
        return "unknown/blank"
    if tl == "b2b":
        return "B2B"
    if "exe" in tl or "comp. account" in tl:
        return "Exe./Comp. Account"
    if "rc team" in tl and "no commission" in tl:
        return "B2C non-commissionable"
    if "rc team" in tl:
        return "B2C - RC Team"
    if tl.startswith("b2c"):
        return "B2C - other (Web)"
    return t


def ar_status(balance: float, status: str, payment_date) -> str:
    s = (status or "").lower()
    if payment_date is not None:
        return "PAID"
    if balance == 0 or "paid" in s or "closed" in s:
        return "PAID"
    if balance > 0:
        return "UNPAID"
    return "REVIEW"


def payable_status(team: str, salesperson: str, ticket: str) -> str:
    """Mirror the engine's routing to say if a line would currently pay."""
    tl = (team or "").strip().lower()
    is_b2b = tl.startswith("b2b") or tl.startswith("exe") or "comp. account" in tl
    if not is_b2b:
        return "excluded (non-B2B sales team)"
    if classify_special_person(salesperson):
        return "held (Bruce/Marshall review)"
    if is_known_inactive(salesperson):
        return "held (inactive/non-B2B name)"
    if is_b2c_coupon_rep(salesperson):
        return "held (B2C coupon rep)"
    if (ticket or "").strip():
        return "held (Ticket#)"
    if resolve_roster_sheet(salesperson) is None:
        return "held (not in roster / pending)"
    return "PAYABLE"


def scan_month(conn, year, month, map_by_sku, tiers):
    meta = _load_invoice_meta_map(conn, year, month)
    inv_ids = list(meta.keys())
    if not inv_ids:
        return []
    ph = ",".join("?" * len(inv_ids))

    info = {}
    so_numbers = set()
    for r in conn.execute(
        f"SELECT invoice_id, invoice_number, salesorder_number, customer_name, "
        f"salesperson_name, balance, status FROM invoices WHERE invoice_id IN ({ph})",
        inv_ids,
    ).fetchall():
        info[str(r["invoice_id"])] = {
            "invoice": r["invoice_number"] or "",
            "so": r["salesorder_number"] or "",
            "customer": r["customer_name"] or "",
            "inv_sp": r["salesperson_name"] or "",
            "balance": float(r["balance"] or 0),
            "status": r["status"] or "",
        }
        if r["salesorder_number"]:
            so_numbers.add(r["salesorder_number"])

    so_sp = {}
    if so_numbers:
        sos = list(so_numbers)
        ph2 = ",".join("?" * len(sos))
        for r in conn.execute(
            f"SELECT salesorder_number, salesperson_name FROM sales_orders "
            f"WHERE salesorder_number IN ({ph2})", sos
        ).fetchall():
            so_sp[r["salesorder_number"]] = r["salesperson_name"] or ""

    pay_dates = _load_payment_dates(conn, inv_ids)

    out = []
    for r in conn.execute(
        f"SELECT invoice_id, sku, item_name, quantity, item_total, rate "
        f"FROM invoice_lines WHERE invoice_id IN ({ph})", inv_ids
    ).fetchall():
        iid = str(r["invoice_id"])
        sku = (r["sku"] or "").strip()
        name = r["item_name"] or ""
        qty = float(r["quantity"] or 0)
        rev = float(r["item_total"] or 0)
        rate0 = float(r["rate"] or 0)
        if rev <= 0:
            continue
        lt = classify_line_type(sku=sku, item_name=name, item_total=rev, quantity=qty, rate=rate0)
        if lt in ("shipping", "credit_card_fee", "discount", "unknown"):
            continue
        mp = map_by_sku.get(sku.upper(), 0.0) if sku else 0.0
        # keep: product with NO map, or other_charge (blank-SKU custom line)
        if lt == "product" and mp > 0:
            continue
        if lt not in ("product", "other_charge"):
            continue

        h = info.get(iid, {})
        team = meta.get(iid, {}).get("sales_team", "")
        ticket = meta.get(iid, {}).get("ticket_number", "")
        salesperson = so_sp.get(h.get("so", ""), "") or h.get("inv_sp", "")
        pstat = payable_status(team, salesperson, ticket)
        ar = ar_status(h.get("balance", 0), h.get("status", ""), pay_dates.get(iid))

        # exposure: only if it would currently be payable; estimate at 0% discount
        exposure = 0.0
        if pstat == "PAYABLE":
            rt = rate_type_for(resolve_roster_sheet(salesperson) or "Paul")
            exposure = rev * commission_rate(0.0, rt, tiers)

        issue = "MISSING_MAP (product, no MAP)" if lt == "product" else "custom/other charge (blank SKU)"
        out.append({
            "month": f"{year}-{month:02d}",
            "sales_order": h.get("so", ""),
            "invoice": h.get("invoice", ""),
            "customer": h.get("customer", ""),
            "salesperson": salesperson,
            "team_bucket": team_bucket(team),
            "team_raw": team,
            "sku": sku,
            "item_name": name,
            "revenue": rev,
            "payment_status": ar,
            "ticket_number": ticket,
            "issue_flag": issue,
            "line_type": lt,
            "payable_status": pstat,
            "est_commission_exposure": round(exposure, 2),
        })
    return out


def write_xlsx(rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        print(f"  (openpyxl unavailable, skipped xlsx: {exc})")
        return
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0E3B66")

    ws = wb.active
    ws.title = "Summary"
    ws.append(["MISSING_MAP / possible-ticket diagnostic"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Metric", "Value"])
    for c in ws[3]:
        c.font, c.fill = head, fill
    payable = [r for r in rows if r["payable_status"] == "PAYABLE"]
    ws.append(["Total lines", len(rows)])
    ws.append(["Sales Orders affected", len({r["sales_order"] for r in rows if r["sales_order"]})])
    ws.append(["Invoices affected", len({r["invoice"] for r in rows if r["invoice"]})])
    ws.append(["Total revenue affected", round(sum(r["revenue"] for r in rows), 2)])
    ws.append(["Currently-payable lines", len(payable)])
    ws.append(["Commission exposure (payable, est.)", round(sum(r["est_commission_exposure"] for r in rows), 2)])
    ws.append([])
    ws.append(["By team bucket", "lines", "revenue", "payable lines", "exposure"])
    for c in ws[ws.max_row]:
        c.font, c.fill = head, fill
    buckets = {}
    for r in rows:
        b = buckets.setdefault(r["team_bucket"], {"n": 0, "rev": 0.0, "pay": 0, "exp": 0.0})
        b["n"] += 1
        b["rev"] += r["revenue"]
        if r["payable_status"] == "PAYABLE":
            b["pay"] += 1
        b["exp"] += r["est_commission_exposure"]
    for name, b in sorted(buckets.items(), key=lambda kv: -kv[1]["rev"]):
        ws.append([name, b["n"], round(b["rev"], 2), b["pay"], round(b["exp"], 2)])

    wd = wb.create_sheet("Detail")
    cols = ["month", "sales_order", "invoice", "customer", "salesperson", "team_bucket",
            "team_raw", "sku", "item_name", "revenue", "payment_status", "ticket_number",
            "issue_flag", "line_type", "payable_status", "est_commission_exposure"]
    wd.append(cols)
    for c in wd[1]:
        c.font, c.fill = head, fill
    for r in sorted(rows, key=lambda x: -x["revenue"]):
        wd.append([r[c] for c in cols])

    wb.save(OUT_XLSX)
    print(f"  wrote {OUT_XLSX}")


def main():
    tiers = load_tiers_from_template(TEMPLATE)
    rlp = load_map_from_template(TEMPLATE)

    print("=" * 72)
    print(" MISSING_MAP / POSSIBLE-TICKET DIAGNOSTIC (read-only)")
    print(f" Months: {MONTHS}  Year: {YEAR}  | all sales teams")
    print("=" * 72)

    all_rows = []
    with get_connection() as conn:
        item_map = _load_item_map(conn)
        map_by_sku = {**item_map, **rlp}
        for m in MONTHS:
            rows = scan_month(conn, YEAR, m, map_by_sku, tiers)
            all_rows.extend(rows)
            print(f"  {calendar.month_name[m]:<9}: {len(rows)} lines, "
                  f"revenue {money(sum(r['revenue'] for r in rows))}")

    sos = {r["sales_order"] for r in all_rows if r["sales_order"]}
    invs = {r["invoice"] for r in all_rows if r["invoice"]}
    rev = sum(r["revenue"] for r in all_rows)
    payable = [r for r in all_rows if r["payable_status"] == "PAYABLE"]
    exposure = sum(r["est_commission_exposure"] for r in all_rows)

    print("\n REQUESTED REPORT")
    print(" " + "-" * 60)
    print(f" 1. MISSING_MAP count               : {len(all_rows)}")
    print(f" 2. Sales Orders affected           : {len(sos)}")
    print(f" 3. Invoices affected               : {len(invs)}")
    print(f" 4. Total revenue affected          : {money(rev)}")
    print(f" 5. Commission exposure (payable)   : {money(exposure)}  "
          f"({len(payable)} payable lines)")

    print("\n By team bucket:")
    by = {}
    for r in all_rows:
        b = by.setdefault(r["team_bucket"], {"n": 0, "rev": 0.0, "pay": 0, "exp": 0.0})
        b["n"] += 1
        b["rev"] += r["revenue"]
        if r["payable_status"] == "PAYABLE":
            b["pay"] += 1
        b["exp"] += r["est_commission_exposure"]
    for name, b in sorted(by.items(), key=lambda kv: -kv[1]["rev"]):
        print(f"   {name:<26} {b['n']:>4} lines  rev {money(b['rev']):>14}  "
              f"payable {b['pay']:>3}  exposure {money(b['exp'])}")

    print("\n 6. EXAMPLES (top 15 by revenue):")
    for r in sorted(all_rows, key=lambda x: -x["revenue"])[:15]:
        print(f"   {r['invoice']}/{r['sales_order']} | {r['customer'][:22]:<22} | "
              f"{r['salesperson'][:16]:<16} | {r['team_bucket'][:20]:<20} | "
              f"{(r['sku'] or '(blank)')[:16]:<16} {r['item_name'][:24]:<24} | "
              f"rev {money(r['revenue']):>12} | {r['payment_status']:<6} | "
              f"tkt={r['ticket_number'] or '-':<6} | {r['payable_status']}")

    print("\n WRITING XLSX")
    write_xlsx(all_rows)
    print("=" * 72)


if __name__ == "__main__":
    main()
