"""
Test build_master_workbook with Brett's real March 2026 data
(other salespeople blank).
"""

from __future__ import annotations

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from src.commission.workbook_builder_v2 import (
    Block,
    DetailRow,
    SalespersonData,
    build_master_workbook,
)


HISTORICAL_BRETT = Path(
    r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001"
    r"\Commissions\2026\2026-3_March_Completed\2026-3_Commission B2B_BRETT BERN.xlsx"
)
MASTER_TEMPLATE = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"
OUTPUT = BASE_DIR / "data" / "output" / "_test_master_march_2026.xlsx"


SALESPEOPLE = {
    "Paul": ("PAUL PERLMAN", "salaried"),
    "Jose": ("JOSE AYALA", "salaried"),
    "Michael": ("MICHAEL AYALA", "salaried"),
    "Jim": ("JIM SUTTON", "salaried"),
    "Weston": ("WESTON FIELDS", "salaried"),
    "Brett": ("BRETT BERN", "non_salaried"),
    "Leslie": ("LESLIE NEIPERT", "non_salaried"),
    "Carmen": ("CARMEN DAETZ", "non_salaried"),
    "Garrett": ("GARRETT LOCKHART", "non_salaried"),
    "Company Acct": ("COMPANY ACCOUNT", "salaried"),
}


def read_block_rows(ws, header_row: int, subtotal_row: int) -> list[DetailRow]:
    rows: list[DetailRow] = []
    for r in range(header_row + 1, subtotal_row):
        sku = ws.cell(r, 11).value
        if not sku:
            continue

        def num(c):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(v) if v not in (None, "") else 0.0
            except (ValueError, TypeError):
                return 0.0

        rows.append(
            DetailRow(
                order_date=ws.cell(r, 4).value,
                sales_order_number=ws.cell(r, 5).value or "",
                invoice_date=ws.cell(r, 6).value,
                invoice_number=ws.cell(r, 7).value or "",
                invoice_status=ws.cell(r, 8).value or "",
                customer_name=ws.cell(r, 9).value or "",
                estimate_number=ws.cell(r, 10).value or "",
                sku=sku or "",
                quantity=num(12),
                item_total=num(13),
                account=ws.cell(r, 14).value or "",
                account_code=str(ws.cell(r, 15).value or ""),
                payment_terms=ws.cell(r, 16).value or "",
                delivery_method=ws.cell(r, 17).value or "",
                shipment_date=ws.cell(r, 19).value,
                shipment_status=ws.cell(r, 20).value or "",
                ar_status=ws.cell(r, 21).value or "",
                payment_date=ws.cell(r, 22).value,
                shipping_method=ws.cell(r, 24).value or "",
                reason=ws.cell(r, 25).value or "",
                include_in_commission=ws.cell(r, 26).value or "No",
                shipping_income=num(27),
                shipping_expenses=num(28),
                discount_rate=ws.cell(r, 34).value if isinstance(ws.cell(r, 34).value, (int, float)) else None,
            )
        )
    return rows


def load_brett_data() -> SalespersonData:
    wb = load_workbook(HISTORICAL_BRETT, data_only=False)
    ws = wb["Brett"]
    sp = SalespersonData(
        name="Brett",
        full_name="BRETT BERN",
        rate_type="non_salaried",
        month_name="March",
        year=2026,
        current_commissionable=Block("ORDERS COMMISSIONABLE - SHIPPED AND INVOICED", read_block_rows(ws, 23, 45)),
        current_shipping=Block("SHIPPING INCOME", read_block_rows(ws, 48, 53)),
        current_other=Block("OTHER CHARGES", read_block_rows(ws, 56, 59)),
        prior_commissionable=Block("ORDERS COMMISSIONABLE", read_block_rows(ws, 64, 66)),
        prior_shipping=Block("SHIPPING CHARGE", read_block_rows(ws, 69, 71)),
    )
    wb.close()
    return sp


def empty_salesperson(name: str, full_name: str, rate_type: str, month_name: str, year: int) -> SalespersonData:
    return SalespersonData(
        name=name,
        full_name=full_name,
        rate_type=rate_type,
        month_name=month_name,
        year=year,
        current_commissionable=Block("ORDERS COMMISSIONABLE - SHIPPED AND INVOICED"),
        current_shipping=Block("SHIPPING INCOME"),
        current_other=Block("OTHER CHARGES"),
        prior_commissionable=Block("ORDERS COMMISSIONABLE"),
        prior_shipping=Block("SHIPPING CHARGE"),
    )


def main() -> None:
    print("Building salespeople data...")
    salespeople = []
    for name, (full_name, rate_type) in SALESPEOPLE.items():
        if name == "Brett":
            salespeople.append(load_brett_data())
        else:
            salespeople.append(empty_salesperson(name, full_name, rate_type, "March", 2026))

    print(f"  Total: {len(salespeople)} salespeople")
    print(f"  Brett rows: I.1={len(salespeople[5].current_commissionable.rows)}")

    print(f"\nBuilding master from: {MASTER_TEMPLATE.name}")
    result = build_master_workbook(
        template_path=MASTER_TEMPLATE,
        output_path=OUTPUT,
        salespeople=salespeople,
        month_name="March",
        year=2026,
    )
    print(f"OK: wrote {result}")
    print(f"  Size: {result.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
