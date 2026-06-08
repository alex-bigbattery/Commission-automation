"""Read-only inventory of FV_PRICE / MAP sources and dry-run load plans."""
from __future__ import annotations

import calendar
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

from dotenv import load_dotenv
load_dotenv(REPO / ".env")

import openpyxl  # noqa: E402

from src.db.connection import get_connection, using_postgres  # noqa: E402


DOWNLOADS = Path.home() / "Downloads"
REPO_DATA = REPO / "data"

CANDIDATES: list[tuple[str, Path, str]] = [
    ("2026-03 primary", DOWNLOADS / "2026-3_Commission B2B.xlsx", "2026-03"),
    ("2026-03 alt", DOWNLOADS / "commission_b2b_march_2026.xlsx", "2026-03"),
    ("2026-04 primary", DOWNLOADS / "2026-4_Commission B2B.xlsx", "2026-04"),
    ("2026-04 alt", DOWNLOADS / "commission_b2b_april_2026.xlsx", "2026-04"),
    ("2026-04 loader default", DOWNLOADS / "2026-04_Commissions_B2B.xlsx", "2026-04"),
    ("2026-05 primary", DOWNLOADS / "commission_b2b_may_2026.xlsx", "2026-05"),
    ("2026-05 alt", DOWNLOADS / "commission_b2b_may_2026 (3).xlsx", "2026-05"),
    ("master template", REPO_DATA / "templates" / "master_template_clean.xlsx", "template"),
]


PRICE_SHEETS = ("FV_PRICE", "Price_Base", "MAP", "R_LP", "items", "B2B_Commission")


def month_bounds(month: str) -> tuple[str, str]:
    y, m = (int(p) for p in month.split("-", 1))
    last = calendar.monthrange(y, m)[1]
    return f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}"


def parse_sheet(ws, sku_col: int, price_col: int) -> tuple[list[tuple[str, float]], dict[str, int]]:
    rows: list[tuple[str, float]] = []
    seen: set[str] = set()
    stats = {"skipped_zero": 0, "skipped_blank": 0, "skipped_dup": 0, "skipped_header": 0}
    for r in range(1, ws.max_row + 1):
        sku_raw = ws.cell(r, sku_col).value
        price_raw = ws.cell(r, price_col).value
        if sku_raw is None:
            continue
        sku = str(sku_raw).strip().upper()
        if not sku or sku in ("SKU", "ITEM", "ITEM SKU", "SKU/ITEM", "PRICES", "COMPONENTS"):
            stats["skipped_header"] += 1
            continue
        try:
            price = float(price_raw)
        except (TypeError, ValueError):
            stats["skipped_blank"] += 1
            continue
        if price <= 0:
            stats["skipped_zero"] += 1
            continue
        if sku in seen:
            stats["skipped_dup"] += 1
            continue
        seen.add(sku)
        rows.append((sku, price))
    return rows, stats


def inspect_workbook(label: str, path: Path, month_hint: str) -> dict[str, Any] | None:
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sheets = list(wb.sheetnames)
    best: dict[str, Any] | None = None
    for sheet in sheets:
        if sheet not in PRICE_SHEETS and sheet != "R_LP":
            continue
        ws = wb[sheet]
        # FV_PRICE: B/C; R_LP: A/G
        if sheet == "R_LP":
            sku_col, price_col = 1, 7
        else:
            sku_col, price_col = 2, 3
        rows, stats = parse_sheet(ws, sku_col, price_col)
        if not rows:
            continue
        sample = rows[:5]
        entry = {
            "label": label,
            "path": str(path),
            "month_hint": month_hint,
            "sheet": sheet,
            "sku_col": sku_col,
            "price_col": price_col,
            "sku_count": len(rows),
            "stats": stats,
            "sample": sample,
            "all_sheets": sheets,
        }
        if sheet == "FV_PRICE" or (best is None) or (best["sheet"] != "FV_PRICE" and len(rows) > best["sku_count"]):
            best = entry
    wb.close()
    return best


def db_state(conn) -> dict[str, Any]:
    total = conn.execute("SELECT COUNT(*) AS c FROM price_history").fetchone()["c"]
    skus = conn.execute(
        "SELECT COUNT(DISTINCT UPPER(sku)) AS c FROM price_history WHERE sku IS NOT NULL AND sku != ''"
    ).fetchone()["c"]
    by_source = conn.execute(
        "SELECT snapshot_month, source, COUNT(*) AS n, MIN(effective_from) AS f, MAX(effective_to) AS t "
        "FROM price_history GROUP BY snapshot_month, source ORDER BY snapshot_month, source"
    ).fetchall()
    bounds = conn.execute(
        "SELECT MIN(effective_from) AS min_f, MAX(effective_to) AS max_t FROM price_history"
    ).fetchone()
    pws = conn.execute(
        "SELECT sku, map_price, effective_from, effective_to, source, snapshot_month "
        "FROM price_history WHERE UPPER(sku)='PWS015' ORDER BY effective_from"
    ).fetchall()
    return {
        "backend": "postgres" if using_postgres() else "sqlite",
        "total": total,
        "skus": skus,
        "by_source": [dict(r) for r in by_source],
        "min_from": bounds["min_f"],
        "max_to": bounds["max_t"],
        "pws015": [dict(r) for r in pws],
    }


def plan_load(
    conn,
    *,
    source_file: dict[str, Any],
    snapshot_month: str,
    source_label: str,
) -> dict[str, Any]:
    eff_from, eff_to = month_bounds(snapshot_month)
    parsed = [(sku, price) for sku, price in source_file["sample"]]  # placeholder
    # re-parse full file
    wb = openpyxl.load_workbook(source_file["path"], data_only=True)
    ws = wb[source_file["sheet"]]
    rows, stats = parse_sheet(ws, source_file["sku_col"], source_file["price_col"])
    wb.close()

    existing = conn.execute(
        "SELECT sku, map_price, effective_from, effective_to, source "
        "FROM price_history WHERE snapshot_month=? AND source=?",
        (snapshot_month, source_label),
    ).fetchall()
    existing_by_sku = {str(r["sku"]).upper(): dict(r) for r in existing}

    would_insert = 0
    unchanged = 0
    price_conflict = 0
    overlap_risk: list[str] = []

    for sku, price in rows:
        ex = existing_by_sku.get(sku)
        if ex is None:
            would_insert += 1
        elif float(ex["map_price"]) == price:
            unchanged += 1
        else:
            price_conflict += 1

    # Check other sources same month different source
    other_same_month = conn.execute(
        "SELECT source, COUNT(*) AS n FROM price_history WHERE snapshot_month=? GROUP BY source",
        (snapshot_month,),
    ).fetchall()

    # Window overlap with live for sample SKU PWS015
    pws_rows = [r for r in rows if r[0] == "PWS015"]
    live_pws = conn.execute(
        "SELECT * FROM price_history WHERE UPPER(sku)='PWS015' AND snapshot_month='live'"
    ).fetchall()

    return {
        "source_file": source_file["path"],
        "sheet": source_file["sheet"],
        "snapshot_month": snapshot_month,
        "source_label": source_label,
        "effective_from": eff_from,
        "effective_to": eff_to,
        "parsed_valid": len(rows),
        "parse_stats": stats,
        "dup_skus_in_file": stats["skipped_dup"],
        "existing_rows_same_source": len(existing),
        "would_insert": would_insert,
        "unchanged": unchanged,
        "price_conflict_same_source": price_conflict,
        "other_sources_same_month": [dict(r) for r in other_same_month],
        "pws015_in_file": pws_rows,
        "pws015_live_exists": len(live_pws) > 0,
        "sample_rows": rows[:5],
        "touches_existing": price_conflict > 0,
        "safe_if_no_conflicts": price_conflict == 0,
    }


def main() -> None:
    print("=" * 72)
    print("STEP 1 — Historical source inventory")
    print("=" * 72)
    found: dict[str, dict[str, Any]] = {}
    for label, path, month in CANDIDATES:
        info = inspect_workbook(label, path, month)
        if info is None:
            print(f"\n[MISSING or no price sheet] {label}: {path}")
            continue
        key = f"{month}:{path.name}"
        found[key] = info
        print(f"\n[FOUND] {label}")
        print(f"  path   : {path}")
        print(f"  sheet  : {info['sheet']}  (cols sku={info['sku_col']} price={info['price_col']})")
        print(f"  month  : {month}")
        print(f"  SKUs   : {info['sku_count']}")
        print(f"  stats  : {info['stats']}")
        print(f"  sample : {info['sample']}")
        print(f"  sheets : {', '.join(info['all_sheets'][:12])}{'...' if len(info['all_sheets'])>12 else ''}")

    print("\n" + "=" * 72)
    print("STEP 2 — Current DB state (read-only)")
    print("=" * 72)
    conn = get_connection()
    state = db_state(conn)
    print(f"database_backend : {state['backend']}")
    print(f"price_history    : {state['total']} rows, {state['skus']} SKUs")
    print(f"date range       : {state['min_from']} .. {state['max_to']}")
    print("by snapshot/source:")
    for r in state["by_source"]:
        print(f"  {r['snapshot_month']:8} | {r['source']:35} | n={r['n']:4} | {r['f']} -> {r['t']}")
    print("PWS015:")
    for r in state["pws015"]:
        print(f"  {r}")

    print("\n" + "=" * 72)
    print("STEP 3-4 — Dry-run load plans (no writes)")
    print("=" * 72)

    plans: list[dict[str, Any]] = []
    month_sources = [
        ("2026-03", "accountant_fvprice_2026_03", [
            DOWNLOADS / "2026-3_Commission B2B.xlsx",
            DOWNLOADS / "commission_b2b_march_2026.xlsx",
        ]),
        ("2026-04", "accountant_fvprice_2026_04", [
            DOWNLOADS / "2026-4_Commission B2B.xlsx",
            DOWNLOADS / "commission_b2b_april_2026.xlsx",
            DOWNLOADS / "2026-04_Commissions_B2B.xlsx",
        ]),
        ("2026-05", "accountant_fvprice_2026_05", [
            DOWNLOADS / "commission_b2b_may_2026.xlsx",
            DOWNLOADS / "commission_b2b_may_2026 (3).xlsx",
        ]),
    ]
    for month, source_label, paths in month_sources:
        print(f"\n--- {month} ({source_label}) ---")
        chosen = None
        for p in paths:
            for k, info in found.items():
                if Path(info["path"]) == p and info["month_hint"] == month:
                    chosen = info
                    break
            if chosen:
                break
        if not chosen:
            print("  MISSING: no workbook found on disk for this month")
            continue
        plan = plan_load(conn, source_file=chosen, snapshot_month=month, source_label=source_label)
        plans.append(plan)
        print(f"  file             : {Path(plan['source_file']).name}")
        print(f"  sheet            : {plan['sheet']}")
        print(f"  window           : {plan['effective_from']} .. {plan['effective_to']}")
        print(f"  parsed valid     : {plan['parsed_valid']}")
        print(f"  would_insert     : {plan['would_insert']}")
        print(f"  unchanged        : {plan['unchanged']}")
        print(f"  price_conflicts  : {plan['price_conflict_same_source']}")
        print(f"  existing same src: {plan['existing_rows_same_source']}")
        print(f"  other src month  : {plan['other_sources_same_month']}")
        print(f"  PWS015 in file   : {plan['pws015_in_file']}")
        print(f"  safe (no conflict): {plan['safe_if_no_conflicts']}")

    conn.close()
    print("\nDone. No data was modified.")


if __name__ == "__main__":
    main()
