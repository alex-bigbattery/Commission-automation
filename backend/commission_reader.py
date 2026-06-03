from __future__ import annotations

import base64
import os
from datetime import date, datetime
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

# Root folder for the historical "Commissions / History" workbooks.
# Override with the COMMISSIONS_ROOT env var when deploying off the local PC
# (e.g. a mounted disk on Render). Falls back to the local Windows path.
_commissions_root_env = os.environ.get("COMMISSIONS_ROOT", "")
COMMISSIONS_DIR = Path(
    _commissions_root_env
    or r"C:\Users\Bigbattery\Downloads\Commissions-20260529T132541Z-3-001\Commissions"
)
if not _commissions_root_env and not COMMISSIONS_DIR.exists():
    import warnings
    warnings.warn(
        "COMMISSIONS_ROOT env var is not set and the default path does not exist. "
        "History workbook lookups will fail. Set COMMISSIONS_ROOT in production.",
        stacklevel=1,
    )

SHEET_GROUPS = {
    "summary": ["B2B Summary"],
    "salespeople": [
        "Paul",
        "Jose",
        "Michael",
        "Jim",
        "Weston",
        "Brett",
        "Carmen",
        "Leslie",
        "Garrett",
        "Sawyer",
        "Company Acct",
    ],
    "reference": ["Table", "R_LP", "R_SO", "R_INV", "R_SH"],
}


def _has_b2b_summary(path: Path) -> bool:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return "B2B Summary" in names
    except Exception:
        return False


def _workbook_score(path: Path) -> int:
    name = path.name.lower()
    score = 0
    if re.search(r"commission[ _]b2b", name):
        score += 50
    if "adjusted" in name:
        score += 20
    # Keep tree loading fast: avoid opening each workbook during scoring.
    if "b2b" in name:
        score += 40
    if "original" in name or "do not pay" in name:
        score -= 80
    if name.count("_") > 3:
        score -= 10
    return score


def _pick_workbook(folder: Path) -> Path | None:
    candidates = _workbook_candidates(folder)
    if not candidates:
        return None
    return sorted(candidates, key=lambda p: (_workbook_score(p), -p.stat().st_mtime), reverse=True)[0]


def _workbook_candidates(folder: Path) -> list[Path]:
    seen: dict[str, Path] = {}
    for pattern in ("*Commission B2B*.xlsx", "*Commission_B2B*.xlsx"):
        for p in folder.glob(pattern):
            seen[str(p.resolve())] = p
    return list(seen.values())


def _workbook_kind(path: Path) -> str:
    name = path.name.lower()
    if re.search(r"commission[ _]b2b(?:_adjusted)?\.xlsx$", name):
        return "full"
    if re.search(r"commission[ _]b2b", name):
        return "individual"
    return "other"


def _encode_workbook_id(path: Path) -> str:
    rel = path.relative_to(COMMISSIONS_DIR).as_posix().encode("utf-8")
    token = base64.urlsafe_b64encode(rel).decode("ascii").rstrip("=")
    return f"b64_{token}"


def _decode_workbook_id(workbook_id: str) -> Path | None:
    if not workbook_id.startswith("b64_"):
        return None
    token = workbook_id[4:]
    padded = token + ("=" * ((4 - len(token) % 4) % 4))
    try:
        rel = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
    except Exception:
        return None
    candidate = (COMMISSIONS_DIR / rel).resolve()
    base = COMMISSIONS_DIR.resolve()
    if str(candidate).startswith(str(base)) and candidate.exists():
        return candidate
    return None


def _month_sort_key(label: str) -> tuple[int, str]:
    m = re.search(r"(\d{1,2})", label)
    month_num = int(m.group(1)) if m else 99
    return (month_num, label.lower())


def build_commissions_tree() -> dict[str, Any]:
    if not COMMISSIONS_DIR.exists():
        return {"root": str(COMMISSIONS_DIR), "years": []}

    years: list[dict[str, Any]] = []
    for year_dir in sorted(COMMISSIONS_DIR.iterdir(), key=lambda p: p.name, reverse=True):
        if not year_dir.is_dir() or not year_dir.name.isdigit():
            continue

        months: list[dict[str, Any]] = []
        for month_dir in sorted(year_dir.iterdir(), key=lambda p: _month_sort_key(p.name)):
            if not month_dir.is_dir():
                continue
            workbook_paths = _workbook_candidates(month_dir)
            if not workbook_paths:
                continue
            workbook_paths = sorted(
                workbook_paths,
                key=lambda p: (
                    0 if _workbook_kind(p) == "full" else 1,
                    -p.stat().st_mtime,
                    p.name.lower(),
                ),
            )
            workbooks = [
                {
                    "id": _encode_workbook_id(path),
                    "label": path.name,
                    "kind": _workbook_kind(path),
                    "workbook_path": str(path),
                }
                for path in workbook_paths
            ]
            default_workbook = next((wb for wb in workbooks if wb["kind"] == "full"), workbooks[0])
            months.append(
                {
                    "id": f"{year_dir.name}/{month_dir.name}",
                    "label": month_dir.name,
                    "workbook_id": default_workbook["id"],  # Backward compatibility.
                    "workbook_path": default_workbook["workbook_path"],  # Backward compatibility.
                    "workbooks": workbooks,
                    "default_workbook_id": default_workbook["id"],
                }
            )

        if months:
            years.append({"year": year_dir.name, "months": months})

    return {"root": str(COMMISSIONS_DIR), "years": years}


def list_workbook_sheets(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    grouped: dict[str, list[str]] = {key: [] for key in SHEET_GROUPS}
    grouped["other"] = []
    known = {name for names in SHEET_GROUPS.values() for name in names}

    for sheet in sheets:
        placed = False
        for group, names in SHEET_GROUPS.items():
            if sheet in names:
                grouped[group].append(sheet)
                placed = True
                break
        if not placed:
            grouped["other"].append(sheet)

    return {"sheets": sheets, "groups": grouped}


def _rgb_to_hex(value: str | None) -> str | None:
    if not value or value in {"00000000", "0", "None"}:
        return None
    text = str(value)
    if len(text) == 8:
        return f"#{text[2:].lower()}"
    if len(text) == 6:
        return f"#{text.lower()}"
    return None


def _cell_style(cell: Cell) -> dict[str, Any]:
    fill_hex = None
    if cell.fill and cell.fill.fill_type == "solid":
        fill_hex = _rgb_to_hex(getattr(cell.fill.start_color, "rgb", None))

    font = cell.font
    font_color = _rgb_to_hex(getattr(font.color, "rgb", None)) if font and font.color else None

    return {
        "bold": bool(font.bold) if font else False,
        "italic": bool(font.italic) if font else False,
        "underline": bool(getattr(font, "underline", None)) if font else False,
        "bg": fill_hex,
        "color": font_color,
        "align": (cell.alignment.horizontal or "general") if cell.alignment else "general",
        "num_fmt": getattr(cell, "number_format", None) or "",
        # MergedCell (non-anchor cells of a merged range) has no `is_date` property.
        "is_date": bool(getattr(cell, "is_date", False)),
        "has_border": bool(
            cell.border and any(getattr(side, "style", None) for side in (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom))
        ),
    }


def _format_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            return int(round(value))
        return round(value, 6)
    return value


def _excel_col_name(col_num: int) -> str:
    name = ""
    n = col_num
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def read_sheet_grid(path: Path, sheet_name: str, *, max_rows: int = 1200, max_cols: int = 80) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    ws = wb[sheet_name]
    used_rows = min(ws.max_row or 1, max_rows)
    used_cols = min(ws.max_column or 1, max_cols)

    rows: list[list[dict[str, Any]]] = []
    last_content_row = 0

    for r in range(1, used_rows + 1):
        row_cells: list[dict[str, Any]] = []
        row_has_content = False
        for c in range(1, used_cols + 1):
            cell = ws.cell(r, c)
            value = _format_value(cell.value)
            if value != "":
                row_has_content = True
            style = _cell_style(cell)
            row_cells.append({"v": value, **style})
        if row_has_content:
            last_content_row = r
        rows.append(row_cells)

    wb.close()

    if last_content_row:
        rows = rows[:last_content_row]

    columns = [_excel_col_name(i) for i in range(1, used_cols + 1)]
    return {
        "sheet": sheet_name,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "col_count": used_cols,
    }


def resolve_workbook(workbook_id: str) -> Path:
    decoded = _decode_workbook_id(workbook_id)
    if decoded is not None:
        return decoded
    matches = sorted(COMMISSIONS_DIR.rglob(workbook_id))
    if not matches:
        raise FileNotFoundError(f"Workbook '{workbook_id}' not found.")
    return matches[0]
