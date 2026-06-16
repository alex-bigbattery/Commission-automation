"""Tests for the B2C RC-Team commission engine.

Pins the rules transcribed from the April 2026 accountant workbook: only RC Team
sales teams are eligible, the rate is a flat 2%, the "(No Commissionable)" RC
variant and other B2C teams are excluded, and the pool is commissionable * rate.
"""
from __future__ import annotations

import json

import pytest

from src.commission.b2c_commission import build_b2c_commission, write_b2c_workbook
from src.db.connection import get_connection, init_database


YEAR, MONTH = 2026, 4
SYNCED = "2026-04-30T00:00:00"


def _seed_line(conn, *, so_id, so_number, invoice_id, invoice_number,
               salesperson, sales_team, sku, qty, rate, item_total, returned=0):
    conn.execute(
        "INSERT OR IGNORE INTO items (item_id, sku, name, rate, raw_json, synced_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (sku, sku, sku, rate, "{}", SYNCED),
    )
    so_raw = {"line_items": [{
        "sku": sku, "line_item_id": f"{so_id}-L1",
        "quantity": qty, "quantity_invoiced": qty,
        "quantity_shipped": qty, "quantity_returned": returned,
    }]}
    conn.execute(
        """
        INSERT INTO sales_orders
          (salesorder_id, salesorder_number, order_date, customer_name,
           salesperson_name, delivery_method, raw_json, synced_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (so_id, so_number, f"{YEAR:04d}-{MONTH:02d}-05", "Acme Co",
         salesperson, "Freight", json.dumps(so_raw), SYNCED),
    )
    inv_raw = {"custom_fields": [{"label": "Sales Team", "value": sales_team}]}
    conn.execute(
        """
        INSERT INTO invoices
          (invoice_id, invoice_number, invoice_date, salesorder_number, status,
           balance, customer_name, salesperson_name, raw_json, synced_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (invoice_id, invoice_number, f"{YEAR:04d}-{MONTH:02d}-10", so_number,
         "paid", 0.0, "Acme Co", "", json.dumps(inv_raw), SYNCED),
    )
    conn.execute(
        """
        INSERT INTO invoice_lines
          (invoice_id, line_index, sku, item_name, quantity, rate, item_total,
           raw_json, synced_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (invoice_id, 0, sku, sku, qty, rate, item_total, "{}", SYNCED),
    )


def _seed(db_path, lines):
    init_database(db_path)
    conn = get_connection(db_path)
    try:
        for i, ln in enumerate(lines):
            _seed_line(conn, so_id=f"SO{i}", so_number=f"SO-{i}",
                       invoice_id=f"INV{i}", invoice_number=f"INV-{i}", **ln)
        conn.commit()
    finally:
        conn.close()


def test_rc_team_line_pays_flat_two_percent(tmp_path):
    db_path = tmp_path / "b2c.sqlite"
    _seed(db_path, [dict(
        salesperson="Dylan Nava", sales_team="B2C Web - RC Team",
        sku="WIDGET", qty=1, rate=650.0, item_total=617.5,
    )])

    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)

    assert result.rate == 0.02
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row["item_total"] == pytest.approx(617.5)
    assert row["commission_rate"] == 0.02
    assert row["commission_amount"] == pytest.approx(12.35)  # 617.5 * 0.02
    assert row["discount_rate"] == pytest.approx(0.05)        # 1 - 617.5/650
    assert result.pool_commissionable == pytest.approx(617.5)
    assert result.pool_commission == pytest.approx(12.35)
    assert result.by_salesperson["Dylan Nava"] == pytest.approx(12.35)


def test_other_b2c_teams_are_excluded(tmp_path):
    db_path = tmp_path / "excluded.sqlite"
    _seed(db_path, [
        dict(salesperson="Customer Service", sales_team="B2C - RC Team (No Commissionable)",
             sku="A", qty=1, rate=100.0, item_total=100.0),
        dict(salesperson="Customer Service", sales_team="B2C Web - Marketing",
             sku="B", qty=1, rate=100.0, item_total=100.0),
        dict(salesperson="Customer Service", sales_team="B2C Web - Affiliate",
             sku="C", qty=1, rate=100.0, item_total=100.0),
        dict(salesperson="Customer Service", sales_team="B2C Web - Website (Organic Sales)",
             sku="D", qty=1, rate=100.0, item_total=100.0),
        dict(salesperson="Paul Perlman", sales_team="B2B",
             sku="E", qty=1, rate=100.0, item_total=100.0),
    ])

    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)
    assert result.rows == []
    assert result.pool_commission == 0.0


def test_both_rc_variants_included_and_pooled(tmp_path):
    db_path = tmp_path / "pool.sqlite"
    _seed(db_path, [
        dict(salesperson="Customer Service", sales_team="B2C Web - RC Team",
             sku="A", qty=1, rate=1000.0, item_total=1000.0),
        dict(salesperson="Dylan Nava", sales_team="B2C - RC Team",
             sku="B", qty=1, rate=500.0, item_total=500.0),
    ])

    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)
    assert len(result.rows) == 2
    assert result.pool_commissionable == pytest.approx(1500.0)
    assert result.pool_commission == pytest.approx(30.0)  # 1500 * 0.02
    # Per-rep subtotals are surfaced, but the split itself is left to Accounting.
    assert result.by_salesperson == {
        "Customer Service": pytest.approx(20.0),
        "Dylan Nava": pytest.approx(10.0),
    }
    assert result.kpis["split_is_manual"] is True


def test_fully_returned_rc_line_not_commissionable(tmp_path):
    db_path = tmp_path / "returned.sqlite"
    _seed(db_path, [dict(
        salesperson="Dylan Nava", sales_team="B2C Web - RC Team",
        sku="WIDGET", qty=1, rate=100.0, item_total=100.0, returned=1,
    )])
    # Return date defaults to the SO with no salesreturns -> qty_returned only;
    # fully returned with no later RMA date is conservatively excluded.
    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)
    assert result.pool_commission == 0.0
    assert result.rows[0]["commission_amount"] == 0.0
    assert "FULLY_RETURNED" in result.rows[0]["flags"]


def test_miscellaneous_sku_excluded(tmp_path):
    # A "MISCELLANEOUS" placeholder SKU on an eligible RC Team invoice must NOT
    # earn commission (matches the accountant's April 2026 treatment).
    db_path = tmp_path / "misc.sqlite"
    _seed(db_path, [
        dict(salesperson="Dylan Nava", sales_team="B2C Web - RC Team",
             sku="MISCELLANEOUS", qty=1, rate=200.0, item_total=200.0),
        dict(salesperson="Dylan Nava", sales_team="B2C Web - RC Team",
             sku="REALSKU", qty=1, rate=100.0, item_total=100.0),
    ])
    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)
    skus = {r["sku"] for r in result.rows}
    assert "MISCELLANEOUS" not in skus
    assert result.pool_commissionable == pytest.approx(100.0)
    assert result.pool_commission == pytest.approx(2.0)


def test_custom_rate_override(tmp_path):
    db_path = tmp_path / "rate.sqlite"
    _seed(db_path, [dict(
        salesperson="Dylan Nava", sales_team="B2C Web - RC Team",
        sku="WIDGET", qty=1, rate=100.0, item_total=100.0,
    )])
    result = build_b2c_commission(YEAR, MONTH, db_path=db_path, rate=0.03)
    assert result.rate == 0.03
    assert result.pool_commission == pytest.approx(3.0)


def test_write_b2c_workbook(tmp_path):
    import openpyxl

    db_path = tmp_path / "wb.sqlite"
    _seed(db_path, [
        dict(salesperson="Customer Service", sales_team="B2C Web - RC Team",
             sku="A", qty=1, rate=1000.0, item_total=1000.0),
        dict(salesperson="Dylan Nava", sales_team="B2C - RC Team",
             sku="B", qty=1, rate=500.0, item_total=500.0),
    ])
    result = build_b2c_commission(YEAR, MONTH, db_path=db_path)

    out = tmp_path / "b2c_report.xlsx"
    write_b2c_workbook(result, out, year=YEAR, month=MONTH)
    assert out.exists()

    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb.sheetnames == ["Summary", "B2C_Commission"]
    # Detail sheet has a header + 2 data rows + a TOTAL row.
    ws = wb["B2C_Commission"]
    assert ws.max_row >= 4
    # Pool total appears on the Summary sheet (1500 * 0.02 = 30.00).
    summary_text = [ws_cell.value for row in wb["Summary"].iter_rows() for ws_cell in row]
    assert 30.0 in summary_text
