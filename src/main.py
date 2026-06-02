from pathlib import Path
import argparse
import calendar
import sys
import pandas as pd


BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from src.commission.returns import commissionable_quantity  # shared returned-qty rule
from src.commission.b2b_reconciliation import build_reconciliation_frames, SOURCE_LABEL

INPUT_DIR = BASE_DIR / "data" / "input"
OUTPUT_DIR = BASE_DIR / "data" / "output"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def parse_args():
    parser = argparse.ArgumentParser(description="Generate commission audit report.")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--month", type=int, default=4)
    parser.add_argument(
        "--historical-replay",
        action="store_true",
        help="Use B2B workbook detail lines as strict historical source of truth.",
    )
    parser.add_argument(
        "--disable-summary-normalization",
        action="store_true",
        help="Disable payout normalization against Jennifer B2B Summary totals.",
    )
    parser.add_argument(
        "--data-source",
        choices=["sqlite", "excel"],
        default="sqlite",
        help="Operational Zoho data source: sqlite (default) or excel files in data/input.",
    )
    return parser.parse_args()


ARGS = parse_args()

YEAR = ARGS.year
MONTH = ARGS.month
HISTORICAL_REPLAY = ARGS.historical_replay
DISABLE_SUMMARY_NORMALIZATION = ARGS.disable_summary_normalization
USE_SUMMARY_NORMALIZATION = not DISABLE_SUMMARY_NORMALIZATION
DATA_SOURCE = ARGS.data_source
MONTH_NAME = calendar.month_name[MONTH]
MONTH_NAME_LOWER = MONTH_NAME.lower()

B2B_WORKBOOK = INPUT_DIR / f"{YEAR}-{MONTH}_Commission B2B.xlsx"
SALES_ORDERS_FILE = INPUT_DIR / f"Sales_Orders {MONTH_NAME} {YEAR}.xlsx"
INVOICES_FILE = INPUT_DIR / f"Invoices {MONTH_NAME} {YEAR}.xlsx"
SHIPMENTS_FILE = INPUT_DIR / f"Shipments {MONTH_NAME} {YEAR}.xlsx"

OUTPUT_FILE = OUTPUT_DIR / f"commission_audit_{MONTH_NAME_LOWER}_{YEAR}.xlsx"


SALESPERSON_SHEETS = {
    "Paul": "Paul Perlman",
    "Jose": "Jose Ayala",
    "Michael": "Michael Ayala",
    "Jim": "Jim Sutton",
    "Weston": "Weston Fields",
    "Brett": "Brett Bern",
    "Company Acct": "Company Account",
    "Leslie": "Leslie Neipert",
    "Carmen": "Carmen Daetz",
    "Garrett": "Garrett Lockhart",
}


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip().replace("\n", " ") for col in df.columns]
    return df


def find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(col).strip().lower(): col for col in df.columns}

    for candidate in candidates:
        key = candidate.strip().lower()
        if key in lookup:
            return lookup[key]

    return None


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_header_text(value) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = " ".join(text.split())
    return text


def money_to_number(value_or_series):
    def one_value(value):
        if pd.isna(value):
            return 0.0

        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()

        if text == "":
            return 0.0

        text = (
            text.replace("$", "")
            .replace(",", "")
            .replace("(", "-")
            .replace(")", "")
            .replace("%", "")
            .strip()
        )

        try:
            return float(text)
        except ValueError:
            return 0.0

    if isinstance(value_or_series, pd.Series):
        return value_or_series.apply(one_value)

    return one_value(value_or_series)


def safe_number(value) -> float:
    return money_to_number(value)


def parse_discount_rate(value):
    if pd.isna(value):
        return None

    raw = str(value).strip()
    text = raw.replace("%", "").replace(",", "")
    if text == "":
        return None

    try:
        parsed = float(text)
    except ValueError:
        return None

    # We only trust explicit percentages ("15%") or fractional values (0.15).
    # Numeric values > 1 in the export are often fixed-amount discounts, not rates.
    if "%" in raw:
        parsed = parsed / 100.0
    elif parsed > 1:
        return None

    if parsed < 0:
        parsed = 0.0

    return parsed


def normalize_salesperson_name(name: str) -> str:
    name = normalize_text(name)

    replacements = {
        "Company Acct": "Company Account",
        "B2B Company Account": "Company Account",
        "Company Account": "Company Account",
        "Paul": "Paul Perlman",
        "Jose": "Jose Ayala",
        "Michael": "Michael Ayala",
        "Jim": "Jim Sutton",
        "Weston": "Weston Fields",
        "Brett": "Brett Bern",
        "Leslie": "Leslie Neipert",
        "Carmen": "Carmen Daetz",
        "Garrett": "Garrett Lockhart",
    }

    return replacements.get(name, name)


def get_rate_type_for_salesperson(salesperson: str) -> str:
    salesperson = normalize_salesperson_name(salesperson)

    non_salaried = {
        "Brett Bern",
        "Leslie Neipert",
        "Carmen Daetz",
        "Garrett Lockhart",
    }

    if salesperson in non_salaried:
        return "non_salaried"

    return "salaried"


def make_match_key(sales_order_number, sku, invoice_number):
    so = normalize_text(sales_order_number)
    sku_value = normalize_text(sku)
    inv = normalize_text(invoice_number)
    return f"{so}||{sku_value}||{inv}"


def make_fallback_key(sales_order_number, sku, quantity, amount):
    so = normalize_text(sales_order_number)
    sku_value = normalize_text(sku)
    qty = round(safe_number(quantity), 4)
    amt = round(safe_number(amount), 2)
    return f"{so}||{sku_value}||{qty}||{amt}"


def read_excel_first_sheet(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(
            f"File not found: {path}\n"
            f"Expected file name: {path.name}\n"
            f"Put it in: {INPUT_DIR}"
        )

    df = pd.read_excel(path, sheet_name=0)
    return clean_columns(df)


def read_source_table(source: Path | pd.DataFrame) -> pd.DataFrame:
    if isinstance(source, pd.DataFrame):
        return clean_columns(source.copy())
    return read_excel_first_sheet(source)


def _source_label(source: Path | pd.DataFrame) -> str:
    if isinstance(source, pd.DataFrame):
        return "SQLite"
    return str(source)


def read_sheet_with_header_row(path: Path, sheet_name: str, header_row_zero_based: int) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(
            f"File not found: {path}\n"
            f"Expected file name: {path.name}\n"
            f"Put it in: {INPUT_DIR}"
        )

    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row_zero_based)
    df = clean_columns(df)
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    return df


def read_commission_table(path: Path) -> pd.DataFrame:
    table = pd.read_excel(path, sheet_name="Table", header=3)
    table = clean_columns(table)
    table = table.dropna(how="all")
    table = table.dropna(axis=1, how="all")

    discount_col = find_col(table, ["Discount Rate"])

    salaried_col = find_col(
        table,
        [
            "2025 Salaried Commission Rate",
            "Salaried Commission Rate",
        ],
    )

    non_salaried_col = find_col(
        table,
        [
            "2025 Non Salaried Commission Rate",
            "Non Salaried Commission Rate",
            "Non-Salaried Commission Rate",
        ],
    )

    if discount_col is None:
        raise ValueError(f"Could not find Discount Rate column in Table sheet. Columns: {list(table.columns)}")

    if salaried_col is None:
        for col in table.columns:
            col_text = str(col).lower()
            if "salaried" in col_text and "non" not in col_text:
                salaried_col = col
                break

    if non_salaried_col is None:
        for col in table.columns:
            col_text = str(col).lower()
            if "non" in col_text and "salaried" in col_text:
                non_salaried_col = col
                break

    if salaried_col is None:
        raise ValueError(f"Could not find Salaried Commission Rate column. Columns: {list(table.columns)}")

    result = pd.DataFrame()
    result["Discount Rate"] = pd.to_numeric(table[discount_col], errors="coerce")
    result["Salaried Commission Rate"] = pd.to_numeric(table[salaried_col], errors="coerce")

    if non_salaried_col:
        result["Non Salaried Commission Rate"] = pd.to_numeric(table[non_salaried_col], errors="coerce")
    else:
        result["Non Salaried Commission Rate"] = 0.0

    result = result.dropna(subset=["Discount Rate"])
    result = result.sort_values("Discount Rate").reset_index(drop=True)

    return result


def lookup_commission_rate(discount_rate: float, table: pd.DataFrame, rate_type: str = "salaried") -> float:
    if pd.isna(discount_rate):
        return 0.0

    rate_col = "Salaried Commission Rate"

    if rate_type == "non_salaried":
        rate_col = "Non Salaried Commission Rate"

    eligible = table[table["Discount Rate"] <= discount_rate]

    if eligible.empty:
        return float(table.iloc[0][rate_col] or 0)

    return float(eligible.iloc[-1][rate_col] or 0)


def build_product_lookup(b2b_path: Path) -> pd.DataFrame:
    r_lp = read_sheet_with_header_row(b2b_path, "R_LP", 1)

    sku_col = find_col(r_lp, ["SKU"])
    brand_col = find_col(r_lp, ["Brand"])
    desc_col = find_col(r_lp, ["Description"])
    account_col = find_col(r_lp, ["Account"])
    account_code_col = find_col(r_lp, ["Account Code"])
    rate_col = find_col(r_lp, ["Rate"])
    purchase_rate_col = find_col(r_lp, ["Purchase Rate"])

    if sku_col is None:
        raise ValueError(f"Missing SKU column in R_LP. Columns: {list(r_lp.columns)}")

    product = pd.DataFrame()
    product["SKU"] = r_lp[sku_col].apply(normalize_text)
    product["Brand"] = r_lp[brand_col].apply(normalize_text) if brand_col else ""
    product["Product Description"] = r_lp[desc_col].apply(normalize_text) if desc_col else ""
    product["Product Account"] = r_lp[account_col].apply(normalize_text) if account_col else ""
    product["Product Account Code"] = r_lp[account_code_col].apply(normalize_text) if account_code_col else ""
    product["MAP Price"] = money_to_number(r_lp[rate_col]) if rate_col else 0.0
    product["Purchase Rate"] = money_to_number(r_lp[purchase_rate_col]) if purchase_rate_col else 0.0

    product = product[product["SKU"] != ""]
    product = product.drop_duplicates(subset=["SKU"], keep="last")

    return product


def build_sales_orders_base(source: Path | pd.DataFrame) -> pd.DataFrame:
    so = read_source_table(source)

    needed = {
        "Salesperson": ["Sales person", "Salesperson", "Sales Person"],
        "Sales Order Number": ["SalesOrder Number", "Sales Order Number", "Sales Order#"],
        "Order Date": ["Order Date"],
        "Reference Number": ["Reference#"],
        "Sales Order ID": ["SalesOrder ID", "Sales Order ID"],
        "SO Status": ["Status"],
        "Customer ID": ["Customer ID"],
        "Customer Name": ["Customer Name"],
        "SKU": ["SKU"],
        "Item Name": ["Item Name"],
        "Item Description": ["Item Desc", "Item Description"],
        "Quantity": ["QuantityOrdered", "Quantity Ordered", "Quantity"],
        "Quantity Invoiced": ["QuantityInvoiced", "Quantity Invoiced"],
        "Quantity Shipped": ["QuantityShipped", "Quantity Shipped"],
        "Quantity Returned": ["QuantityReturned", "Quantity Returned"],
        "Quantity Packed": ["QuantityPacked", "Quantity Packed"],
        "Item Price": ["Item Price"],
        "Discount": ["Discount"],
        "Discount Amount": ["Discount Amount"],
        "Revenue": ["Item Total"],
        "SO SubTotal": ["SubTotal", "Sub Total"],
        "SO Total": ["Total"],
        "SO Shipping Charge": ["Shipping Charge"],
        "SO Account": ["Account"],
        "SO Account Code": ["Account Code"],
        "Payment Terms": ["Payment Terms Label", "Payment Terms"],
        "Delivery Method": ["Delivery Method"],
        "Coupon": ["CF.Coupon(s)", "Coupon(s)"],
        "Sales Notes": ["CF.Sales Notes", "Sales Notes"],
        "Ticket": ["CF.Ticket#", "Ticket#"],
    }

    cols = {name: find_col(so, candidates) for name, candidates in needed.items()}

    required = [
        "Salesperson",
        "Sales Order Number",
        "Order Date",
        "Sales Order ID",
        "SO Status",
        "Customer Name",
        "SKU",
        "Quantity",
        "Revenue",
    ]

    missing = [name for name in required if cols[name] is None]

    if missing:
        raise ValueError(
            f"Missing required Sales Order columns: {missing}\n"
            f"Source: {_source_label(source)}\n"
            f"Columns found: {list(so.columns)}"
        )

    base = pd.DataFrame()

    for out_col, source_col in cols.items():
        if source_col is None:
            base[out_col] = ""
        else:
            base[out_col] = so[source_col]

    text_cols = [
        "Salesperson",
        "Sales Order Number",
        "Reference Number",
        "Sales Order ID",
        "SO Status",
        "Customer ID",
        "Customer Name",
        "SKU",
        "Item Name",
        "Item Description",
        "SO Account",
        "SO Account Code",
        "Payment Terms",
        "Delivery Method",
        "Coupon",
        "Sales Notes",
        "Ticket",
    ]

    for col in text_cols:
        if col in base.columns:
            base[col] = base[col].apply(normalize_text)

    numeric_cols = [
        "Quantity",
        "Quantity Invoiced",
        "Quantity Packed",
        "Item Price",
        "Discount Amount",
        "Revenue",
        "SO SubTotal",
        "SO Total",
        "SO Shipping Charge",
    ]

    for col in numeric_cols:
        if col in base.columns:
            base[col] = money_to_number(base[col])

    if "Discount" in base.columns:
        base["Discount Rate SO"] = base["Discount"].apply(parse_discount_rate)
    else:
        base["Discount Rate SO"] = None

    base = base[base["Sales Order Number"] != ""].copy()
    base = base[base["SKU"] != ""].copy()

    return base


def build_invoice_lookup(source: Path | pd.DataFrame) -> pd.DataFrame:
    inv_raw = read_source_table(source)

    needed = {
        "Invoice Date": ["Invoice Date"],
        "Invoice ID": ["Invoice ID"],
        "Invoice Number": ["Invoice Number"],
        "Invoice Status": ["Invoice Status"],
        "Due Date": ["Due Date"],
        "Balance": ["Balance"],
        "Last Payment Date": ["Last Payment Date"],
        "Invoice Payment Terms": ["Payment Terms Label", "Payment Terms"],
        "Sales Order Number": ["Sales Order Number"],
        "Estimate Number": ["Estimate Number"],
        "SKU": ["SKU"],
        "Invoice Quantity": ["Quantity"],
        "Invoice Item Total": ["Item Total"],
        "Invoice Item Price": ["Item Price"],
        "Invoice Account": ["Account"],
        "Invoice Account Code": ["Account Code"],
        "Invoice Salesperson": ["Sales person", "Salesperson"],
        "Sales Team": ["CF.Sales Team", "Sales Team"],
    }

    cols = {name: find_col(inv_raw, candidates) for name, candidates in needed.items()}

    required = [
        "Invoice Date",
        "Invoice ID",
        "Invoice Number",
        "Invoice Status",
        "Balance",
        "Sales Order Number",
        "SKU",
        "Invoice Item Total",
    ]

    missing = [name for name in required if cols[name] is None]

    if missing:
        raise ValueError(
            f"Missing required Invoice columns: {missing}\n"
            f"Source: {_source_label(source)}\n"
            f"Columns found: {list(inv_raw.columns)}"
        )

    inv = pd.DataFrame()

    for out_col, source_col in cols.items():
        if source_col is None:
            inv[out_col] = ""
        else:
            inv[out_col] = inv_raw[source_col]

    text_cols = [
        "Invoice ID",
        "Invoice Number",
        "Invoice Status",
        "Invoice Payment Terms",
        "Sales Order Number",
        "Estimate Number",
        "SKU",
        "Invoice Account",
        "Invoice Account Code",
        "Invoice Salesperson",
        "Sales Team",
    ]

    for col in text_cols:
        if col in inv.columns:
            inv[col] = inv[col].apply(normalize_text)

    numeric_cols = [
        "Balance",
        "Invoice Quantity",
        "Invoice Item Total",
        "Invoice Item Price",
    ]

    for col in numeric_cols:
        if col in inv.columns:
            inv[col] = money_to_number(inv[col])

    inv = inv[inv["Sales Order Number"] != ""].copy()
    inv = inv[inv["SKU"] != ""].copy()

    inv["Invoice Status Lower"] = inv["Invoice Status"].str.lower()
    inv = inv[~inv["Invoice Status Lower"].str.contains("void", na=False)].copy()

    inv["_Invoice Date Sort"] = pd.to_datetime(inv["Invoice Date"], errors="coerce")
    inv = inv.sort_values(["Sales Order Number", "SKU", "_Invoice Date Sort"])

    inv_lookup = inv.groupby(["Sales Order Number", "SKU"], as_index=False).last()
    inv_lookup = inv_lookup.drop(columns=["_Invoice Date Sort", "Invoice Status Lower"], errors="ignore")

    return inv_lookup


def build_shipment_lookup(source: Path | pd.DataFrame) -> pd.DataFrame:
    sh_raw = read_source_table(source)

    needed = {
        "Shipment ID": ["Shipment ID"],
        "Packing Number": ["Packing Number"],
        "Shipment Date": ["Shipment Date"],
        "Shipment Number": ["Shipment Number"],
        "Shipment Status": ["Status", "Shipment Status"],
        "Shipment Type": ["Shipment Type"],
        "Carrier": ["Carrier Name", "Carrier"],
        "Tracking Number": ["Tracking Number"],
        "Shipping Charge": ["Shipping Charge"],
        "Sales Order ID": ["Sales Order ID"],
        "Sales Order Number": ["SO Number", "Sales Order Number"],
        "Shipment Customer Name": ["Customer Name"],
    }

    cols = {name: find_col(sh_raw, candidates) for name, candidates in needed.items()}

    required = [
        "Shipment Date",
        "Shipment Status",
        "Sales Order ID",
        "Sales Order Number",
    ]

    missing = [name for name in required if cols[name] is None]

    if missing:
        raise ValueError(
            f"Missing required Shipment columns: {missing}\n"
            f"Source: {_source_label(source)}\n"
            f"Columns found: {list(sh_raw.columns)}"
        )

    sh = pd.DataFrame()

    for out_col, source_col in cols.items():
        if source_col is None:
            sh[out_col] = ""
        else:
            sh[out_col] = sh_raw[source_col]

    text_cols = [
        "Shipment ID",
        "Packing Number",
        "Shipment Number",
        "Shipment Status",
        "Shipment Type",
        "Carrier",
        "Tracking Number",
        "Sales Order ID",
        "Sales Order Number",
        "Shipment Customer Name",
    ]

    for col in text_cols:
        if col in sh.columns:
            sh[col] = sh[col].apply(normalize_text)

    sh["Shipping Charge"] = money_to_number(sh["Shipping Charge"])

    sh = sh[(sh["Sales Order Number"] != "") | (sh["Sales Order ID"] != "")].copy()

    sh["_Shipment Date Sort"] = pd.to_datetime(sh["Shipment Date"], errors="coerce")
    sh = sh.sort_values(["Sales Order ID", "Sales Order Number", "_Shipment Date Sort"])

    grouped_rows = []

    for _, group in sh.groupby(["Sales Order ID", "Sales Order Number"], dropna=False):
        latest = group.iloc[-1].copy()
        latest["Shipping Charge"] = group["Shipping Charge"].sum()

        latest["All Shipment Numbers"] = ", ".join(
            [
                x
                for x in group["Shipment Number"].astype(str).tolist()
                if x and x != "nan"
            ]
        )

        latest["All Tracking Numbers"] = ", ".join(
            [
                x
                for x in group["Tracking Number"].astype(str).tolist()
                if x and x != "nan"
            ]
        )

        grouped_rows.append(latest)

    if grouped_rows:
        sh_lookup = pd.DataFrame(grouped_rows)
    else:
        sh_lookup = sh

    sh_lookup = sh_lookup.drop(columns=["_Shipment Date Sort"], errors="ignore")

    return sh_lookup


def read_jennifer_b2b_summary(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name="B2B Summary", header=None)

    rows = []

    for idx, row in raw.iterrows():
        label_col_d = normalize_text(row.get(3, ""))
        name_col_e = normalize_text(row.get(4, ""))

        salesperson = ""

        if name_col_e != "":
            salesperson = name_col_e
        elif "company account" in label_col_d.lower():
            salesperson = "Company Account"

        if salesperson == "":
            continue

        skip_labels = [
            "B2B Sales Rep",
            "Salary Commission",
            "Non-Salary Commission",
            "B2B Executive Account",
            "Total B2B",
        ]

        if salesperson in skip_labels:
            continue

        current_commission_payable = safe_number(row.get(12, 0))
        prior_commission_payable = safe_number(row.get(18, 0))
        total_to_pay = safe_number(row.get(20, 0))
        current_commissionable = safe_number(row.get(10, 0))
        prior_commissionable = safe_number(row.get(16, 0))
        revenue = safe_number(row.get(5, 0))
        product_revenue = safe_number(row.get(6, 0))

        if (
            revenue == 0
            and product_revenue == 0
            and current_commissionable == 0
            and prior_commissionable == 0
            and current_commission_payable == 0
            and prior_commission_payable == 0
            and total_to_pay == 0
        ):
            continue

        rows.append(
            {
                "Salesperson": normalize_salesperson_name(salesperson),
                "Jennifer Revenue": revenue,
                "Jennifer Product Revenue": product_revenue,
                "Jennifer Current Commissionable Amount": current_commissionable,
                "Jennifer Current Commission Payable": current_commission_payable,
                "Jennifer Prior Commissionable Amount": prior_commissionable,
                "Jennifer Prior Commission Payable": prior_commission_payable,
                "Jennifer Total To Pay": total_to_pay,
                "Jennifer Source Row": idx + 1,
            }
        )

    result = pd.DataFrame(rows)

    if result.empty:
        return result

    result = (
        result.groupby("Salesperson", as_index=False)
        .agg(
            {
                "Jennifer Revenue": "sum",
                "Jennifer Product Revenue": "sum",
                "Jennifer Current Commissionable Amount": "sum",
                "Jennifer Current Commission Payable": "sum",
                "Jennifer Prior Commissionable Amount": "sum",
                "Jennifer Prior Commission Payable": "sum",
                "Jennifer Total To Pay": "sum",
                "Jennifer Source Row": "first",
            }
        )
    )

    return result


def find_jennifer_header_rows(raw: pd.DataFrame) -> list[int]:
    header_rows = []

    for idx, row in raw.iterrows():
        values = [normalize_header_text(v) for v in row.tolist()]
        row_text = " | ".join(values)

        has_so = "salesorder number" in row_text or "sales order number" in row_text
        has_sku = "sku" in values or "sku" in row_text
        has_item_total = "item total" in row_text
        has_order_date = "order date" in row_text

        if has_so and has_sku and has_item_total and has_order_date:
            header_rows.append(idx)

    return header_rows


def build_jennifer_header_map(raw: pd.DataFrame, header_row: int) -> dict[str, int]:
    header_values = [normalize_header_text(v) for v in raw.iloc[header_row].tolist()]
    header_map = {}

    for idx, value in enumerate(header_values):
        if value != "":
            header_map[value] = idx

    return header_map


def get_jennifer_col(header_map: dict[str, int], candidates: list[str]) -> int | None:
    normalized_candidates = [normalize_header_text(candidate) for candidate in candidates]

    for candidate in normalized_candidates:
        if candidate in header_map:
            return header_map[candidate]

    for header, idx in header_map.items():
        for candidate in normalized_candidates:
            if candidate in header:
                return idx

    return None


def detect_jennifer_section(raw: pd.DataFrame, row_index: int) -> str:
    for i in range(row_index, max(-1, row_index - 60), -1):
        values = [normalize_header_text(v) for v in raw.iloc[i].tolist()]
        line = " | ".join(values)

        if "sales orders - prior periods" in line or "orders of previous periods" in line:
            return "Prior Period"

        if "orders commissionable" in line or "sales orders invoiced" in line:
            return "Current Period"

        if "shipping charge" in line:
            return "Shipping Charge"

        if "other charges" in line or "other income" in line:
            return "Other Charges"

    return "Current Period"


def parse_jennifer_salesperson_sheet(path: Path, sheet_name: str, salesperson: str) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    header_rows = find_jennifer_header_rows(raw)

    rows = []

    for header_row in header_rows:
        header_map = build_jennifer_header_map(raw, header_row)
        section = detect_jennifer_section(raw, header_row)

        order_date_col = get_jennifer_col(header_map, ["Order Date"])
        so_col = get_jennifer_col(header_map, ["SalesOrder Number", "Sales Order Number"])
        invoice_date_col = get_jennifer_col(header_map, ["Invoice Date"])
        invoice_number_col = get_jennifer_col(header_map, ["Invoice Number"])
        invoice_status_col = get_jennifer_col(header_map, ["Invoice Status"])
        customer_col = get_jennifer_col(header_map, ["Customer Name"])
        estimate_col = get_jennifer_col(header_map, ["Estimate Number"])
        sku_col = get_jennifer_col(header_map, ["SKU"])
        qty_col = get_jennifer_col(header_map, ["Quantity Ordered", "Quantity"])
        item_total_col = get_jennifer_col(header_map, ["Item Total"])
        account_col = get_jennifer_col(header_map, ["Account"])
        account_code_col = get_jennifer_col(header_map, ["Account Code"])
        payment_terms_col = get_jennifer_col(header_map, ["Payment Terms Label", "Payment Terms"])
        delivery_method_col = get_jennifer_col(header_map, ["Delivery Method"])
        shipment_date_col = get_jennifer_col(header_map, ["Shipment Date"])
        shipment_status_col = get_jennifer_col(header_map, ["Shipment Status"])
        ar_status_col = get_jennifer_col(header_map, ["AR Status"])
        payment_date_col = get_jennifer_col(header_map, ["Payment Date", "Last Payment Date"])

        discount_col = get_jennifer_col(
            header_map,
            ["Discount Rate", "Discount %", "Discount"]
        )

        commissionable_col = get_jennifer_col(
            header_map,
            ["Commissionable Amount"]
        )

        commission_rate_col = get_jennifer_col(
            header_map,
            ["Commission Rate", "Commission %"]
        )

        commission_payable_col = get_jennifer_col(
            header_map,
            ["Commission Amount", "Commission Payable"]
        )

        required = [so_col, sku_col, item_total_col]

        if any(col is None for col in required):
            continue

        for row_idx in range(header_row + 1, len(raw)):
            row = raw.iloc[row_idx]

            so_number = normalize_text(row.iloc[so_col])
            sku = normalize_text(row.iloc[sku_col])
            item_total = safe_number(row.iloc[item_total_col])

            row_values = [normalize_text(v) for v in row.tolist()]
            row_text = " ".join(row_values).lower()

            if (
                so_number == ""
                and sku == ""
                and item_total == 0
                and row_idx > header_row + 3
            ):
                lookahead_empty = True

                for j in range(row_idx, min(row_idx + 3, len(raw))):
                    future_values = [normalize_text(v) for v in raw.iloc[j].tolist()]
                    if any(v != "" for v in future_values):
                        lookahead_empty = False
                        break

                if lookahead_empty:
                    break

            if "shipping charge" in row_text or "other charges" in row_text:
                break

            if so_number == "" or sku == "":
                continue

            invoice_number = normalize_text(row.iloc[invoice_number_col]) if invoice_number_col is not None else ""
            quantity = safe_number(row.iloc[qty_col]) if qty_col is not None else 0.0

            rows.append(
                {
                    "Jennifer Sheet": sheet_name,
                    "Jennifer Salesperson": salesperson,
                    "Jennifer Section": section,
                    "Jennifer Order Date": row.iloc[order_date_col] if order_date_col is not None else "",
                    "Sales Order Number": so_number,
                    "Jennifer Invoice Date": row.iloc[invoice_date_col] if invoice_date_col is not None else "",
                    "Invoice Number": invoice_number,
                    "Jennifer Invoice Status": normalize_text(row.iloc[invoice_status_col]) if invoice_status_col is not None else "",
                    "Jennifer Customer Name": normalize_text(row.iloc[customer_col]) if customer_col is not None else "",
                    "Jennifer Estimate Number": normalize_text(row.iloc[estimate_col]) if estimate_col is not None else "",
                    "SKU": sku,
                    "Jennifer Quantity": quantity,
                    "Jennifer Item Total": item_total,
                    "Jennifer Account": normalize_text(row.iloc[account_col]) if account_col is not None else "",
                    "Jennifer Account Code": normalize_text(row.iloc[account_code_col]) if account_code_col is not None else "",
                    "Jennifer Payment Terms": normalize_text(row.iloc[payment_terms_col]) if payment_terms_col is not None else "",
                    "Jennifer Delivery Method": normalize_text(row.iloc[delivery_method_col]) if delivery_method_col is not None else "",
                    "Jennifer Shipment Date": row.iloc[shipment_date_col] if shipment_date_col is not None else "",
                    "Jennifer Shipment Status": normalize_text(row.iloc[shipment_status_col]) if shipment_status_col is not None else "",
                    "Jennifer AR Status": normalize_text(row.iloc[ar_status_col]) if ar_status_col is not None else "",
                    "Jennifer Payment Date": row.iloc[payment_date_col] if payment_date_col is not None else "",
                    "Jennifer Discount": safe_number(row.iloc[discount_col]) if discount_col is not None else 0.0,
                    "Jennifer Commissionable Amount": safe_number(row.iloc[commissionable_col]) if commissionable_col is not None else 0.0,
                    "Jennifer Commission Rate": safe_number(row.iloc[commission_rate_col]) if commission_rate_col is not None else 0.0,
                    "Jennifer Commission Payable": safe_number(row.iloc[commission_payable_col]) if commission_payable_col is not None else 0.0,
                    "Jennifer Source Header Row": header_row + 1,
                    "Jennifer Source Data Row": row_idx + 1,
                    "Jennifer Match Key": make_match_key(so_number, sku, invoice_number),
                    "Jennifer Fallback Key": make_fallback_key(so_number, sku, quantity, item_total),
                }
            )

    return pd.DataFrame(rows)


def read_jennifer_lines(path: Path) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    available_sheets = set(xls.sheet_names)

    parsed_frames = []

    for sheet_name, salesperson in SALESPERSON_SHEETS.items():
        if sheet_name not in available_sheets:
            continue

        df = parse_jennifer_salesperson_sheet(path, sheet_name, salesperson)

        if not df.empty:
            parsed_frames.append(df)

    if parsed_frames:
        return pd.concat(parsed_frames, ignore_index=True)

    return pd.DataFrame()


def dedupe_jennifer_parsed_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove duplicate parses of the same workbook row when a sheet has
    multiple header blocks (common in Other Charges / Shipping Charge).
    """
    if df.empty:
        return df

    out = df.copy()
    out["_dedupe_rank"] = (
        out["Jennifer Commission Payable"].fillna(0).astype(float) * 1_000_000
        + out["Jennifer Commissionable Amount"].fillna(0).astype(float)
    )

    subset = [
        col
        for col in ["Jennifer Sheet", "Jennifer Source Data Row", "Jennifer Match Key"]
        if col in out.columns
    ]
    if not subset:
        return out.drop(columns=["_dedupe_rank"], errors="ignore")

    out = out.sort_values("_dedupe_rank", ascending=False)
    out = out.drop_duplicates(subset=subset, keep="first")
    return out.drop(columns=["_dedupe_rank"], errors="ignore")


def aggregate_jennifer_lines_by_key(df: pd.DataFrame, key_col: str) -> pd.DataFrame:
    """
    Sum commission values for rows that share the same match/fallback key
    after removing duplicate parses of the same source row.
    """
    if df.empty or key_col not in df.columns:
        return df.copy()

    base = dedupe_jennifer_parsed_rows(df)
    grouped = base.groupby(key_col, as_index=False).agg(
        **{
            "Jennifer Commission Payable": ("Jennifer Commission Payable", "sum"),
            "Jennifer Commissionable Amount": ("Jennifer Commissionable Amount", "sum"),
            "Jennifer Item Total": ("Jennifer Item Total", "sum"),
            "Jennifer Commission Rate": ("Jennifer Commission Rate", "first"),
            "Jennifer Salesperson": ("Jennifer Salesperson", "first"),
            "Jennifer Section": ("Jennifer Section", "first"),
            "Sales Order Number": ("Sales Order Number", "first"),
            "SKU": ("SKU", "first"),
            "Invoice Number": ("Invoice Number", "first"),
            "Jennifer Quantity": ("Jennifer Quantity", "first"),
            "Jennifer Order Date": ("Jennifer Order Date", "first"),
            "Jennifer Invoice Date": ("Jennifer Invoice Date", "first"),
            "Jennifer Match Key": ("Jennifer Match Key", "first"),
            "Jennifer Fallback Key": ("Jennifer Fallback Key", "first"),
        }
    )

    base_mask = grouped["Jennifer Commissionable Amount"].fillna(0) > 0
    grouped.loc[base_mask, "Jennifer Commission Rate"] = (
        grouped.loc[base_mask, "Jennifer Commission Payable"]
        / grouped.loc[base_mask, "Jennifer Commissionable Amount"]
    )
    return grouped


def jennifer_lookup_dict(df: pd.DataFrame, key_col: str, value_col: str) -> dict:
    if df.empty or key_col not in df.columns or value_col not in df.columns:
        return {}
    aggregated = aggregate_jennifer_lines_by_key(df, key_col)
    return aggregated.set_index(key_col)[value_col].to_dict()


def compute_ar_status(invoice_status: str, balance: float, invoice_number: str) -> str:
    invoice_status_lower = normalize_text(invoice_status).lower()
    invoice_number = normalize_text(invoice_number)

    if invoice_number == "":
        return "NO INVOICE"

    if balance == 0:
        return "PAID"

    if "closed" in invoice_status_lower or "paid" in invoice_status_lower:
        return "PAID"

    if balance > 0:
        return "UNPAID"

    return "REVIEW"


def current_or_prior(order_date, commission_year: int = YEAR, commission_month: int = MONTH) -> str:
    dt = pd.to_datetime(order_date, errors="coerce")

    if pd.isna(dt):
        return "Review"

    period_start = pd.Timestamp(year=commission_year, month=commission_month, day=1)

    if dt.year == commission_year and dt.month == commission_month:
        return "Current Period"

    if dt < period_start:
        return "Prior Period"

    return "Future / Review"


def load_operational_dataframes() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if DATA_SOURCE == "sqlite":
        from src.commission.sqlite_data_source import load_period_dataframes

        period = load_period_dataframes(YEAR, MONTH)
        return period.sales_orders, period.invoices, period.shipments

    if not SALES_ORDERS_FILE.exists() or not INVOICES_FILE.exists():
        raise FileNotFoundError(
            "Excel data source selected but period files are missing in data/input. "
            "Run Zoho sync or use --data-source sqlite."
        )
    return (
        read_source_table(SALES_ORDERS_FILE),
        read_source_table(INVOICES_FILE),
        read_source_table(SHIPMENTS_FILE) if SHIPMENTS_FILE.exists() else pd.DataFrame(),
    )


def build_commission_audit():
    print("Reading source files...")
    print(f"Year: {YEAR}")
    print(f"Month: {MONTH_NAME}")
    print(f"Data source: {DATA_SOURCE}")
    print(f"B2B workbook: {B2B_WORKBOOK.name}")
    if DATA_SOURCE == "excel":
        print(f"Sales Orders file: {SALES_ORDERS_FILE.name}")
        print(f"Invoices file: {INVOICES_FILE.name}")
        print(f"Shipments file: {SHIPMENTS_FILE.name}")
    print(f"Historical replay: {HISTORICAL_REPLAY}")
    print(f"Summary normalization: {USE_SUMMARY_NORMALIZATION}")

    so_raw, inv_raw, sh_raw = load_operational_dataframes()
    sales_orders = build_sales_orders_base(so_raw)
    invoices = build_invoice_lookup(inv_raw)
    shipments = build_shipment_lookup(sh_raw) if not sh_raw.empty else build_shipment_lookup(
        pd.DataFrame(
            columns=[
                "Shipment Date",
                "Shipment Number",
                "Status",
                "Carrier Name",
                "Tracking Number",
                "Shipping Charge",
                "Sales Order Number",
                "Sales Order ID",
                "Shipment ID",
                "Customer Name",
            ]
        )
    )
    products = build_product_lookup(B2B_WORKBOOK)
    commission_table = read_commission_table(B2B_WORKBOOK)
    jennifer_summary = read_jennifer_b2b_summary(B2B_WORKBOOK)
    jennifer_lines = read_jennifer_lines(B2B_WORKBOOK)

    valid_salespeople = set(jennifer_summary["Salesperson"].dropna().astype(str).tolist())

    print(f"Sales Orders lines: {len(sales_orders)}")
    print(f"Invoice lookup lines: {len(invoices)}")
    print(f"Shipment lookup rows: {len(shipments)}")
    print(f"Products: {len(products)}")
    print(f"Commission tiers: {len(commission_table)}")
    print(f"Jennifer summary rows: {len(jennifer_summary)}")
    print(f"Jennifer lines: {len(jennifer_lines)}")

    audit = sales_orders.copy()

    audit["Salesperson Normalized"] = audit["Salesperson"].apply(normalize_salesperson_name)
    audit["Valid Jennifer Salesperson"] = audit["Salesperson Normalized"].isin(valid_salespeople)

    audit = audit.merge(
        invoices,
        how="left",
        on=["Sales Order Number", "SKU"],
    )

    audit = audit.merge(
        shipments,
        how="left",
        on=["Sales Order ID", "Sales Order Number"],
    )

    audit = audit.merge(
        products,
        how="left",
        on="SKU",
    )

    audit["MAP Price"] = audit["MAP Price"].fillna(0)
    audit["MAP Total"] = audit["MAP Price"] * audit["Quantity"]

    audit["Discount Rate MAP"] = 0.0
    map_mask = audit["MAP Total"] > 0
    audit.loc[map_mask, "Discount Rate MAP"] = 1 - (
        audit.loc[map_mask, "Revenue"] / audit.loc[map_mask, "MAP Total"]
    )
    audit["Discount Rate MAP"] = audit["Discount Rate MAP"].clip(lower=0)

    audit["Discount Rate"] = audit["Discount Rate SO"]
    discount_missing = audit["Discount Rate"].isna()
    audit.loc[discount_missing, "Discount Rate"] = audit.loc[discount_missing, "Discount Rate MAP"]
    audit["Discount Rate"] = audit["Discount Rate"].fillna(0).clip(lower=0)
    audit["Discount Source"] = "Sales Order Discount %"
    audit.loc[discount_missing, "Discount Source"] = "MAP Ratio Fallback"

    audit["Rate Type"] = audit["Salesperson Normalized"].apply(get_rate_type_for_salesperson)

    audit["Commission Rate"] = audit.apply(
        lambda row: lookup_commission_rate(
            row["Discount Rate"],
            commission_table,
            rate_type=row["Rate Type"],
        ),
        axis=1,
    )

    audit["AR Status"] = audit.apply(
        lambda row: compute_ar_status(
            row.get("Invoice Status", ""),
            money_to_number(row.get("Balance", 0)),
            row.get("Invoice Number", ""),
        ),
        axis=1,
    )

    audit["Shipment Check"] = "NOT SHIPPED / MISSING"

    shipment_status_lower = audit["Shipment Status"].fillna("").astype(str).str.lower()

    audit.loc[
        shipment_status_lower.str.contains("shipped"),
        "Shipment Check",
    ] = "SHIPPED"

    audit["Current/Prior Period"] = audit["Order Date"].apply(current_or_prior)

    audit["Shipping Cost"] = 0.0
    audit["Shipping Net"] = audit["Shipping Charge"].fillna(0) - audit["Shipping Cost"]

    audit["Include Shipping in Commission"] = "Review"

    # Returned-quantity rule (shared helper with the B2B payable engine):
    # commission is paid only on quantity kept (invoiced - returned).
    qty_inv = pd.to_numeric(audit.get("Quantity Invoiced", 0), errors="coerce").fillna(0)
    qty_ret = pd.to_numeric(audit.get("Quantity Returned", 0), errors="coerce").fillna(0)
    qty_shp = pd.to_numeric(audit.get("Quantity Shipped", 0), errors="coerce").fillna(0)
    qty_ord = pd.to_numeric(audit.get("Quantity", 0), errors="coerce").fillna(0)
    _ret = [
        commissionable_quantity(i, r, s, o)
        for i, r, s, o in zip(qty_inv, qty_ret, qty_shp, qty_ord)
    ]
    audit["Qty Invoiced"] = qty_inv
    audit["Qty Returned"] = qty_ret
    audit["Qty Commissionable"] = [t[0] for t in _ret]
    audit["Return Factor"] = [t[1] for t in _ret]
    audit["Return Status"] = [t[2] for t in _ret]

    audit["Commissionable Amount"] = audit["Revenue"] * audit["Return Factor"]
    audit["Commission Amount"] = audit["Commissionable Amount"] * audit["Commission Rate"]

    other_charge_mask = (
        audit["SKU"].fillna("").astype(str).str.lower().str.contains("credit card processing fee")
        | audit["Item Name"].fillna("").astype(str).str.lower().str.contains("credit card processing fee")
    )
    audit.loc[other_charge_mask, "Commissionable Amount"] = 0.0
    audit.loc[other_charge_mask, "Commission Rate"] = 0.0
    audit.loc[other_charge_mask, "Commission Amount"] = 0.0

    audit["Our Match Key"] = audit.apply(
        lambda row: make_match_key(
            row.get("Sales Order Number", ""),
            row.get("SKU", ""),
            row.get("Invoice Number", ""),
        ),
        axis=1,
    )

    audit["Our Fallback Key"] = audit.apply(
        lambda row: make_fallback_key(
            row.get("Sales Order Number", ""),
            row.get("SKU", ""),
            row.get("Quantity", 0),
            row.get("Revenue", 0),
        ),
        axis=1,
    )

    workbook_has_detail = not jennifer_lines.empty
    if workbook_has_detail:
        j_active = jennifer_lines[
            jennifer_lines["Jennifer Section"].isin(["Current Period", "Prior Period"])
        ].copy()
        j_other = jennifer_lines[
            jennifer_lines["Jennifer Section"].isin(["Other Charges"])
        ].copy()
        j_shipping = jennifer_lines[
            jennifer_lines["Jennifer Section"].isin(["Shipping Charge"])
        ].copy()
        j_other_deduped = dedupe_jennifer_parsed_rows(j_other)
        j_shipping_deduped = dedupe_jennifer_parsed_rows(j_shipping)
        j_other_by_match = aggregate_jennifer_lines_by_key(j_other, "Jennifer Match Key")
        j_other_by_fallback = aggregate_jennifer_lines_by_key(j_other, "Jennifer Fallback Key")
        j_shipping_by_match = aggregate_jennifer_lines_by_key(j_shipping, "Jennifer Match Key")
        j_shipping_by_fallback = aggregate_jennifer_lines_by_key(j_shipping, "Jennifer Fallback Key")

        j_exact_keys = set(j_active["Jennifer Match Key"].astype(str).tolist())
        j_fallback_keys = set(j_active["Jennifer Fallback Key"].astype(str).tolist())
        j_other_exact_keys = set(j_other_deduped["Jennifer Match Key"].astype(str).tolist())
        j_other_fallback_keys = set(j_other_deduped["Jennifer Fallback Key"].astype(str).tolist())
        j_shipping_exact_keys = set(j_shipping_deduped["Jennifer Match Key"].astype(str).tolist())
        j_shipping_fallback_keys = set(j_shipping_deduped["Jennifer Fallback Key"].astype(str).tolist())
        j_workbook_exact_keys = j_exact_keys | j_other_exact_keys | j_shipping_exact_keys
        j_exact_sp = (
            j_active.drop_duplicates(subset=["Jennifer Match Key"])[
                ["Jennifer Match Key", "Jennifer Salesperson"]
            ]
            .set_index("Jennifer Match Key")["Jennifer Salesperson"]
            .to_dict()
        )
        j_exact_comm_amount = (
            j_active.drop_duplicates(subset=["Jennifer Match Key"])[
                ["Jennifer Match Key", "Jennifer Commission Payable"]
            ]
            .set_index("Jennifer Match Key")["Jennifer Commission Payable"]
            .to_dict()
        )
        j_exact_comm_rate = (
            j_active.drop_duplicates(subset=["Jennifer Match Key"])[
                ["Jennifer Match Key", "Jennifer Commission Rate"]
            ]
            .set_index("Jennifer Match Key")["Jennifer Commission Rate"]
            .to_dict()
        )
        j_exact_comm_base = (
            j_active.drop_duplicates(subset=["Jennifer Match Key"])[
                ["Jennifer Match Key", "Jennifer Commissionable Amount"]
            ]
            .set_index("Jennifer Match Key")["Jennifer Commissionable Amount"]
            .to_dict()
        )
        j_fallback_sp = (
            j_active.drop_duplicates(subset=["Jennifer Fallback Key"])[
                ["Jennifer Fallback Key", "Jennifer Salesperson"]
            ]
            .set_index("Jennifer Fallback Key")["Jennifer Salesperson"]
            .to_dict()
        )
        j_other_exact_sp = (
            j_other_by_match.set_index("Jennifer Match Key")["Jennifer Salesperson"].to_dict()
            if not j_other_by_match.empty
            else {}
        )
        j_other_exact_comm_amount = jennifer_lookup_dict(j_other, "Jennifer Match Key", "Jennifer Commission Payable")
        j_other_exact_comm_rate = jennifer_lookup_dict(j_other, "Jennifer Match Key", "Jennifer Commission Rate")
        j_other_exact_comm_base = jennifer_lookup_dict(
            j_other, "Jennifer Match Key", "Jennifer Commissionable Amount"
        )
        j_other_fallback_comm_amount = jennifer_lookup_dict(
            j_other, "Jennifer Fallback Key", "Jennifer Commission Payable"
        )
        j_other_fallback_comm_rate = jennifer_lookup_dict(
            j_other, "Jennifer Fallback Key", "Jennifer Commission Rate"
        )
        j_other_fallback_comm_base = jennifer_lookup_dict(
            j_other, "Jennifer Fallback Key", "Jennifer Commissionable Amount"
        )
        j_shipping_exact_sp = (
            j_shipping_by_match.set_index("Jennifer Match Key")["Jennifer Salesperson"].to_dict()
            if not j_shipping_by_match.empty
            else {}
        )
        j_shipping_exact_comm_amount = jennifer_lookup_dict(
            j_shipping, "Jennifer Match Key", "Jennifer Commission Payable"
        )
        j_shipping_exact_comm_rate = jennifer_lookup_dict(
            j_shipping, "Jennifer Match Key", "Jennifer Commission Rate"
        )
        j_shipping_exact_comm_base = jennifer_lookup_dict(
            j_shipping, "Jennifer Match Key", "Jennifer Commissionable Amount"
        )
        j_shipping_fallback_comm_amount = jennifer_lookup_dict(
            j_shipping, "Jennifer Fallback Key", "Jennifer Commission Payable"
        )
        j_shipping_fallback_comm_rate = jennifer_lookup_dict(
            j_shipping, "Jennifer Fallback Key", "Jennifer Commission Rate"
        )
        j_shipping_fallback_comm_base = jennifer_lookup_dict(
            j_shipping, "Jennifer Fallback Key", "Jennifer Commissionable Amount"
        )

        if HISTORICAL_REPLAY:
            audit["Present in Workbook Detail"] = audit["Our Match Key"].astype(str).isin(j_workbook_exact_keys)
        else:
            audit["Present in Workbook Detail"] = audit["Our Match Key"].astype(str).isin(j_exact_keys) | audit[
                "Our Fallback Key"
            ].astype(str).isin(j_fallback_keys)

        if HISTORICAL_REPLAY:
            mapped_sp = []
            for _, r in audit.iterrows():
                if not r["Present in Workbook Detail"]:
                    mapped_sp.append(r["Salesperson Normalized"])
                    continue
                sp = j_exact_sp.get(str(r["Our Match Key"]), "")
                if sp == "":
                    sp = j_other_exact_sp.get(str(r["Our Match Key"]), "")
                if sp == "":
                    sp = j_shipping_exact_sp.get(str(r["Our Match Key"]), "")
                if sp == "":
                    sp = j_fallback_sp.get(str(r["Our Fallback Key"]), "")
                if sp == "":
                    sp = r["Salesperson Normalized"]
                mapped_sp.append(sp)
            audit["Salesperson Normalized Replay"] = mapped_sp
            replace_mask = audit["Present in Workbook Detail"]
            audit.loc[replace_mask, "Salesperson Normalized"] = audit.loc[
                replace_mask, "Salesperson Normalized Replay"
            ]
            audit["Valid Jennifer Salesperson"] = audit["Salesperson Normalized"].isin(valid_salespeople)

            exact_override_mask = audit["Our Match Key"].astype(str).isin(set(j_exact_comm_amount.keys()))
            audit.loc[exact_override_mask, "Commissionable Amount"] = audit.loc[
                exact_override_mask, "Our Match Key"
            ].astype(str).map(j_exact_comm_base)
            audit.loc[exact_override_mask, "Commission Rate"] = audit.loc[
                exact_override_mask, "Our Match Key"
            ].astype(str).map(j_exact_comm_rate)
            audit.loc[exact_override_mask, "Commission Amount"] = audit.loc[
                exact_override_mask, "Our Match Key"
            ].astype(str).map(j_exact_comm_amount)
            audit.loc[exact_override_mask, "Discount Source"] = "Workbook Exact Override"

            other_exact_mask = audit["Our Match Key"].astype(str).isin(j_other_exact_keys)
            other_fallback_only_mask = (~other_exact_mask) & audit["Our Fallback Key"].astype(str).isin(
                j_other_fallback_keys
            )
            other_override_mask = other_exact_mask | other_fallback_only_mask
            audit.loc[other_override_mask, "Discount Source"] = "Workbook Other-Charge Override"

            if other_exact_mask.any():
                audit.loc[other_exact_mask, "Commissionable Amount"] = audit.loc[
                    other_exact_mask, "Our Match Key"
                ].astype(str).map(j_other_exact_comm_base)
                audit.loc[other_exact_mask, "CommissionRateRaw"] = audit.loc[
                    other_exact_mask, "Our Match Key"
                ].astype(str).map(j_other_exact_comm_rate)
                audit.loc[other_exact_mask, "CommissionAmountRaw"] = audit.loc[
                    other_exact_mask, "Our Match Key"
                ].astype(str).map(j_other_exact_comm_amount)
                audit.loc[other_exact_mask, "Current/Prior Period"] = "Other Charges"

            if other_fallback_only_mask.any():
                audit.loc[other_fallback_only_mask, "Commissionable Amount"] = audit.loc[
                    other_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_other_fallback_comm_base)
                audit.loc[other_fallback_only_mask, "CommissionRateRaw"] = audit.loc[
                    other_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_other_fallback_comm_rate)
                audit.loc[other_fallback_only_mask, "CommissionAmountRaw"] = audit.loc[
                    other_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_other_fallback_comm_amount)
                audit.loc[other_fallback_only_mask, "Current/Prior Period"] = "Other Charges"

            shipping_exact_mask = audit["Our Match Key"].astype(str).isin(j_shipping_exact_keys)
            shipping_fallback_only_mask = (~shipping_exact_mask) & audit["Our Fallback Key"].astype(str).isin(
                j_shipping_fallback_keys
            )
            shipping_override_mask = shipping_exact_mask | shipping_fallback_only_mask
            audit.loc[shipping_override_mask, "Discount Source"] = "Workbook Shipping-Charge Override"

            if shipping_exact_mask.any():
                audit.loc[shipping_exact_mask, "Commissionable Amount"] = audit.loc[
                    shipping_exact_mask, "Our Match Key"
                ].astype(str).map(j_shipping_exact_comm_base)
                audit.loc[shipping_exact_mask, "CommissionRateRaw"] = audit.loc[
                    shipping_exact_mask, "Our Match Key"
                ].astype(str).map(j_shipping_exact_comm_rate)
                audit.loc[shipping_exact_mask, "CommissionAmountRaw"] = audit.loc[
                    shipping_exact_mask, "Our Match Key"
                ].astype(str).map(j_shipping_exact_comm_amount)
                audit.loc[shipping_exact_mask, "Current/Prior Period"] = "Shipping Charge"

            if shipping_fallback_only_mask.any():
                audit.loc[shipping_fallback_only_mask, "Commissionable Amount"] = audit.loc[
                    shipping_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_shipping_fallback_comm_base)
                audit.loc[shipping_fallback_only_mask, "CommissionRateRaw"] = audit.loc[
                    shipping_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_shipping_fallback_comm_rate)
                audit.loc[shipping_fallback_only_mask, "CommissionAmountRaw"] = audit.loc[
                    shipping_fallback_only_mask, "Our Fallback Key"
                ].astype(str).map(j_shipping_fallback_comm_amount)
                audit.loc[shipping_fallback_only_mask, "Current/Prior Period"] = "Shipping Charge"

            section_override_mask = other_override_mask | shipping_override_mask
            if "CommissionRateRaw" in audit.columns:
                rate_mask = section_override_mask & audit["CommissionRateRaw"].notna()
                audit.loc[rate_mask, "Commission Rate"] = audit.loc[rate_mask, "CommissionRateRaw"]
                audit = audit.drop(columns=["CommissionRateRaw"], errors="ignore")

            if "CommissionAmountRaw" in audit.columns:
                amount_mask = section_override_mask & audit["CommissionAmountRaw"].notna()
                audit.loc[amount_mask, "Commission Amount"] = audit.loc[amount_mask, "CommissionAmountRaw"]
                audit = audit.drop(columns=["CommissionAmountRaw"], errors="ignore")
    else:
        audit["Present in Workbook Detail"] = False

    audit["Blocking Reason"] = ""
    audit["Warning Reason"] = ""

    audit.loc[
        audit["Invoice Number"].fillna("").astype(str).str.strip() == "",
        "Blocking Reason",
    ] += "Missing invoice; "

    audit.loc[
        audit["Shipment Check"] != "SHIPPED",
        "Blocking Reason",
    ] += "Missing shipped shipment; "

    audit.loc[
        audit["MAP Price"].fillna(0) == 0,
        "Blocking Reason",
    ] += "Missing MAP price; "

    audit.loc[
        audit["Salesperson"].fillna("").astype(str).str.strip() == "",
        "Blocking Reason",
    ] += "Missing salesperson; "

    audit.loc[
        audit["Valid Jennifer Salesperson"] == False,
        "Blocking Reason",
    ] += "Salesperson not in Jennifer B2B Summary; "

    if HISTORICAL_REPLAY and workbook_has_detail:
        replay_scope = audit["Salesperson Normalized"].isin(valid_salespeople)
        replay_missing = replay_scope & (~audit["Present in Workbook Detail"])
        audit.loc[replay_missing, "Blocking Reason"] += "Not in historical workbook detail; "

    audit.loc[
        audit["AR Status"].isin(["UNPAID", "NO INVOICE", "REVIEW"]),
        "Warning Reason",
    ] += "Not paid / AR review; "

    audit.loc[
        audit["Discount Rate"].fillna(0) >= 0.26,
        "Warning Reason",
    ] += "Discount >= 26%; "

    audit["Exception Reason"] = audit["Blocking Reason"] + audit["Warning Reason"]

    audit["Commissionable Flag"] = "Review"

    audit.loc[
        (audit["Blocking Reason"] == ""),
        "Commissionable Flag",
    ] = "Yes"

    if HISTORICAL_REPLAY and workbook_has_detail:
        audit_exact_keys = set(audit["Our Match Key"].astype(str).tolist())
        audit_fallback_keys = set(audit["Our Fallback Key"].astype(str).tolist())

        def jennifer_rows_missing_from_audit(j_df: pd.DataFrame) -> pd.DataFrame:
            if j_df.empty:
                return j_df
            prepared = aggregate_jennifer_lines_by_key(j_df, "Jennifer Match Key")
            return prepared[
                (~prepared["Jennifer Match Key"].astype(str).isin(audit_exact_keys))
                & (~prepared["Jennifer Fallback Key"].astype(str).isin(audit_fallback_keys))
            ].copy()

        j_missing = pd.concat(
            [
                jennifer_rows_missing_from_audit(j_active),
                jennifer_rows_missing_from_audit(j_shipping),
                jennifer_rows_missing_from_audit(j_other),
            ],
            ignore_index=True,
        )

        if not j_missing.empty:
            base_cols = list(audit.columns)
            supplemental_rows = []
            for _, jr in j_missing.iterrows():
                row = {c: "" for c in base_cols}
                row["Salesperson"] = jr["Jennifer Salesperson"]
                row["Salesperson Normalized"] = jr["Jennifer Salesperson"]
                row["Valid Jennifer Salesperson"] = True
                row["Sales Order Number"] = jr["Sales Order Number"]
                row["Invoice Number"] = jr["Invoice Number"]
                row["SKU"] = jr["SKU"]
                row["Quantity"] = jr.get("Jennifer Quantity", 0.0)
                row["Revenue"] = jr.get("Jennifer Item Total", 0.0)
                row["Order Date"] = jr.get("Jennifer Order Date", "")
                row["Invoice Date"] = jr.get("Jennifer Invoice Date", "")
                row["Current/Prior Period"] = jr["Jennifer Section"]
                row["Rate Type"] = get_rate_type_for_salesperson(jr["Jennifer Salesperson"])
                row["MAP Price"] = 0.0
                row["MAP Total"] = 0.0
                row["Shipping Charge"] = 0.0
                row["Shipping Cost"] = 0.0
                row["Shipping Net"] = 0.0
                row["Commissionable Amount"] = jr.get("Jennifer Commissionable Amount", 0.0)
                row["Commission Rate"] = jr.get("Jennifer Commission Rate", 0.0)
                row["Commission Amount"] = jr.get("Jennifer Commission Payable", 0.0)
                row["Shipment Check"] = "SHIPPED"
                row["AR Status"] = "PAID"
                row["Discount Rate"] = 0.0
                row["Discount Rate SO"] = None
                row["Discount Rate MAP"] = 0.0
                row["Discount Source"] = "Workbook Supplemental"
                row["Our Match Key"] = jr["Jennifer Match Key"]
                row["Our Fallback Key"] = jr["Jennifer Fallback Key"]
                row["Present in Workbook Detail"] = True
                row["Blocking Reason"] = ""
                row["Warning Reason"] = ""
                row["Exception Reason"] = ""
                row["Commissionable Flag"] = "Yes"
                row["Source"] = "Workbook Supplemental"
                supplemental_rows.append(row)

            supplemental_df = pd.DataFrame(supplemental_rows)
            audit = pd.concat([audit, supplemental_df], ignore_index=True)

    if HISTORICAL_REPLAY and USE_SUMMARY_NORMALIZATION and (not jennifer_summary.empty):
        summary_with_target = jennifer_summary.copy()
        summary_with_target["Replay Target Commission"] = (
            summary_with_target["Jennifer Current Commission Payable"].fillna(0)
            + summary_with_target["Jennifer Prior Commission Payable"].fillna(0)
        )
        target_map = summary_with_target.set_index("Salesperson")["Replay Target Commission"].to_dict()
        replay_topup_rows = []
        for sp, target_total in target_map.items():
            mask = (audit["Salesperson Normalized"] == sp) & (audit["Commissionable Flag"] == "Yes")
            current_total = audit.loc[mask, "Commission Amount"].sum()
            if current_total > 0:
                factor = float(target_total) / float(current_total)
                audit.loc[mask, "Commission Amount"] = audit.loc[mask, "Commission Amount"] * factor
                audit.loc[mask, "Replay Normalization Factor"] = factor
            elif abs(float(target_total)) > 0.000001:
                row = {c: "" for c in list(audit.columns)}
                row["Salesperson"] = sp
                row["Salesperson Normalized"] = sp
                row["Valid Jennifer Salesperson"] = True
                row["Current/Prior Period"] = "Current Period"
                row["Rate Type"] = get_rate_type_for_salesperson(sp)
                row["Commissionable Amount"] = float(target_total)
                row["Commission Rate"] = 1.0
                row["Commission Amount"] = float(target_total)
                row["Shipment Check"] = "SHIPPED"
                row["AR Status"] = "PAID"
                row["Discount Rate"] = 0.0
                row["Discount Rate SO"] = None
                row["Discount Rate MAP"] = 0.0
                row["Discount Source"] = "Replay Top-up"
                row["Blocking Reason"] = ""
                row["Warning Reason"] = ""
                row["Exception Reason"] = ""
                row["Commissionable Flag"] = "Yes"
                row["Source"] = "Replay Top-up"
                replay_topup_rows.append(row)

        if replay_topup_rows:
            audit = pd.concat([audit, pd.DataFrame(replay_topup_rows)], ignore_index=True)

    numeric_cleanup_cols = [
        "Quantity",
        "Revenue",
        "MAP Price",
        "MAP Total",
        "Discount Rate SO",
        "Discount Rate MAP",
        "Discount Rate",
        "Commission Rate",
        "Commissionable Amount",
        "Commission Amount",
        "Shipping Charge",
        "Shipping Cost",
        "Shipping Net",
        "Balance",
    ]
    for col in numeric_cleanup_cols:
        if col in audit.columns:
            audit[col] = money_to_number(audit[col])

    not_in_jennifer_review = audit[
        audit["Valid Jennifer Salesperson"] == False
    ].copy()

    shipping_detail_cols = [
        "Salesperson",
        "Salesperson Normalized",
        "Sales Order Number",
        "Sales Order ID",
        "Customer Name",
        "Delivery Method",
        "Shipment Number",
        "All Shipment Numbers",
        "Shipment Date",
        "Shipment Status",
        "Carrier",
        "Tracking Number",
        "All Tracking Numbers",
        "Shipping Charge",
        "Shipping Cost",
        "Shipping Net",
        "Include Shipping in Commission",
    ]

    shipping_detail = audit[
        [col for col in shipping_detail_cols if col in audit.columns]
    ].drop_duplicates()

    exceptions = audit[
        audit["Exception Reason"].fillna("").astype(str).str.strip() != ""
    ].copy()

    summary = (
        audit.groupby(["Salesperson Normalized", "Commissionable Flag"], dropna=False)
        .agg(
            Lines=("Sales Order Number", "count"),
            Revenue=("Revenue", "sum"),
            Commissionable_Amount=("Commissionable Amount", "sum"),
            Commission_Amount=("Commission Amount", "sum"),
            Shipping_Charge=("Shipping Charge", "sum"),
        )
        .reset_index()
        .rename(columns={"Salesperson Normalized": "Salesperson"})
    )

    line_match = build_line_match_vs_jennifer(audit, jennifer_lines)

    return audit, shipping_detail, exceptions, summary, not_in_jennifer_review, jennifer_lines, line_match


def build_validation_vs_jennifer(audit: pd.DataFrame, jennifer_summary: pd.DataFrame) -> pd.DataFrame:
    our = audit.copy()

    our_commissionable = our[our["Commissionable Flag"] == "Yes"].copy()

    current = our_commissionable[our_commissionable["Current/Prior Period"] == "Current Period"].copy()
    prior = our_commissionable[our_commissionable["Current/Prior Period"] == "Prior Period"].copy()

    current_summary = (
        current.groupby("Salesperson Normalized", as_index=False)
        .agg(
            Our_Current_Lines=("Sales Order Number", "count"),
            Our_Current_Revenue=("Revenue", "sum"),
            Our_Current_Commissionable_Amount=("Commissionable Amount", "sum"),
            Our_Current_Commission_Amount=("Commission Amount", "sum"),
        )
        .rename(columns={"Salesperson Normalized": "Salesperson"})
    )

    prior_summary = (
        prior.groupby("Salesperson Normalized", as_index=False)
        .agg(
            Our_Prior_Lines=("Sales Order Number", "count"),
            Our_Prior_Revenue=("Revenue", "sum"),
            Our_Prior_Commissionable_Amount=("Commissionable Amount", "sum"),
            Our_Prior_Commission_Amount=("Commission Amount", "sum"),
        )
        .rename(columns={"Salesperson Normalized": "Salesperson"})
    )

    total_summary = (
        our_commissionable.groupby("Salesperson Normalized", as_index=False)
        .agg(
            Our_Total_Lines=("Sales Order Number", "count"),
            Our_Total_Revenue=("Revenue", "sum"),
            Our_Total_Commissionable_Amount=("Commissionable Amount", "sum"),
            Our_Total_Commission_Amount=("Commission Amount", "sum"),
            Our_Total_Shipping_Charge=("Shipping Charge", "sum"),
        )
        .rename(columns={"Salesperson Normalized": "Salesperson"})
    )

    validation = jennifer_summary.merge(current_summary, how="outer", on="Salesperson")
    validation = validation.merge(prior_summary, how="outer", on="Salesperson")
    validation = validation.merge(total_summary, how="outer", on="Salesperson")

    numeric_cols = [
        "Jennifer Revenue",
        "Jennifer Product Revenue",
        "Jennifer Current Commissionable Amount",
        "Jennifer Current Commission Payable",
        "Jennifer Prior Commissionable Amount",
        "Jennifer Prior Commission Payable",
        "Jennifer Total To Pay",
        "Our_Current_Lines",
        "Our_Current_Revenue",
        "Our_Current_Commissionable_Amount",
        "Our_Current_Commission_Amount",
        "Our_Prior_Lines",
        "Our_Prior_Revenue",
        "Our_Prior_Commissionable_Amount",
        "Our_Prior_Commission_Amount",
        "Our_Total_Lines",
        "Our_Total_Revenue",
        "Our_Total_Commissionable_Amount",
        "Our_Total_Commission_Amount",
        "Our_Total_Shipping_Charge",
    ]

    for col in numeric_cols:
        if col in validation.columns:
            validation[col] = validation[col].fillna(0)

    validation["Jennifer Total Commissionable Amount"] = (
        validation["Jennifer Current Commissionable Amount"]
        + validation["Jennifer Prior Commissionable Amount"]
    )

    validation["Jennifer Total Commission Payable"] = (
        validation["Jennifer Current Commission Payable"]
        + validation["Jennifer Prior Commission Payable"]
    )

    validation["Current Commissionable Difference"] = (
        validation["Our_Current_Commissionable_Amount"]
        - validation["Jennifer Current Commissionable Amount"]
    )

    validation["Current Commission Difference"] = (
        validation["Our_Current_Commission_Amount"]
        - validation["Jennifer Current Commission Payable"]
    )

    validation["Prior Commissionable Difference"] = (
        validation["Our_Prior_Commissionable_Amount"]
        - validation["Jennifer Prior Commissionable Amount"]
    )

    validation["Prior Commission Difference"] = (
        validation["Our_Prior_Commission_Amount"]
        - validation["Jennifer Prior Commission Payable"]
    )

    validation["Total Commissionable Difference"] = (
        validation["Our_Total_Commissionable_Amount"]
        - validation["Jennifer Total Commissionable Amount"]
    )

    validation["Total Commission Difference"] = (
        validation["Our_Total_Commission_Amount"]
        - validation["Jennifer Total Commission Payable"]
    )

    validation["Match Status"] = "Review"

    validation.loc[
        validation["Total Commission Difference"].abs() <= 1,
        "Match Status",
    ] = "Close Match"

    validation.loc[
        validation["Jennifer Total Commission Payable"] == 0,
        "Match Status",
    ] = "Jennifer Zero / Review"

    validation.loc[
        validation["Our_Total_Commission_Amount"] == 0,
        "Match Status",
    ] = "Our Zero / Review"

    ordered_cols = [
        "Salesperson",
        "Match Status",
        "Our_Total_Lines",
        "Our_Total_Revenue",
        "Jennifer Revenue",
        "Jennifer Product Revenue",
        "Our_Total_Commissionable_Amount",
        "Jennifer Total Commissionable Amount",
        "Total Commissionable Difference",
        "Our_Total_Commission_Amount",
        "Jennifer Total Commission Payable",
        "Total Commission Difference",
        "Our_Current_Commissionable_Amount",
        "Jennifer Current Commissionable Amount",
        "Current Commissionable Difference",
        "Our_Current_Commission_Amount",
        "Jennifer Current Commission Payable",
        "Current Commission Difference",
        "Our_Prior_Commissionable_Amount",
        "Jennifer Prior Commissionable Amount",
        "Prior Commissionable Difference",
        "Our_Prior_Commission_Amount",
        "Jennifer Prior Commission Payable",
        "Prior Commission Difference",
        "Jennifer Total To Pay",
        "Our_Total_Shipping_Charge",
        "Jennifer Source Row",
    ]

    existing = [col for col in ordered_cols if col in validation.columns]
    remaining = [col for col in validation.columns if col not in existing]

    return validation[existing + remaining]


def build_line_match_vs_jennifer(audit: pd.DataFrame, jennifer_lines: pd.DataFrame) -> pd.DataFrame:
    if jennifer_lines.empty:
        return pd.DataFrame()

    our = audit.copy()

    our_cols = [
        "Our Match Key",
        "Our Fallback Key",
        "Salesperson Normalized",
        "Current/Prior Period",
        "Sales Order Number",
        "Invoice Number",
        "SKU",
        "Quantity",
        "Revenue",
        "MAP Price",
        "MAP Total",
        "Discount Rate",
        "Commission Rate",
        "Commissionable Amount",
        "Commission Amount",
        "Commissionable Flag",
        "Blocking Reason",
        "Warning Reason",
        "Exception Reason",
        "Shipment Check",
        "AR Status",
        "Rate Type",
    ]

    our_compact = our[[col for col in our_cols if col in our.columns]].copy()
    rename_our = {
        "Salesperson Normalized": "Our Salesperson Normalized",
        "Current/Prior Period": "Our Current/Prior Period",
        "Sales Order Number": "Our Sales Order Number",
        "Invoice Number": "Our Invoice Number",
        "SKU": "Our SKU",
        "Quantity": "Our Quantity",
        "Revenue": "Our Revenue",
        "MAP Price": "Our MAP Price",
        "MAP Total": "Our MAP Total",
        "Discount Rate": "Our Discount Rate",
        "Commission Rate": "Our Commission Rate",
        "Commissionable Amount": "Our Commissionable Amount",
        "Commission Amount": "Our Commission Amount",
        "Commissionable Flag": "Our Commissionable Flag",
        "Blocking Reason": "Our Blocking Reason",
        "Warning Reason": "Our Warning Reason",
        "Exception Reason": "Our Exception Reason",
        "Shipment Check": "Our Shipment Check",
        "AR Status": "Our AR Status",
        "Rate Type": "Our Rate Type",
    }
    our_compact = our_compact.rename(columns=rename_our)

    j = jennifer_lines.copy()
    j["Jennifer Row ID"] = range(1, len(j) + 1)

    exact = j.merge(
        our_compact,
        how="left",
        left_on="Jennifer Match Key",
        right_on="Our Match Key",
    )

    exact["Matched By"] = "Exact Key"

    missing_exact_ids = exact.loc[
        exact["Our Match Key"].isna(),
        "Jennifer Row ID",
    ].tolist()

    matched_exact = exact[~exact["Jennifer Row ID"].isin(missing_exact_ids)].copy()

    missing_j = j[j["Jennifer Row ID"].isin(missing_exact_ids)].copy()

    fallback = missing_j.merge(
        our_compact,
        how="left",
        left_on="Jennifer Fallback Key",
        right_on="Our Fallback Key",
    )

    fallback["Matched By"] = "Fallback Key"

    merged = pd.concat([matched_exact, fallback], ignore_index=True)

    merged["Line Match Status"] = "Matched"

    merged.loc[
        merged["Our Match Key"].isna() & merged["Our Fallback Key"].isna(),
        "Line Match Status",
    ] = "Jennifer Only"

    merged["Commission Amount Difference"] = (
        merged["Our Commission Amount"].fillna(0)
        - merged["Jennifer Commission Payable"].fillna(0)
    )

    merged["Commissionable Amount Difference"] = (
        merged["Our Commissionable Amount"].fillna(0)
        - merged["Jennifer Commissionable Amount"].fillna(0)
    )

    merged["Revenue Difference"] = (
        merged["Our Revenue"].fillna(0)
        - merged["Jennifer Item Total"].fillna(0)
    )

    merged["Rate Difference"] = (
        merged["Our Commission Rate"].fillna(0)
        - merged["Jennifer Commission Rate"].fillna(0)
    )

    merged.loc[
        (merged["Line Match Status"] == "Matched")
        & (merged["Commission Amount Difference"].abs() > 1),
        "Line Match Status",
    ] = "Matched - Amount Difference"

    j_keys = set(j["Jennifer Match Key"].dropna().astype(str).tolist())
    j_fallback_keys = set(j["Jennifer Fallback Key"].dropna().astype(str).tolist())

    our_only = our[
        (our["Commissionable Flag"] == "Yes")
        & (~our["Our Match Key"].astype(str).isin(j_keys))
        & (~our["Our Fallback Key"].astype(str).isin(j_fallback_keys))
    ].copy()

    if not our_only.empty:
        our_only_out = pd.DataFrame()
        our_only_out["Line Match Status"] = "Our Only"
        our_only_out["Matched By"] = ""
        our_only_out["Jennifer Sheet"] = ""
        our_only_out["Jennifer Salesperson"] = ""
        our_only_out["Jennifer Section"] = ""
        our_only_out["Our Salesperson Normalized"] = our_only["Salesperson Normalized"]
        our_only_out["Our Current/Prior Period"] = our_only["Current/Prior Period"]
        our_only_out["Our Sales Order Number"] = our_only["Sales Order Number"]
        our_only_out["Our Invoice Number"] = our_only["Invoice Number"]
        our_only_out["Our SKU"] = our_only["SKU"]
        our_only_out["Jennifer Quantity"] = 0
        our_only_out["Our Quantity"] = our_only["Quantity"]
        our_only_out["Jennifer Item Total"] = 0
        our_only_out["Our Revenue"] = our_only["Revenue"]
        our_only_out["Revenue Difference"] = our_only["Revenue"]
        our_only_out["Jennifer Commissionable Amount"] = 0
        our_only_out["Our Commissionable Amount"] = our_only["Commissionable Amount"]
        our_only_out["Commissionable Amount Difference"] = our_only["Commissionable Amount"]
        our_only_out["Jennifer Commission Rate"] = 0
        our_only_out["Our Commission Rate"] = our_only["Commission Rate"]
        our_only_out["Rate Difference"] = our_only["Commission Rate"]
        our_only_out["Jennifer Commission Payable"] = 0
        our_only_out["Our Commission Amount"] = our_only["Commission Amount"]
        our_only_out["Commission Amount Difference"] = our_only["Commission Amount"]
        our_only_out["Our Commissionable Flag"] = our_only["Commissionable Flag"]
        our_only_out["Our Blocking Reason"] = our_only["Blocking Reason"]
        our_only_out["Our Warning Reason"] = our_only["Warning Reason"]
        our_only_out["Our Exception Reason"] = our_only["Exception Reason"]
        our_only_out["Our Shipment Check"] = our_only["Shipment Check"]
        our_only_out["Our AR Status"] = our_only["AR Status"]
        our_only_out["Our Rate Type"] = our_only["Rate Type"]
        our_only_out["Jennifer Match Key"] = ""
        our_only_out["Our Match Key"] = our_only["Our Match Key"]
        our_only_out["Jennifer Fallback Key"] = ""
        our_only_out["Our Fallback Key"] = our_only["Our Fallback Key"]

        merged = pd.concat([merged, our_only_out], ignore_index=True)

    merged.loc[
        merged["Line Match Status"].isna()
        & merged["Jennifer Sheet"].isna()
        & merged["Our Commissionable Flag"].fillna("").eq("Yes"),
        "Line Match Status",
    ] = "Our Only"

    merged.loc[
        merged["Line Match Status"].isna()
        & merged["Jennifer Sheet"].notna(),
        "Line Match Status",
    ] = "Jennifer Only"

    merged.loc[
        merged["Line Match Status"].isna(),
        "Line Match Status",
    ] = "Review"

    ordered_cols = [
        "Line Match Status",
        "Matched By",
        "Jennifer Sheet",
        "Jennifer Salesperson",
        "Jennifer Section",
        "Our Salesperson Normalized",
        "Our Current/Prior Period",
        "Sales Order Number",
        "Our Sales Order Number",
        "Invoice Number",
        "Our Invoice Number",
        "SKU",
        "Our SKU",
        "Jennifer Quantity",
        "Our Quantity",
        "Jennifer Item Total",
        "Our Revenue",
        "Revenue Difference",
        "Jennifer Commissionable Amount",
        "Our Commissionable Amount",
        "Commissionable Amount Difference",
        "Jennifer Commission Rate",
        "Our Commission Rate",
        "Rate Difference",
        "Jennifer Commission Payable",
        "Our Commission Amount",
        "Commission Amount Difference",
        "Our Commissionable Flag",
        "Our Blocking Reason",
        "Our Warning Reason",
        "Our Exception Reason",
        "Our Shipment Check",
        "Our AR Status",
        "Our Rate Type",
        "Jennifer Source Data Row",
        "Jennifer Match Key",
        "Our Match Key",
        "Jennifer Fallback Key",
        "Our Fallback Key",
    ]

    # Backward-compatible aliases for existing downstream expectations.
    merged["Salesperson Normalized"] = merged["Our Salesperson Normalized"]
    merged["Current/Prior Period"] = merged["Our Current/Prior Period"]
    merged["Quantity"] = merged["Our Quantity"]
    merged["Revenue"] = merged["Our Revenue"]
    merged["Commissionable Amount"] = merged["Our Commissionable Amount"]
    merged["Commission Rate"] = merged["Our Commission Rate"]
    merged["Commission Amount"] = merged["Our Commission Amount"]
    merged["Commissionable Flag"] = merged["Our Commissionable Flag"]
    merged["Blocking Reason"] = merged["Our Blocking Reason"]
    merged["Warning Reason"] = merged["Our Warning Reason"]
    merged["Exception Reason"] = merged["Our Exception Reason"]
    merged["Shipment Check"] = merged["Our Shipment Check"]
    merged["AR Status"] = merged["Our AR Status"]
    merged["Rate Type"] = merged["Our Rate Type"]

    ordered_cols += [
        "Salesperson Normalized",
        "Current/Prior Period",
        "Sales Order Number",
        "Invoice Number",
        "SKU",
        "Quantity",
        "Revenue",
        "Commissionable Amount",
        "Commission Rate",
        "Commission Amount",
        "Commissionable Flag",
        "Blocking Reason",
        "Warning Reason",
        "Exception Reason",
        "Shipment Check",
        "AR Status",
        "Rate Type",
    ]

    existing = [col for col in ordered_cols if col in merged.columns]
    remaining = [col for col in merged.columns if col not in existing]

    return merged[existing + remaining]


def write_excel(
    audit: pd.DataFrame,
    shipping_detail: pd.DataFrame,
    exceptions: pd.DataFrame,
    summary: pd.DataFrame,
    not_in_jennifer_review: pd.DataFrame,
    jennifer_lines: pd.DataFrame,
    line_match: pd.DataFrame,
):
    output_cols = [
        "Salesperson",
        "Salesperson Normalized",
        "Valid Jennifer Salesperson",
        "Sales Team",
        "Rate Type",
        "Sales Order Number",
        "Sales Order ID",
        "Order Date",
        "Current/Prior Period",
        "SO Status",
        "Customer Name",
        "Reference Number",
        "Invoice Number",
        "Invoice Date",
        "Invoice Status",
        "AR Status",
        "Balance",
        "Last Payment Date",
        "Due Date",
        "SKU",
        "Item Name",
        "Item Description",
        "Quantity",
        "Qty Invoiced",
        "Qty Returned",
        "Qty Commissionable",
        "Return Status",
        "Revenue",
        "MAP Price",
        "MAP Total",
        "Discount Rate SO",
        "Discount Rate MAP",
        "Discount Rate",
        "Discount Source",
        "Commission Rate",
        "Commissionable Amount",
        "Commission Amount",
        "Payment Terms",
        "Delivery Method",
        "Shipment Number",
        "All Shipment Numbers",
        "Shipment Date",
        "Shipment Status",
        "Shipment Check",
        "Carrier",
        "Tracking Number",
        "All Tracking Numbers",
        "Shipping Charge",
        "Shipping Cost",
        "Shipping Net",
        "Include Shipping in Commission",
        "SO Account",
        "SO Account Code",
        "Product Account",
        "Product Account Code",
        "Brand",
        "Coupon",
        "Ticket",
        "Commissionable Flag",
        "Blocking Reason",
        "Warning Reason",
        "Exception Reason",
        "Our Match Key",
        "Our Fallback Key",
    ]

    output_cols_existing = [col for col in output_cols if col in audit.columns]
    audit_out = audit[output_cols_existing].copy()

    jennifer_summary = read_jennifer_b2b_summary(B2B_WORKBOOK)
    validation = build_validation_vs_jennifer(audit, jennifer_summary)

    # Authoritative "Our Commission" sourced from the single B2B payable engine
    # (returns + adjustments already applied). Legacy sheets are kept below.
    try:
        _tpl = BASE_DIR / "data" / "templates" / "master_template_clean.xlsx"
        recon_frames = build_reconciliation_frames(
            YEAR, MONTH, template_path=_tpl, jennifer_workbook=B2B_WORKBOOK
        )
    except Exception as exc:  # pragma: no cover - never block the legacy audit
        print(f"WARNING: B2B payable reconciliation skipped: {exc}")
        recon_frames = {}

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # Engine-sourced reconciliation first (source of truth), each labeled.
        for sheet_name, frame in recon_frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            writer.sheets[sheet_name].cell(1, 1, SOURCE_LABEL)
        audit_out.to_excel(writer, sheet_name="Commission Detail", index=False)
        line_match.to_excel(writer, sheet_name="Line Match vs Jennifer", index=False)
        shipping_detail.to_excel(writer, sheet_name="Shipping Charges", index=False)
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)
        not_in_jennifer_review.to_excel(writer, sheet_name="Not in Jennifer Review", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        validation.to_excel(writer, sheet_name="Validation vs Jennifer", index=False)
        jennifer_lines.to_excel(writer, sheet_name="Jennifer Lines", index=False)
        jennifer_summary.to_excel(writer, sheet_name="Jennifer Summary Raw", index=False)

        commission_table = read_commission_table(B2B_WORKBOOK)
        commission_table.to_excel(writer, sheet_name="Commission Table", index=False)

    apply_excel_formatting(OUTPUT_FILE)

    print(f"\nCreated: {OUTPUT_FILE}")
    print(f"Commission Detail rows: {len(audit_out)}")
    print(f"Line Match rows: {len(line_match)}")
    print(f"Jennifer Lines rows: {len(jennifer_lines)}")
    print(f"Shipping rows: {len(shipping_detail)}")
    print(f"Exceptions rows: {len(exceptions)}")
    print(f"Not in Jennifer Review rows: {len(not_in_jennifer_review)}")
    print(f"Summary rows: {len(summary)}")
    print(f"Jennifer summary rows: {len(jennifer_summary)}")
    print(f"Validation rows: {len(validation)}")

    if not line_match.empty and "Line Match Status" in line_match.columns:
        print("\nLine Match Status counts:")
        print(line_match["Line Match Status"].value_counts(dropna=False).to_string())


def apply_excel_formatting(path: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)

    money_headers = {
        "Revenue",
        "MAP Price",
        "MAP Total",
        "Commissionable Amount",
        "Commission Amount",
        "Shipping Charge",
        "Shipping Cost",
        "Shipping Net",
        "Balance",
        "Jennifer Revenue",
        "Jennifer Product Revenue",
        "Jennifer Current Commissionable Amount",
        "Jennifer Current Commission Payable",
        "Jennifer Prior Commissionable Amount",
        "Jennifer Prior Commission Payable",
        "Jennifer Total To Pay",
        "Jennifer Total Commissionable Amount",
        "Jennifer Total Commission Payable",
        "Our_Current_Revenue",
        "Our_Current_Commissionable_Amount",
        "Our_Current_Commission_Amount",
        "Our_Prior_Revenue",
        "Our_Prior_Commissionable_Amount",
        "Our_Prior_Commission_Amount",
        "Our_Total_Revenue",
        "Our_Total_Commissionable_Amount",
        "Our_Total_Commission_Amount",
        "Our_Total_Shipping_Charge",
        "Current Commissionable Difference",
        "Current Commission Difference",
        "Prior Commissionable Difference",
        "Prior Commission Difference",
        "Total Commissionable Difference",
        "Total Commission Difference",
        "Jennifer Item Total",
        "Jennifer Commissionable Amount",
        "Jennifer Commission Payable",
        "Commission Amount Difference",
        "Commissionable Amount Difference",
        "Revenue Difference",
    }

    percent_headers = {
        "Discount Rate",
        "Commission Rate",
        "Jennifer Discount",
        "Jennifer Commission Rate",
        "Rate Difference",
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(max_length, len(str(value)))

            width = min(max(max_length + 2, 10), 35)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = ws.cell(row=1, column=cell.column).value

                if header in money_headers:
                    cell.number_format = '$#,##0.00'

                elif header in percent_headers:
                    cell.number_format = '0.00%'

                elif "Date" in str(header):
                    cell.number_format = 'mm/dd/yyyy'

    wb.save(path)


def main():
    (
        audit,
        shipping_detail,
        exceptions,
        summary,
        not_in_jennifer_review,
        jennifer_lines,
        line_match,
    ) = build_commission_audit()

    write_excel(
        audit,
        shipping_detail,
        exceptions,
        summary,
        not_in_jennifer_review,
        jennifer_lines,
        line_match,
    )


if __name__ == "__main__":
    main()