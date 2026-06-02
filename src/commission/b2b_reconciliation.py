"""
B2B Payable reconciliation — sources "Our Commission" from the SINGLE B2B payable
engine (``build_salespeople_from_sqlite``), then compares it to Jennifer's
historical workbook.

This does NOT recompute commission with separate logic (no third calculation
path): it reads the exact lines/totals used by the payable workbook (returns,
manual adjustments and all rules already applied) and lines them up against
Jennifer's numbers.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SOURCE_LABEL = "Our calculated commission is sourced from the B2B payable engine."

# Jennifer display name -> payable sheet key
NAME_TO_SHEET = {
    "Paul Perlman": "Paul", "Jose Ayala": "Jose", "Michael Ayala": "Michael",
    "Jim Sutton": "Jim", "Weston Fields": "Weston", "Brett Bern": "Brett",
    "Leslie Neipert": "Leslie", "Carmen Daetz": "Carmen", "Garrett Lockhart": "Garrett",
    "B2B Company Account": "Company Acct", "Company Account": "Company Acct",
}


def jennifer_summary_totals(workbook_path: Path) -> dict[str, float]:
    """Per-salesperson total (current + prior) from Jennifer's B2B Summary sheet."""
    out: dict[str, float] = {}
    if not workbook_path or not Path(workbook_path).exists():
        return out
    wb = load_workbook(workbook_path, data_only=True)
    if "B2B Summary" not in wb.sheetnames:
        wb.close()
        return out
    ws = wb["B2B Summary"]
    skip = {"salary commission", "non-salary commission", "b2b sales rep",
            "total b2b", "b2b executive account"}
    for r in range(20, 45):
        name = ws.cell(r, 5).value or ws.cell(r, 4).value
        if not name:
            continue
        name = str(name).strip()
        if name.lower() in skip:
            continue
        cur = ws.cell(r, 13).value or 0
        pri = ws.cell(r, 19).value or 0
        try:
            tot = float(cur) + float(pri)
        except (TypeError, ValueError):
            continue
        if abs(tot) > 0.005:
            key = NAME_TO_SHEET.get(name, name)
            out[key] = round(out.get(key, 0.0) + tot, 2)
    wb.close()
    return out


def build_reconciliation_frames(
    year: int,
    month: int,
    *,
    template_path: Path,
    jennifer_workbook: Path,
    db_path: Path | None = None,
) -> dict[str, pd.DataFrame]:
    """Return {sheet_name: DataFrame} sourced from the B2B payable engine."""
    from src.commission.sqlite_to_workbook import (
        build_salespeople_from_sqlite,
        load_map_from_template,
        load_tiers_from_template,
    )

    tiers = load_tiers_from_template(template_path)
    rlp = load_map_from_template(template_path)
    result = build_salespeople_from_sqlite(year, month, db_path=db_path, tiers=tiers, rlp_map=rlp)
    audit_rows = result.audit_rows
    our_totals = {k: round(v, 2) for k, v in result.totals_by_sheet.items()}
    jen_totals = jennifer_summary_totals(jennifer_workbook)

    # --- 1) Per-salesperson comparison (Our payable vs Jennifer) ---
    keys = sorted(set(our_totals) | set(jen_totals))
    cmp_rows = []
    for k in keys:
        ours = round(our_totals.get(k, 0.0), 2)
        jen = round(jen_totals.get(k, 0.0), 2)
        if abs(ours) < 0.005 and abs(jen) < 0.005:
            continue
        cmp_rows.append({
            "Salesperson": k,
            "Our Commission (B2B Payable)": ours,
            "Jennifer Commission": jen,
            "Difference": round(ours - jen, 2),
        })
    cmp_df = pd.DataFrame(cmp_rows)
    if not cmp_df.empty:
        total = {
            "Salesperson": "TOTAL",
            "Our Commission (B2B Payable)": round(cmp_df["Our Commission (B2B Payable)"].sum(), 2),
            "Jennifer Commission": round(cmp_df["Jennifer Commission"].sum(), 2),
            "Difference": round(cmp_df["Difference"].sum(), 2),
        }
        cmp_df = pd.concat([cmp_df, pd.DataFrame([total])], ignore_index=True)

    # --- 2) Payable detail (exact lines that earn commission) ---
    pay_cols = [
        "salesperson", "sales_order", "invoice", "sku", "customer",
        "qty_invoiced", "qty_returned", "qty_commissionable", "return_status",
        "revenue", "final_commissionable", "final_rate", "final_commission",
    ]
    payable = [
        a for a in audit_rows
        if a.get("block") == "commissionable" and not a.get("pending") and not a.get("excluded")
    ]
    detail_df = pd.DataFrame([{c: a.get(c, "") for c in pay_cols} for a in payable])

    # --- 3) Excluded-from-payable, categorized ---
    def category(a: dict) -> str:
        cls = (a.get("classification") or "").lower()
        if cls == "company":
            return "Company Account Review"
        if cls == "executive":
            return "Executive Account Review"
        if a.get("excluded"):
            return "Excluded by Accounting"
        if a.get("salesperson") == "(unassigned)":
            return "Not Payable Yet (unassigned)"
        if a.get("pending"):
            return "Pending Review"
        return ""

    excl_rows = []
    for a in audit_rows:
        cat = category(a)
        if not cat:
            continue
        excl_rows.append({
            "Category": cat,
            "Salesperson": a.get("salesperson"),
            "Sales Team": a.get("sales_team"),
            "Sales Order": a.get("sales_order"),
            "Invoice": a.get("invoice"),
            "SKU": a.get("sku"),
            "Revenue": a.get("revenue"),
            "Est. Commission (system)": a.get("system_commission"),
            "Return Status": a.get("return_status"),
        })
    excl_df = pd.DataFrame(excl_rows)

    return {
        "B2B Payable vs Jennifer": cmp_df,
        "B2B Payable Detail": detail_df,
        "Excluded from Payable": excl_df,
    }
