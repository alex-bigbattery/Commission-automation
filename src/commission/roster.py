"""
Approved B2B commission roster and assignment display helpers.

Roster members are configurable via COMMISSION_ROSTER (comma-separated sheet keys,
e.g. ``Paul,Jose,Michael``). Full-name aliases are defined in DEFAULT_ROSTER_ENTRIES.

Business rules confirmed by Accounting (June 2026 governance form):
  - Exe./Comp. Account → exclude from payable unless manually approved.
  - Bruce Taylor = Company Account (special arrangement, not paid direct commissions).
  - Marshall Neipert = Executive Account (exception approver = Marshall himself).
  - Known inactive names must never auto-assign — flag for manual review.
  - Dylan Nava / Customer Service → B2C coupon-based rule, not standard B2B.
  - Ticket numbers on an order = usually noncommissionable (flag, don't auto-exclude).
  - Roster changes effective on date of hire; owner = Bruce Taylor.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

# Load local .env so COMMISSION_ROSTER / COMPANY_ACCOUNT_NAMES / EXECUTIVE_ACCOUNT_NAMES
# are honored regardless of import order (Render uses real env vars directly).
try:
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).resolve().parent.parent.parent / ".env")
except ImportError:
    pass

# Internal routing bucket for non-roster lines (never shown to Accounting).
ROUTING_UNASSIGNED = "(unassigned)"

COMPANY_SHEET = "Company Acct"
COMPANY_FULL = "Company Account"

# (sheet_key, zoho_full_name) — active B2B reps only.
# Confirmed active roster per Accounting governance form (June 2026).
DEFAULT_ROSTER_ENTRIES: list[tuple[str, str]] = [
    ("Paul",    "Paul Perlman"),
    ("Jose",    "Jose Ayala"),
    ("Michael", "Michael Ayala"),
    ("Jim",     "Jim Sutton"),
    ("Weston",  "Weston Fields"),
    ("Brett",   "Brett Bern"),
    ("Leslie",  "Leslie Neipert"),
    ("Carmen",  "Carmen Daetz"),
    ("Garrett", "Garrett Lockhart"),
]

EXTRA_NAME_ALIASES: dict[str, str] = {
    "Company Account":     COMPANY_SHEET,
    "B2B Company Account": COMPANY_SHEET,
}

NON_SALARIED_SHEETS = frozenset({"Brett", "Leslie", "Carmen", "Garrett"})

# ---- Known-inactive / non-B2B names (confirmed by Accounting, June 2026) ----
# If any of these appear as the Zoho salesperson, the line is flagged for
# manual review and never auto-assigned.
KNOWN_INACTIVE_NAMES: frozenset[str] = frozenset({
    "BB Affiliate Investment",
    "Executive Account",
    # "Marshall Neipert" is now handled as an Executive Account (see below), not inactive.
    "Dmitry Gorobets",
    "Michael Northcutt",
    "Joseph Mohney",
    "Jessica Key",
    "Kara Wagner",
    "Alexis McCarthy",
    "River Michelle Harrington",
    "Ryan Murphy",
})

# ---- Special-person routing rules (configurable) -------------------------
# Per Accounting (Marshall, June 2026): pay based on who the Zoho salesperson is.
#   Company Account (Bruce Taylor): route to the Company Acct sheet at NORMAL
#     commission; Bruce's payout is 20% of that (applied in B2B Summary J13).
#   Executive Account (Marshall Neipert, Eric): revenue tracked, commission = 0.
# Names are configurable via env (comma-separated). "Eric" — exact Zoho spelling
# is unconfirmed; set EXECUTIVE_ACCOUNT_NAMES with the precise name once known.
def _parse_names_env(var: str, default: list[str]) -> list[str]:
    raw = os.environ.get(var, "")
    if raw.strip():
        return [n.strip() for n in raw.split(",") if n.strip()]
    return list(default)


COMPANY_ACCOUNT_NAMES = _parse_names_env("COMPANY_ACCOUNT_NAMES", ["Bruce Taylor"])
EXECUTIVE_ACCOUNT_NAMES = _parse_names_env("EXECUTIVE_ACCOUNT_NAMES", ["Marshall Neipert", "Eric"])


def _build_special_routing() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for n in COMPANY_ACCOUNT_NAMES:
        if n.strip():
            out[n.strip().lower()] = {
                "flag": "COMPANY_ACCOUNT",
                "category": "company",
                "issue": "Company Account / Bruce special commission",
                "action": "Review only if exception",
            }
    for n in EXECUTIVE_ACCOUNT_NAMES:
        if n.strip():
            out[n.strip().lower()] = {
                "flag": "EXECUTIVE_ACCOUNT",
                "category": "executive",
                "issue": "Executive Account / no commission payable",
                "action": "Track revenue, no commission unless exception approved",
            }
    return out


SPECIAL_PERSON_ROUTING: dict[str, dict[str, str]] = _build_special_routing()

# ---- B2C / coupon-based reps ---------------------------------------------
# Dylan Nava and Customer Service commissions depend on coupon codes.
# Invoices without the B2C-RC Team coupon are organic sales (commissionable
# differently); invoices with B2C-Web Marketing coupon are NOT commissionable.
# These names must never land in the standard B2B payable automatically.
B2C_COUPON_REPS: frozenset[str] = frozenset({
    "dylan nava",
    "customer service",
})

MISSING_ZOHO_LABEL  = "(missing in Zoho)"
ISSUE_NOT_IN_ROSTER = "Salesperson not in commission roster"
ISSUE_MISSING_ZOHO  = "Missing salesperson in Zoho"
ACTION_NOT_IN_ROSTER = (
    "Classify as Company Account, Executive Account, Bruce Commission, "
    "assign to a salesperson, or add to roster"
)
ACTION_MISSING_ZOHO = "Assign salesperson"


def _parse_env_roster(raw: str) -> list[tuple[str, str]] | None:
    if not raw.strip():
        return None
    keys = [p.strip() for p in raw.split(",") if p.strip()]
    if not keys:
        return None
    default_by_sheet = {s: f for s, f in DEFAULT_ROSTER_ENTRIES}
    return [(k, default_by_sheet.get(k, k)) for k in keys]


def roster_rep_entries() -> list[tuple[str, str]]:
    if globals().get("_TPL_REPS"):            # business config (template) wins
        return list(_TPL_REPS)
    override = _parse_env_roster(os.environ.get("COMMISSION_ROSTER", ""))
    return override if override is not None else list(DEFAULT_ROSTER_ENTRIES)


def _parse_aliases_env() -> dict[str, str]:
    """Optional extra Zoho full-name -> sheet-key aliases, for spelling variants
    or new reps without a code change. Keeps COMMISSION_ROSTER as sheet keys only.

    Format (semicolon-separated pairs):
        SALESPERSON_ALIASES="Paul Perlman=Paul;Jose Ayala=Jose;Brett Bern=Brett"
    """
    raw = os.environ.get("SALESPERSON_ALIASES", "")
    out: dict[str, str] = {}
    for pair in raw.split(";"):
        if "=" in pair:
            full, key = pair.split("=", 1)
            full, key = full.strip(), key.strip()
            if full and key:
                out[full] = key
    return out


def build_catalog() -> tuple[list[tuple[str, str]], dict[str, str], frozenset[str]]:
    reps = roster_rep_entries()
    all_ordered = reps + [(COMPANY_SHEET, COMPANY_FULL)]
    # Base full-name -> sheet aliases come from the roster entries (e.g. the
    # built-in "Paul Perlman" -> "Paul"). EXTRA_NAME_ALIASES and the optional
    # SALESPERSON_ALIASES env var extend/override them (env wins).
    full_to_sheet = {full: sheet for sheet, full in reps}
    full_to_sheet.update(EXTRA_NAME_ALIASES)
    full_to_sheet.update(_parse_aliases_env())
    _tpl = globals().get("_TPL_PEOPLE")
    if _tpl and _tpl.get("aliases"):
        full_to_sheet.update(_tpl["aliases"])
    roster_sheets = frozenset(s for s, _ in reps)
    return all_ordered, full_to_sheet, roster_sheets


_TEMPLATE_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "templates" / "master_template_clean.xlsx"


def _load_people_config_from_template(path: Path = _TEMPLATE_PATH) -> dict[str, Any] | None:
    """Business-maintained people config from the template's 'Config_People' sheet.

    Columns: Full Name | Role(rep/company/executive/inactive/b2c/alias) | Sheet Key | Pay Type.
    Returns parsed buckets, or None if the sheet/file is missing or unreadable, in
    which case callers fall back to the in-code defaults / .env (nothing breaks).
    """
    try:
        if os.environ.get("COMMISSION_NO_TEMPLATE_CONFIG", "").strip():
            return None      # kill-switch: ignore template config, use in-code / .env
        from openpyxl import load_workbook
        if not Path(path).exists():
            return None
        wb = load_workbook(path, data_only=True, read_only=True)
        if "Config_People" not in wb.sheetnames:
            wb.close()
            return None
        ws = wb["Config_People"]
        reps: list[tuple[str, str]] = []
        company: list[str] = []
        executive: list[str] = []
        non_sal: set[str] = set()
        inactive: set[str] = set()
        b2c: set[str] = set()
        aliases: dict[str, str] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            full = str(row[0]).strip()
            role = str(row[1] or "").strip().lower()
            sheet = str(row[2] or "").strip() if len(row) > 2 else ""
            pay = str(row[3] or "").strip().lower() if len(row) > 3 else ""
            if not full or not role:
                continue
            if role == "rep" and sheet:
                reps.append((sheet, full))
                if pay == "non_salaried":
                    non_sal.add(sheet)
            elif role == "company":
                company.append(full)
                if sheet:
                    aliases[full] = sheet
            elif role == "executive":
                executive.append(full)
            elif role == "inactive":
                inactive.add(full)
            elif role == "b2c":
                b2c.add(full.lower())
            elif role == "alias" and sheet:
                aliases[full] = sheet
        wb.close()
        if not reps:                      # empty / garbled -> use code defaults
            return None
        return {
            "reps": reps, "non_salaried": non_sal, "company": company,
            "executive": executive, "inactive": inactive, "b2c": b2c, "aliases": aliases,
        }
    except Exception:
        return None


# Business config (the template's Config_People sheet) overrides the in-code
# defaults / .env when present. Falls back silently to the in-code values.
_TPL_PEOPLE = _load_people_config_from_template()
if _TPL_PEOPLE:
    _TPL_REPS: list[tuple[str, str]] | None = _TPL_PEOPLE["reps"]
    NON_SALARIED_SHEETS = frozenset(_TPL_PEOPLE["non_salaried"])
    KNOWN_INACTIVE_NAMES = frozenset(_TPL_PEOPLE["inactive"])
    B2C_COUPON_REPS = frozenset(_TPL_PEOPLE["b2c"])
    COMPANY_ACCOUNT_NAMES = list(_TPL_PEOPLE["company"])
    EXECUTIVE_ACCOUNT_NAMES = list(_TPL_PEOPLE["executive"])
    SPECIAL_PERSON_ROUTING = _build_special_routing()
else:
    _TPL_REPS = None


ALL_SHEETS_ORDERED, SALESPERSON_FULL_TO_SHEET, ROSTER_SHEETS = build_catalog()


def roster_rep_sheet_keys() -> list[str]:
    return [s for s, _ in roster_rep_entries()]


def format_zoho_salesperson(raw: str | None) -> str:
    text = (raw or "").strip()
    return text if text else MISSING_ZOHO_LABEL


def classify_special_person(name: str | None) -> dict[str, str] | None:
    """Return special routing rules if the name is Bruce, Marshall, etc. else None."""
    if not name:
        return None
    return SPECIAL_PERSON_ROUTING.get(str(name).strip().lower())


def is_known_inactive(name: str | None) -> bool:
    """True if the Zoho name is a known-inactive or non-B2B person."""
    if not name:
        return False
    n = str(name).strip()
    return n in KNOWN_INACTIVE_NAMES or n.lower() in {x.lower() for x in KNOWN_INACTIVE_NAMES}


def is_b2c_coupon_rep(name: str | None) -> bool:
    """True if the Zoho salesperson requires coupon-based B2C commission rules."""
    if not name:
        return False
    return str(name).strip().lower() in B2C_COUPON_REPS


def resolve_roster_sheet(name: str | None) -> str | None:
    """Map a Zoho full name or sheet key to a roster sheet key, or None if not on roster."""
    if not name:
        return None
    n = str(name).strip()
    if n in SALESPERSON_FULL_TO_SHEET:
        sheet = SALESPERSON_FULL_TO_SHEET[n]
        return sheet if sheet in ROSTER_SHEETS else None
    if n in ROSTER_SHEETS:
        return n
    low = n.lower()
    for sheet, _full in roster_rep_entries():
        if sheet.lower() == low:
            return sheet
    return None


def accounting_category(classification: str) -> str:
    cls = (classification or "").lower()
    if cls == "company":
        return "Company Account"
    if cls == "executive":
        return "Executive Account"
    return ""


def final_commission_assignment(
    *,
    excluded: bool,
    classification: str,
    pending: bool,
    sheet: str,
) -> str:
    if excluded:
        return "Excluded"
    cat = accounting_category(classification)
    if cat:
        return cat
    if pending or sheet == ROUTING_UNASSIGNED:
        return "Pending"
    return sheet


def issue_found(row: dict[str, Any]) -> str:
    if row.get("issue_found"):
        return str(row["issue_found"])
    flags = str(row.get("flags") or "")

    # Special persons — highest priority
    if "COMPANY_ACCOUNT" in flags:
        return "Company Account / Bruce special commission"
    if "EXECUTIVE_ACCOUNT" in flags:
        return "Executive Account / no commission payable"

    # Inactive / non-B2B names
    if "KNOWN_INACTIVE" in flags:
        return "Salesperson is inactive or non-B2B — requires assignment or exclusion"

    # B2C coupon rule
    if "B2C_COUPON_RULE" in flags:
        return "B2C / coupon-based commission rule — verify coupon before including in B2B payable"

    # Ticket number — classified in ticket_classification.py
    if "REAL_TICKET" in flags:
        return "Real support ticket present — non-commissionable"
    if "QUOTE_REFERENCE_IN_TICKET_FIELD" in flags:
        return "Quote reference in Ticket# field — not automatically excluded"
    if "OTHER_TICKET_REFERENCE" in flags:
        return "Unrecognized Ticket# format — review required"

    # Possible ticket via price anomaly (cf_ticket may be empty)
    if "PRICE_ANOMALY" in flags:
        return "Possible ticket / price anomaly — invoiced far above MAP"

    # Lower-side MAP anomaly: resolved snapshot/R_LP MAP suspiciously below items.rate
    if "MAP_ANOMALY_LOW" in flags:
        return "Resolved MAP unusually low vs. live item rate — possible snapshot typo"

    # No sale date to apply a snapshot
    if "MISSING_SALE_DATE" in flags:
        return "No SO/Invoice/Shipment date — snapshot MAP could not be applied"

    # SKU has price_history rows but none cover the sale date
    if "PRICE_HISTORY_NO_WINDOW" in flags:
        return "SKU has snapshot rows but none cover the sale date — used fallback MAP"

    # Discount above the commission-table limit
    if "DISCOUNT_OVER_60" in flags:
        return "Discount over 60% — non-commissionable"
    if "DISCOUNT_OVER_30" in flags:
        return "Discount above commission table limit"

    # Adjustment salesperson not resolved
    if "ADJ_SALESPERSON_NOT_IN_ROSTER" in flags:
        return "Adjusted salesperson name not found in roster — check spelling"

    # Negative balance
    if "NEGATIVE_BALANCE" in flags:
        return "Negative balance (credit / over-payment) — verify before payout"

    # Returns
    if "FULLY_RETURNED" in flags:
        return "Fully returned — not commissionable"
    if "PARTIALLY_RETURNED" in flags:
        return "Partially returned"

    team = str(row.get("sales_team") or "").lower()
    zoho = str(row.get("original_zoho_salesperson") or "")
    if row.get("pending"):
        if zoho == MISSING_ZOHO_LABEL:
            return ISSUE_MISSING_ZOHO
        if "UNASSIGNED" in flags and zoho not in ("", MISSING_ZOHO_LABEL):
            return ISSUE_NOT_IN_ROSTER
        if "exe" in team or "comp" in team:
            return "Company / Executive account needs classification"
        if "UNASSIGNED" in flags:
            return ISSUE_NOT_IN_ROSTER
        return ISSUE_MISSING_ZOHO
    if "MISSING_MAP" in flags:
        return "MAP / discount difference"
    if "UNPAID" in flags:
        return "Invoice not paid yet"
    if row.get("block") == "shipping":
        return "Shipping line"
    if row.get("section") == "II":
        return "Prior-period order"
    return ""


def suggested_action(row: dict[str, Any]) -> str:
    if row.get("suggested_action"):
        return str(row["suggested_action"])
    flags = str(row.get("flags") or "")

    # Special persons
    if "COMPANY_ACCOUNT" in flags:
        return "Review only if exception"
    if "EXECUTIVE_ACCOUNT" in flags:
        return "Track revenue, no commission unless exception approved"

    # Inactive names
    if "KNOWN_INACTIVE" in flags:
        return "Assign to active salesperson, classify as Company/Executive, or exclude"

    # B2C coupon
    if "B2C_COUPON_RULE" in flags:
        return "Review coupon on order — B2C-RC Team coupon = commissionable; B2C-Web Marketing = not commissionable"

    # Ticket number
    if "REAL_TICKET" in flags:
        return "Exclude — real support/warranty ticket (numeric 1–4 digits)"
    if "QUOTE_REFERENCE_IN_TICKET_FIELD" in flags:
        return "No action required — quote reference is not a support ticket"
    if "OTHER_TICKET_REFERENCE" in flags:
        return "Review Ticket# format — classify, exclude, or approve manually"

    # Possible ticket via price anomaly
    if "PRICE_ANOMALY" in flags:
        return "Verify invoice — likely a ticket/keying error; exclude if it is a ticket, else approve"

    # Lower-side MAP anomaly
    if "MAP_ANOMALY_LOW" in flags:
        return "Verify the snapshot price for this SKU; if wrong, correct price_history and re-run"

    # No sale date for snapshot
    if "MISSING_SALE_DATE" in flags:
        return "Resolve SO/Invoice/Shipment date or accept fallback MAP; approve manually"

    # snapshot rows exist but no covering window
    if "PRICE_HISTORY_NO_WINDOW" in flags:
        return "Verify fallback MAP for this sale date; consider loading a covering snapshot"

    # Discount above the commission-table limit
    if "DISCOUNT_OVER_60" in flags:
        return "Non-commissionable — do not pay unless management approves"
    if "DISCOUNT_OVER_30" in flags:
        return "Confirm written approval and applicable commission rate"

    # Adjustment salesperson typo
    if "ADJ_SALESPERSON_NOT_IN_ROSTER" in flags:
        return "Fix adjusted salesperson name to match an active roster entry"

    # Negative balance
    if "NEGATIVE_BALANCE" in flags:
        return "Confirm payment status — negative balance may indicate a credit note"

    # Returns
    if "FULLY_RETURNED" in flags:
        return "Returned — verify $0 commission"
    if "PARTIALLY_RETURNED" in flags:
        return "Partial return — verify kept qty"

    if row.get("pending"):
        zoho = str(row.get("original_zoho_salesperson") or "")
        if zoho == MISSING_ZOHO_LABEL:
            return ACTION_MISSING_ZOHO
        if "UNASSIGNED" in flags and zoho not in ("", MISSING_ZOHO_LABEL):
            return ACTION_NOT_IN_ROSTER
        team = str(row.get("sales_team") or "").lower()
        if "exe" in team or "comp" in team:
            return "Classify as Company / Executive"
        if "UNASSIGNED" in flags:
            return ACTION_NOT_IN_ROSTER
        return ACTION_MISSING_ZOHO
    if "MISSING_MAP" in flags:
        return "Review MAP / discount"
    if "UNPAID" in flags:
        return "Confirm payment, then approve"
    if row.get("excluded"):
        return "Review exclusion"
    if (row.get("approval_status") or "").lower() == "approved":
        return "—"
    return "Approve if correct"


def enrich_audit_fields(
    *,
    zoho_salesperson: str,
    sys_sheet: str,
    sheet: str,
    excluded: bool,
    classification: str,
    pending: bool,
    flags: str,
    block: str,
    section: str,
    sales_team: str,
    approval_status: str,
) -> dict[str, str]:
    """Build display/audit columns shared by API, UI, and workbook export."""
    final_asgn = final_commission_assignment(
        excluded=excluded,
        classification=classification,
        pending=pending,
        sheet=sheet,
    )
    row = {
        "original_zoho_salesperson": zoho_salesperson,
        "final_commission_assignment": final_asgn,
        "accounting_category": accounting_category(classification),
        "salesperson": final_asgn,
        "system_salesperson": sys_sheet,
        "flags": flags,
        "pending": pending,
        "excluded": excluded,
        "classification": classification,
        "block": block,
        "section": section,
        "sales_team": sales_team,
        "approval_status": approval_status,
    }
    return {
        **row,
        "issue_found": issue_found(row),
        "suggested_action": suggested_action(row),
    }
