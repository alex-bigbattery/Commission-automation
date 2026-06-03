"""
Commission workbook builder v2 — template-based.

Strategy:
  1. Load a clean template (data/templates/salesperson_template_clean.xlsx)
     that preserves ALL styles, fonts, borders, fills, hidden columns, etc.
  2. For each detail block, ensure enough reserved rows for the actual data.
     Insert rows (copying formulas/styles from the previous row) if needed.
  3. Fill data into the rows.
  4. Recompute and rewrite formulas at the new dynamic anchor positions
     (subtotal rows + summary cells at top of sheet).
  5. Save the result.

This avoids the openpyxl limitation that `insert_rows()` does NOT update
existing formula references — we recompute anchor-dependent formulas ourselves.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- Data models -------------------------------------------------------------


@dataclass
class DetailRow:
    """One line of detail. Maps to columns D..AH (data columns)."""

    order_date: Any = None
    sales_order_number: str = ""
    invoice_date: Any = None
    invoice_number: str = ""
    invoice_status: str = ""
    customer_name: str = ""
    estimate_number: str = ""
    sku: str = ""
    quantity: float = 0.0
    item_total: float = 0.0
    account: str = ""
    account_code: str = ""
    payment_terms: str = ""
    delivery_method: str = ""
    shipment_date: Any = None
    shipment_status: str = ""
    ar_status: str = ""
    payment_date: Any = None
    shipping_method: str = ""
    reason: str = ""
    include_in_commission: str = "No"
    shipping_income: float = 0.0
    shipping_expenses: float = 0.0
    discount_rate: float | None = None
    # Computed in the bridge (Python) and written as VALUES so the workbook does
    # not depend on Excel evaluating VLOOKUPs against R_LP / Table.
    map_price: float = 0.0          # MAP unit price (col AE)
    commission_rate: float = 0.0    # commission rate for this line (col AK)


@dataclass
class Block:
    label: str
    rows: list[DetailRow] = field(default_factory=list)


@dataclass
class SalespersonData:
    name: str                   # sheet name (e.g. "Brett")
    full_name: str              # display (e.g. "BRETT BERN")
    rate_type: str              # "salaried" | "non_salaried"
    month_name: str             # e.g. "March"
    year: int
    current_commissionable: Block
    current_shipping: Block
    current_other: Block
    prior_commissionable: Block
    prior_shipping: Block


# Layout of the clean template (Brett-derived). These are the row positions
# of each block's header and subtotal AS THEY EXIST IN THE TEMPLATE FILE.
@dataclass(frozen=True)
class TemplateBlock:
    name: str
    header_row: int
    subtotal_row: int

    @property
    def first_data_row(self) -> int:
        return self.header_row + 1

    @property
    def last_data_row(self) -> int:
        return self.subtotal_row - 1

    @property
    def capacity(self) -> int:
        return self.last_data_row - self.first_data_row + 1


# Default template anchors (matches the clean Brett-derived template).
DEFAULT_TEMPLATE_LAYOUT = [
    TemplateBlock("I.1", header_row=23, subtotal_row=45),
    TemplateBlock("I.2", header_row=48, subtotal_row=53),
    TemplateBlock("I.3", header_row=56, subtotal_row=59),
    TemplateBlock("II.1", header_row=64, subtotal_row=66),
    TemplateBlock("II.2", header_row=69, subtotal_row=71),
]


# --- Column helpers ----------------------------------------------------------


def col(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


DATA_COLS_MAP = {
    "order_date": "D",
    "sales_order_number": "E",
    "invoice_date": "F",
    "invoice_number": "G",
    "invoice_status": "H",
    "customer_name": "I",
    "estimate_number": "J",
    "sku": "K",
    "quantity": "L",
    "item_total": "M",
    "account": "N",
    "account_code": "O",
    "payment_terms": "P",
    "delivery_method": "Q",
    "shipment_date": "S",
    "shipment_status": "T",
    "ar_status": "U",
    "payment_date": "V",
    "shipping_method": "X",
    "reason": "Y",
    "include_in_commission": "Z",
    "shipping_income": "AA",
    "shipping_expenses": "AB",
}

# Subtotal columns
SUBTOTAL_SUM_COLS = ("M", "AA", "AB", "AC", "AF", "AG", "AJ", "AL")
SUBTOTAL_WEIGHTED = {
    "AH": "AG",
    "AK": "AJ",
}


# --- Helpers for row insertion with style copy -------------------------------


def _copy_cell_format(src_cell, dst_cell) -> None:
    """Copy style from src to dst (font, fill, border, alignment, number_format)."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _replicate_row_formulas_and_format(
    ws: Worksheet,
    template_row: int,
    target_row: int,
    max_col: int = 38,
) -> None:
    """Copy formulas (with row reference adjusted) and styles from template_row to target_row."""
    for c in range(1, max_col + 1):
        src = ws.cell(template_row, c)
        dst = ws.cell(target_row, c)
        _copy_cell_format(src, dst)
        v = src.value
        if isinstance(v, str) and v.startswith("="):
            # Adjust row references in the formula to point to target_row
            adjusted = _shift_row_refs(v, template_row, target_row)
            dst.value = adjusted


def _shift_row_refs(formula: str, from_row: int, to_row: int) -> str:
    """
    Shift relative-row references in a formula by (to_row - from_row).
    Naively transforms 'AE24' -> 'AE25' etc. for direct cell refs without `$`.
    """
    import re

    delta = to_row - from_row
    if delta == 0:
        return formula

    def _replace(match: re.Match[str]) -> str:
        col_part = match.group(1)
        dollar = match.group(2) or ""
        row_part = int(match.group(3))
        if dollar == "$":
            return match.group(0)
        # Only shift refs that match from_row
        if row_part == from_row:
            return f"{col_part}{dollar}{row_part + delta}"
        return match.group(0)

    # Pattern: optional $ col letters + optional $ + row number
    pattern = re.compile(r"(\$?[A-Z]+)(\$?)(\d+)")
    return pattern.sub(_replace, formula)


# --- Block manipulation ------------------------------------------------------


def _ensure_block_capacity(
    ws: Worksheet,
    block: TemplateBlock,
    needed_rows: int,
    *,
    row_offset_so_far: int = 0,
) -> tuple[TemplateBlock, int]:
    """
    Make sure `block` can hold `needed_rows` data rows. If not, insert additional
    rows just above the subtotal, copying formulas/styles from the last existing
    data row.

    `row_offset_so_far` accumulates the total inserts from previous blocks so
    we can shift this block's position accordingly.

    Returns (updated_block_with_shifted_positions, new_row_offset).
    """
    shifted = TemplateBlock(
        name=block.name,
        header_row=block.header_row + row_offset_so_far,
        subtotal_row=block.subtotal_row + row_offset_so_far,
    )

    if needed_rows <= shifted.capacity:
        return shifted, row_offset_so_far

    extra = needed_rows - shifted.capacity
    insert_at = shifted.subtotal_row  # insert just above subtotal
    template_row = shifted.last_data_row  # row to copy format from

    ws.insert_rows(insert_at, amount=extra)

    # openpyxl insert_rows does NOT copy formulas/styles to new rows.
    # We manually replicate from `template_row` (which is now also shifted by `extra`? No —
    # insert_rows shifts EXISTING rows down, so the row that was at `template_row`
    # is still at the SAME position because we inserted AFTER it... wait, no:
    # insert_rows(idx) inserts BEFORE row idx, so rows >= idx get shifted by +extra.
    # template_row < insert_at, so template_row is unchanged.
    for i in range(extra):
        new_row = template_row + 1 + i
        _replicate_row_formulas_and_format(ws, template_row, new_row)
        # Clear data values in the new row (only formulas/styles inherited)
        for data_col in ("D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
                         "N", "O", "P", "Q", "S", "T", "U", "V", "X", "Y",
                         "Z", "AA", "AB", "AH"):
            c = ws.cell(new_row, col(data_col))
            if not (isinstance(c.value, str) and c.value.startswith("=")):
                c.value = None

    new_subtotal_row = shifted.subtotal_row + extra
    updated = TemplateBlock(
        name=block.name,
        header_row=shifted.header_row,
        subtotal_row=new_subtotal_row,
    )
    return updated, row_offset_so_far + extra


def _write_row_calc_formulas(ws: Worksheet, r: int, *, map_price: float, commission_rate: float) -> None:
    """
    Write the per-row calculation cells.

    MAP price (AE) and commission rate (AK) are written as VALUES (computed in
    Python) so the result does not depend on Excel evaluating VLOOKUPs. The
    arithmetic columns stay as formulas so block subtotals and the B2B Summary
    keep recomputing if a reviewer tweaks a number.
        AC = AA - AB                 (shipping net)
        AE = map_price               (value)
        AF = AE * L                  (total MAP)
        AG = M                       (revenue)
        AH = 1 - AG/AF               (implied discount rate, for audit)
        AJ = IF(Z="Yes", AG+AC, AG)  (commissionable amount)
        AK = commission_rate         (value)
        AL = AJ * AK                 (commission amount)
    """
    ws.cell(r, col("AC"), f"=AA{r}-AB{r}")
    ws.cell(r, col("AE"), float(map_price or 0))
    ws.cell(r, col("AF"), f"=AE{r}*L{r}")
    ws.cell(r, col("AG"), f"=M{r}")
    ws.cell(r, col("AH"), f"=IFERROR(1-(AG{r}/AF{r}),0)")
    ws.cell(r, col("AJ"), f'=IF(Z{r}="Yes",AG{r}+AC{r},AG{r})')
    ws.cell(r, col("AK"), float(commission_rate or 0))
    ws.cell(r, col("AL"), f"=AJ{r}*AK{r}")


def _write_detail_row(ws: Worksheet, r: int, d: DetailRow) -> None:
    """Write one populated data row (data values + calc cells)."""
    for attr, col_letter in DATA_COLS_MAP.items():
        val = getattr(d, attr)
        if attr in ("quantity", "item_total", "shipping_income", "shipping_expenses"):
            val = float(val or 0)
        if attr == "include_in_commission" and not val:
            val = "No"
        ws.cell(r, col(col_letter)).value = val
    _write_row_calc_formulas(ws, r, map_price=d.map_price, commission_rate=d.commission_rate)


def _write_empty_row(ws: Worksheet, r: int) -> None:
    """Blank a reserved row's data and (re)set its calc formulas so it stays consistent."""
    for col_letter in (
        "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
        "S", "T", "U", "V", "X", "Y", "Z", "AA", "AB",
    ):
        ws.cell(r, col(col_letter)).value = None
    _write_row_calc_formulas(ws, r, map_price=0, commission_rate=0)


def _clear_unused_data_rows(ws: Worksheet, first_unused: int, subtotal_row: int) -> None:
    """For reserved-but-unused rows, ensure formulas are present but data is blank."""
    for r in range(first_unused, subtotal_row):
        for data_col in ("D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
                         "N", "O", "P", "Q", "S", "T", "U", "V", "X", "Y",
                         "Z", "AA", "AB", "AH"):
            c = ws.cell(r, col(data_col))
            if not (isinstance(c.value, str) and c.value.startswith("=")):
                c.value = None


def _write_subtotal_formulas(ws: Worksheet, block: TemplateBlock) -> None:
    """Rewrite SUM and weighted-average formulas at the subtotal row for the current data range."""
    first = block.first_data_row
    last = block.last_data_row
    sr = block.subtotal_row
    ws.cell(sr, col("K"), "Subtotal")
    for c in SUBTOTAL_SUM_COLS:
        ws.cell(sr, col(c), f"=SUM({c}{first}:{c}{last})")
    for value_col, weight_col in SUBTOTAL_WEIGHTED.items():
        formula = (
            f"=IFERROR(SUMPRODUCT(${weight_col}${first}:${weight_col}${last},"
            f"{value_col}{first}:{value_col}{last})/SUM(${weight_col}${first}:${weight_col}${last}),0)"
        )
        ws.cell(sr, col(value_col), formula)


def _write_top_summary_formulas(ws: Worksheet, anchors: list[TemplateBlock]) -> None:
    """Rewrite the top-of-sheet summary cells with formulas pointing to current anchors.

    anchors order: I.1, I.2, I.3, II.1, II.2
    """
    i1, i2, i3, ii1, ii2 = anchors

    # Row 7: Shipped and Invoiced (current period)
    sr = i1.subtotal_row
    ws.cell(7, col("E"), f"=AF{sr}")
    ws.cell(7, col("F"), f"=AG{sr}")
    ws.cell(7, col("G"), f"=AH{sr}")
    ws.cell(7, col("H"), f"=AJ{sr}")
    ws.cell(7, col("I"), f"=AK{sr}")
    ws.cell(7, col("J"), f"=AL{sr}")

    # Row 10: Other Income (current period) = sum of shipping + other_charges subtotals (col M)
    ws.cell(10, col("F"), f"=M{i2.subtotal_row}+M{i3.subtotal_row}")

    # Row 14: Prior period shipped and invoiced
    psr = ii1.subtotal_row
    ws.cell(14, col("E"), f"=AF{psr}")
    ws.cell(14, col("F"), f"=AG{psr}")
    ws.cell(14, col("G"), f"=AH{psr}")
    ws.cell(14, col("H"), f"=AJ{psr}")
    ws.cell(14, col("I"), f"=AK{psr}")
    ws.cell(14, col("J"), f"=AL{psr}")

    # Row 17: Prior period Other Income (just shipping for now)
    ws.cell(17, col("F"), f"=M{ii2.subtotal_row}")


# --- Main entry point --------------------------------------------------------


def _update_salesperson_sheet_in_place(ws: Worksheet, data: SalespersonData) -> None:
    """
    Modify an existing salesperson sheet (in a loaded workbook) with the given data.

    Assumes the sheet starts with the canonical Brett-derived layout
    (DEFAULT_TEMPLATE_LAYOUT). Inserts rows as needed, fills data, recomputes
    subtotal and top-summary formulas.
    """
    # Title cells
    ws.cell(2, col("D"), data.full_name)
    ws.cell(3, col("D"), f"COMMISSION REPORT - {data.month_name.upper()} {data.year}")
    rate_label = "Non-Salary Commission Rate" if data.rate_type == "non_salaried" else "Salary Commission Rate"
    ws.cell(4, col("D"), rate_label)
    ws.cell(20, col("C"), f"SALES ORDERS INVOICED IN {data.month_name.upper()} {data.year}")

    blocks_in_order = [
        ("I.1", data.current_commissionable),
        ("I.2", data.current_shipping),
        ("I.3", data.current_other),
        ("II.1", data.prior_commissionable),
        ("II.2", data.prior_shipping),
    ]

    updated_anchors: list[TemplateBlock] = []
    row_offset = 0
    for tmpl, (_name, block_data) in zip(DEFAULT_TEMPLATE_LAYOUT, blocks_in_order):
        needed = max(1, len(block_data.rows))
        updated, row_offset = _ensure_block_capacity(ws, tmpl, needed, row_offset_so_far=row_offset)
        for i, drow in enumerate(block_data.rows):
            target_row = updated.first_data_row + i
            _write_detail_row(ws, target_row, drow)
        first_unused = updated.first_data_row + len(block_data.rows)
        for r in range(first_unused, updated.subtotal_row):
            _write_empty_row(ws, r)
        _write_subtotal_formulas(ws, updated)
        updated_anchors.append(updated)

    _write_top_summary_formulas(ws, updated_anchors)


def build_salesperson_workbook(
    template_path: Path,
    output_path: Path,
    data: SalespersonData,
) -> Path:
    """Build a workbook for ONE salesperson (single-sheet output)."""
    wb = load_workbook(template_path)
    ws = wb["Brett"]
    ws.title = data.name
    _update_salesperson_sheet_in_place(ws, data)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _reorder_master_sheets(wb) -> None:
    """Reorder sheets to: B2B Summary, salespeople (canonical order), Table, reference."""
    desired_order = [
        "B2B Summary",
        "Paul", "Jose", "Michael", "Jim", "Weston",
        "Brett", "Leslie", "Carmen", "Garrett",
        "Company Acct",
        "Table", "R_LP", "R_SO", "R_INV", "R_SH",
    ]
    existing = list(wb.sheetnames)
    ordered = [name for name in desired_order if name in existing]
    leftover = [name for name in existing if name not in desired_order]
    final = ordered + leftover
    # Reorder by moving each sheet to its target position
    wb._sheets = [wb[name] for name in final]


def _write_adjustments_audit_sheet(wb, audit_rows: list[dict]) -> None:
    """Append an 'Adjustments Audit' sheet showing system / adjustment / final per line."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    name = "Adjustments Audit"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    from src.commission.roster import issue_found as roster_issue, suggested_action as roster_action

    headers = [
        ("Pending", "pending"),
        ("Original Zoho Salesperson", "original_zoho_salesperson"),
        ("Final Commission Assignment", "final_commission_assignment"),
        ("Accounting Category", "accounting_category"),
        ("Issue Found", "issue_found"),
        ("Suggested Action", "__action"),
        ("System Salesperson", "system_salesperson"),
        ("Sales Team", "sales_team"),
        ("Sales Order", "sales_order"), ("Invoice", "invoice"), ("SKU", "sku"),
        ("Qty Ordered", "qty_ordered"), ("Qty Shipped", "qty_shipped"),
        ("Qty Invoiced", "qty_invoiced"), ("Qty Returned", "qty_returned"),
        ("Qty Commissionable", "qty_commissionable"), ("Return Status", "return_status"),
        ("Revenue", "revenue"),
        ("System Commissionable", "system_commissionable"), ("System Rate", "system_rate"),
        ("System Amount", "system_commission"),
        ("Final Commissionable", "final_commissionable"), ("Final Rate", "final_rate"),
        ("Final Amount", "final_commission"), ("Adjustment", "adjustment"),
        ("Excluded", "excluded"), ("Classification", "classification"),
        ("Flags", "flags"), ("Reason", "reason"), ("Reviewer", "reviewer"), ("Approval", "approval_status"),
    ]

    def suggested_action(row: dict) -> str:
        return roster_action(row) or ("" if not row.get("adjusted") else "Adjusted — review")

    # Sort so the rows needing attention float to the top.
    def sort_key(row: dict):
        return (
            0 if row.get("pending") else 1,
            0 if str(row.get("flags") or "") else 1,
            0 if row.get("adjusted") else 1,
            str(row.get("final_commission_assignment") or row.get("salesperson") or ""),
        )
    ordered = sorted(audit_rows, key=sort_key)

    # Backfill issue/suggested for rows generated before enrich_audit_fields existed.
    for row in ordered:
        if not row.get("issue_found"):
            row["issue_found"] = roster_issue(row)
        if not row.get("suggested_action"):
            row["suggested_action"] = roster_action(row)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for c, (label, _key) in enumerate(headers, start=1):
        cell = ws.cell(1, c, label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = max(11, min(len(label) + 2, 34))

    money = {"revenue", "system_commissionable", "system_commission",
             "final_commissionable", "final_commission", "adjustment"}
    pct = {"system_rate", "final_rate"}
    warn_fill = PatternFill("solid", fgColor="FCE8E6")
    for r, row in enumerate(ordered, start=2):
        for c, (_label, key) in enumerate(headers, start=1):
            if key == "__action":
                v = suggested_action(row)
            elif key == "issue_found":
                v = row.get("issue_found") or roster_issue(row)
            else:
                v = row.get(key)
            if isinstance(v, bool):
                v = "Yes" if v else ""
            cell = ws.cell(r, c, v)
            if key in money:
                cell.number_format = "$#,##0.00"
            elif key in pct:
                cell.number_format = "0.00%"
            if row.get("pending"):
                cell.fill = warn_fill
    ws.freeze_panes = "C2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions


RECON_REPS = ["Paul", "Jose", "Michael", "Jim", "Weston", "Brett", "Leslie", "Carmen", "Garrett"]


def _clean_summary_formulas(ws) -> None:
    """Remove the inherited Executive Account hardcode and fix the prior-period
    non-salaried aggregation that only summed Carmen (O32) in the old template."""
    from openpyxl.utils import get_column_letter
    # Executive Account is a manual line — never carry a hardcoded number.
    ws["F38"] = 0
    ws["G38"] = 0
    # Prior-period non-salaried totals: sum ALL non-salaried reps (rows 30-33), not just row 32.
    for c in ("O", "Q", "S", "U"):
        ws[f"{c}34"] = f"=SUM({c}30:{c}33)"
    # Weighted-average prior discount / commission % (mirror the current-period row 27 pattern).
    ws["P34"] = "=IFERROR(SUMPRODUCT($O$30:$O$33,P30:P33)/SUM($O$30:$O$33),0)"
    ws["R34"] = "=IFERROR(S34/Q34,0)"


def _write_status_banner(ws, status_info: dict, month_name: str, year: int) -> None:
    """Draft/Final status banner (row 1) + a status detail block below the summary."""
    from openpyxl.styles import Alignment, Font, PatternFill

    info = status_info or {}
    is_draft = bool(info.get("is_draft"))
    pending = int(info.get("pending_lines", 0) or 0)
    label = ("DRAFT — NEEDS ACCOUNTING REVIEW" if is_draft else "FINAL — APPROVED")
    banner = f"{label}   ·   {month_name} {year}"
    if is_draft:
        banner += f"   ·   {pending} pending line(s) to resolve"

    cell = ws.cell(1, 4, banner)  # D1
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=("C0392B" if is_draft else "1E8449"))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    try:
        ws.merge_cells("D1:N1")
        ws.row_dimensions[1].height = 22
    except Exception:
        pass

    # Prominent payment warning while draft.
    if is_draft:
        warn = ws.cell(42, 4, "This workbook is a DRAFT and should not be used for payment until pending "
                              "lines and shipment data are resolved.")
        warn.font = Font(bold=True, italic=True, color="C0392B")

    # Status detail block (below the summary content).
    base = 43
    bold = Font(bold=True)
    ws.cell(base, 4, "GENERATION STATUS").font = Font(bold=True, size=11)
    rows = [
        ("Status", label),
        ("Pending lines", pending),
        ("Pending revenue", round(float(info.get("pending_revenue", 0) or 0), 2)),
        ("Pending est. commission", round(float(info.get("pending_commission", 0) or 0), 2)),
        ("Shipment data synced", "Yes" if info.get("shipment_data_present") else "No — sync shipments before finalizing"),
        ("Adjusted lines", int(info.get("adjusted_lines", 0) or 0)),
    ]
    for i, (k, v) in enumerate(rows, start=base + 1):
        ws.cell(i, 4, k).font = bold
        c = ws.cell(i, 6, v)
        if k in ("Pending revenue", "Pending est. commission"):
            c.number_format = "$#,##0.00"


def _write_reference_sheets(wb, reference_sheets: dict, shipments_present: bool) -> None:
    """Replace the stale template R_SO/R_INV/R_SH/R_LP with period-correct data from SQLite."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    for name, payload in (reference_sheets or {}).items():
        headers, rows = payload
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        if name == "R_SH" and not shipments_present:
            msg = ws.cell(1, 1, "Shipment data not synced for this period")
            msg.font = Font(bold=True, color="C0392B", size=12)
            continue
        if not headers:
            ws.cell(1, 1, f"No {name} data for this period").font = Font(italic=True, color="666666")
            continue
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(c)].width = max(11, min(len(str(h)) + 2, 26))
        for r, row in enumerate(rows, start=2):
            for c, v in enumerate(row, start=1):
                ws.cell(r, c, v)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def _write_reconciliation_sheet(wb, recon: dict) -> None:
    """Reconciliation with engine-computed VALUES (always visible) + live cross-check formulas."""
    from openpyxl.styles import Alignment, Font

    recon = recon or {}
    name = "Reconciliation"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    ws.cell(1, 1, "COMMISSION RECONCILIATION").font = Font(bold=True, size=13)
    ws.cell(2, 2, "Value").font = Font(bold=True)
    ws.cell(2, 3, "Live (Excel)").font = Font(bold=True)

    reps = [s for s in RECON_REPS if s in wb.sheetnames]
    sheets_sum = "+".join(f"({s}!J7+{s}!J14)" for s in reps) or "0"

    def row(r, label, value, formula=None, bold=False):
        ws.cell(r, 1, label).font = Font(bold=bold)
        c = ws.cell(r, 2, value)
        c.number_format = "$#,##0.00"
        if formula is not None:
            f = ws.cell(r, 3, formula)
            f.number_format = "$#,##0.00"

    row(3, "Sum of salesperson sheets (J7 + J14)", recon.get("rep_commission", 0), f"={sheets_sum}")
    row(4, "B2B Sales Rep Commission (Summary I11)", recon.get("rep_commission", 0), "='B2B Summary'!I11")
    row(5, "  Check A: sheets − sales-rep commission (= 0)", recon.get("check_a", 0),
        f"=({sheets_sum})-'B2B Summary'!I11", bold=True)
    row(7, "Company Account Commission (Summary J11)", recon.get("company_commission", 0), "='B2B Summary'!J11")
    row(8, "Bruce Commission (15% rep + 20% company)", recon.get("bruce", 0), "='B2B Summary'!K13")
    row(9, "Executive Account — manual", recon.get("executive", 0), "='B2B Summary'!F38")
    row(11, "Total to Pay", recon.get("total_to_pay", 0), "='B2B Summary'!M10", bold=True)
    row(12, "  Check B: Rep + Company + Bruce − Total to Pay (= 0)", recon.get("check_b", 0),
        "='B2B Summary'!I11+'B2B Summary'!J11+'B2B Summary'!K13-'B2B Summary'!M10", bold=True)

    note = ws.cell(14, 1, "Column B = engine-computed values (always populated). Column C = the same figures "
                          "pulled live from B2B Summary. Both checks must be 0.")
    note.alignment = Alignment(wrap_text=True)
    note.font = Font(italic=True, color="666666")


def build_master_workbook(
    template_path: Path,
    output_path: Path,
    salespeople: list[SalespersonData],
    month_name: str,
    year: int,
    audit_rows: list[dict] | None = None,
    status_info: dict | None = None,
    reconciliation: dict | None = None,
    reference_sheets: dict | None = None,
    shipments_present: bool = False,
) -> Path:
    """
    Build the full master workbook (B2B Summary + all salesperson sheets).

    Per-salesperson sheets reflect FINAL (post-adjustment) values. The B2B Summary
    gets a Draft/Final status banner; the Executive Account hardcode is cleared;
    prior-period non-salaried aggregation is fixed. Raw reference sheets
    (R_SO/R_INV/R_SH/R_LP) are refreshed for the period, and 'Adjustments Audit'
    and 'Reconciliation' sheets are appended.
    """
    wb = load_workbook(template_path)

    if "B2B Summary" in wb.sheetnames:
        summary_ws = wb["B2B Summary"]
        summary_ws.cell(18, 9, f"{month_name.upper()} {year} ORDERS - SHIPPED AND INVOICED")
        _clean_summary_formulas(summary_ws)

    missing = [d.name for d in salespeople if d.name not in wb.sheetnames]
    if missing:
        import logging
        logging.getLogger(__name__).warning(
            "Roster entries not in template (skipped): %s. "
            "Fix COMMISSION_ROSTER env var or add sheets to template.",
            missing,
        )
    for data in salespeople:
        if data.name not in wb.sheetnames:
            continue  # skip roster members without a template sheet
        ws = wb[data.name]
        _update_salesperson_sheet_in_place(ws, data)

    if "B2B Summary" in wb.sheetnames:
        _write_status_banner(wb["B2B Summary"], status_info or {}, month_name, year)

    # Refresh raw reference sheets from SQLite (no stale template snapshots).
    if reference_sheets:
        _write_reference_sheets(wb, reference_sheets, shipments_present)

    _reorder_master_sheets(wb)

    if audit_rows:
        _write_adjustments_audit_sheet(wb, audit_rows)
    _write_reconciliation_sheet(wb, reconciliation or {})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
