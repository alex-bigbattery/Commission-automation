"""Read-only price_history matrix, detail list, and CSV/XLSX export."""
from __future__ import annotations

import csv
import io
from datetime import date as _date, timedelta
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.commission.settings_read import (
    FAR_FUTURE,
    IMPORTED_RLP_CAUTION,
    _detect_coverage_gaps,
    _enrich_price_row,
    _is_open_end,
    _parse_iso_date,
    _source_kind,
    _today_iso,
)
from src.commission.sqlite_to_workbook import load_map_from_template
from src.db.connection import DbConnection, get_connection

Granularity = Literal["daily", "weekly", "monthly"]
ExportMode = Literal["detail", "matrix"]
ExportFormat = Literal["csv", "xlsx"]

MATRIX_PREVIEW_LIMIT = 2000
MATRIX_EXPORT_LIMIT = 2000
DETAIL_PREVIEW_LIMIT = 500
DETAIL_EXPORT_LIMIT = 5000
MAX_DAILY_COLUMNS_UI = 45

LEGEND_ROWS = [
    ("accountant_fvprice_*", "Accountant FV_PRICE snapshot"),
    ("imported_rlp_*", "Imported R_LP snapshot / fallback — not confirmed FV_PRICE"),
    ("zoho_sync_*", "Zoho live sync (forward from sync date)"),
    ("zoho_catalog_snapshot_*", "Zoho catalog snapshot (items.rate backfill)"),
    ("R_LP_template", "Template fallback — not from price_history (only when Include fallback is enabled)"),
    ("(blank)", "No price_history coverage for that date"),
]

RLP_TEMPLATE_SOURCE_TYPE = "rlp_template_fallback"


def default_granularity(from_iso: str, to_iso: str) -> str:
    start = _parse_iso_date(from_iso)
    end = _parse_iso_date(to_iso)
    if not start or not end:
        return "daily"
    days = (end - start).days + 1
    return "daily" if days <= MAX_DAILY_COLUMNS_UI else "monthly"


def build_date_columns(from_iso: str, to_iso: str, granularity: str) -> list[str]:
    start = _parse_iso_date(from_iso)
    end = _parse_iso_date(to_iso)
    if not start or not end or start > end:
        return []
    gran = (granularity or "daily").strip().lower()
    cols: list[str] = []

    if gran == "daily":
        d = start
        while d <= end:
            cols.append(d.isoformat())
            d += timedelta(days=1)
        return cols

    if gran == "weekly":
        d = start
        while d.weekday() != 0:
            d += timedelta(days=1)
        if d > end:
            return [start.isoformat()]
        while d <= end:
            cols.append(d.isoformat())
            d += timedelta(days=7)
        if cols[0] != start.isoformat():
            cols.insert(0, start.isoformat())
        return cols

    if gran == "monthly":
        y, m = start.year, start.month
        while True:
            col = _date(y, m, 1)
            if col > end:
                break
            if col >= start:
                cols.append(col.isoformat())
            if m == 12:
                y, m = y + 1, 1
            else:
                m += 1
        return cols

    raise ValueError(f"Unknown granularity: {granularity!r}")


def _covers_date(effective_from: str, effective_to: str, date_iso: str) -> bool:
    if not effective_from or not date_iso:
        return False
    if date_iso < effective_from:
        return False
    if _is_open_end(effective_to):
        return True
    return date_iso <= effective_to


def resolve_price_for_date(
    history: list[dict[str, Any]],
    date_iso: str,
) -> dict[str, Any] | None:
    """Pick the price_history row covering date_iso (closed-month wins over live)."""
    candidates: list[dict[str, Any]] = []
    for row in history:
        if _covers_date(str(row.get("effective_from") or ""), str(row.get("effective_to") or ""), date_iso):
            candidates.append(row)
    if not candidates:
        return None
    closed = [r for r in candidates if str(r.get("snapshot_month") or "") != "live"]
    pool = closed if closed else candidates
    pool.sort(key=lambda r: (str(r.get("effective_from") or ""), int(r.get("id") or 0)), reverse=True)
    row = pool[0]
    kind = _source_kind(str(row.get("source") or ""), str(row.get("snapshot_month") or ""))
    return {
        "map_price": float(row["map_price"]),
        "source": str(row.get("source") or ""),
        "source_type": kind,
        "snapshot_month": str(row.get("snapshot_month") or ""),
        "source_caution": IMPORTED_RLP_CAUTION if kind == "imported_rlp" else None,
    }


def _fetch_sku_page(
    conn: DbConnection,
    *,
    q: str,
    limit: int,
    offset: int,
) -> tuple[list[str], int]:
    needle = (q or "").strip().upper()
    clauses = ["sku IS NOT NULL", "sku != ''"]
    params: list[Any] = []
    if needle:
        like = f"%{needle}%"
        clauses.append("(UPPER(sku) LIKE ? OR UPPER(COALESCE(item_id, '')) LIKE ?)")
        params.extend([like, like])
    where = " AND ".join(clauses)
    total = int(conn.execute(f"SELECT COUNT(DISTINCT UPPER(sku)) AS c FROM price_history WHERE {where}", tuple(params)).fetchone()["c"])
    rows = conn.execute(
        f"SELECT DISTINCT UPPER(sku) AS sku_u FROM price_history WHERE {where} "
        f"ORDER BY sku_u LIMIT ? OFFSET ?",
        tuple(params + [limit, offset]),
    ).fetchall()
    return [str(r["sku_u"]) for r in rows], total


def _fetch_history_by_skus(conn: DbConnection, sku_list: list[str]) -> dict[str, list[dict[str, Any]]]:
    if not sku_list:
        return {}
    placeholders = ",".join("?" for _ in sku_list)
    rows = conn.execute(
        f"SELECT id, sku, item_id, map_price, effective_from, effective_to, source, snapshot_month, captured_at "
        f"FROM price_history WHERE UPPER(sku) IN ({placeholders}) "
        f"ORDER BY sku, effective_from ASC, id ASC",
        tuple(sku_list),
    ).fetchall()
    grouped: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        key = str(r["sku"]).upper()
        grouped.setdefault(key, []).append({
            "id": int(r["id"]),
            "sku": str(r["sku"] or ""),
            "item_id": r["item_id"],
            "map_price": float(r["map_price"]),
            "effective_from": str(r["effective_from"] or ""),
            "effective_to": str(r["effective_to"] or ""),
            "source": str(r["source"] or ""),
            "snapshot_month": str(r["snapshot_month"] or ""),
            "captured_at": str(r["captured_at"] or ""),
        })
    return grouped


def _fetch_item_meta(conn: DbConnection, sku_list: list[str]) -> dict[str, dict[str, str | None]]:
    if not sku_list:
        return {}
    placeholders = ",".join("?" for _ in sku_list)
    rows = conn.execute(
        f"SELECT sku, item_id, name FROM items WHERE UPPER(sku) IN ({placeholders})",
        tuple(sku_list),
    ).fetchall()
    out: dict[str, dict[str, str | None]] = {}
    for r in rows:
        key = str(r["sku"] or "").strip().upper()
        if key and key not in out:
            out[key] = {
                "item_id": str(r["item_id"] or "") or None,
                "item_name": str(r["name"] or "").strip() or None,
            }
    return out


def _summarize_current(history: list[dict[str, Any]], today: str) -> tuple[float | None, str | None]:
    enriched = [_enrich_price_row(h, today) for h in history]
    active = [r for r in enriched if r.get("is_active_for_today")]
    live = [r for r in active if r.get("is_current_live")]
    pick = live[0] if live else (active[0] if active else (enriched[-1] if enriched else None))
    if not pick:
        return None, None
    return float(pick["map_price"]), str(pick.get("source") or "")


def _gaps_in_range(history: list[dict[str, Any]], from_iso: str, to_iso: str, today: str) -> list[tuple[str, str]]:
    enriched = [_enrich_price_row(h, today) for h in history]
    gaps = _detect_coverage_gaps(enriched)
    start = _parse_iso_date(from_iso)
    end = _parse_iso_date(to_iso)
    if not start or not end:
        return gaps
    clipped: list[tuple[str, str]] = []
    for gf, gt in gaps:
        gs = _parse_iso_date(gf)
        ge = _parse_iso_date(gt)
        if not gs or not ge:
            continue
        clip_from = max(gs, start)
        clip_to = min(ge, end)
        if clip_from <= clip_to:
            clipped.append((clip_from.isoformat(), clip_to.isoformat()))
    return clipped


def _summarize_price_change(prices: dict[str, Any | None], dates: list[str]) -> tuple[bool, str, str]:
    """Distinct MAP levels in range order; label whether price changed across the window."""
    levels: list[float] = []
    for d in dates:
        cell = prices.get(d)
        if not cell:
            continue
        p = float(cell["map_price"])
        if not levels or levels[-1] != p:
            levels.append(p)
    if not levels:
        return False, "No coverage", "—"
    if len(levels) == 1:
        return False, "No change", "No"
    detail = f"${levels[0]:,.2f} → ${levels[-1]:,.2f}"
    return True, detail, f"Yes ({detail})"


def build_matrix_row(
    sku: str,
    history: list[dict[str, Any]],
    dates: list[str],
    *,
    item_id: str | None,
    item_name: str | None,
    today: str,
    from_iso: str,
    to_iso: str,
    include_fallback: bool,
    rlp_map: dict[str, float],
) -> dict[str, Any]:
    current_map, _latest_source = _summarize_current(history, today)
    prices: dict[str, Any] = {}
    for d in dates:
        hit = resolve_price_for_date(history, d)
        if hit:
            prices[d] = hit
        elif include_fallback and sku in rlp_map:
            prices[d] = {
                "map_price": rlp_map[sku],
                "source": "R_LP_template",
                "source_type": RLP_TEMPLATE_SOURCE_TYPE,
                "is_fallback": True,
                "snapshot_month": "",
                "source_caution": "Template R_LP fallback — not from price_history",
            }
        else:
            prices[d] = None

    price_changed, price_change_label, price_changed_display = _summarize_price_change(prices, dates)

    return {
        "sku": sku,
        "item_id": item_id or (history[0]["item_id"] if history else None),
        "item_name": item_name,
        "current_map": current_map,
        "price_changed": price_changed,
        "price_change_label": price_change_label,
        "price_changed_display": price_changed_display,
        "prices": prices,
        "coverage_gaps": [{"from": a, "to": b} for a, b in _gaps_in_range(history, from_iso, to_iso, today)],
    }


def get_price_history_matrix(
    *,
    q: str = "",
    from_date: str | None = None,
    to_date: str | None = None,
    granularity: str | None = None,
    include_fallback: bool = False,
    limit: int = MATRIX_PREVIEW_LIMIT,
    offset: int = 0,
    template_path: Path | None = None,
) -> dict[str, Any]:
    today = _today_iso()
    to_iso = (to_date or today).strip()
    from_iso = (from_date or "2026-03-01").strip()
    gran = (granularity or default_granularity(from_iso, to_iso)).strip().lower()
    if gran not in ("daily", "weekly", "monthly"):
        raise ValueError(f"Invalid granularity: {granularity!r}")

    dates = build_date_columns(from_iso, to_iso, gran)
    safe_limit = max(1, min(int(limit), MATRIX_EXPORT_LIMIT))
    safe_offset = max(0, int(offset))

    conn = get_connection()
    try:
        sku_list, total = _fetch_sku_page(conn, q=q, limit=safe_limit, offset=safe_offset)
        history_by_sku = _fetch_history_by_skus(conn, sku_list)
        item_meta = _fetch_item_meta(conn, sku_list)
    finally:
        conn.close()

    rlp_map = load_map_from_template(template_path) if include_fallback and template_path else {}

    rows = []
    for sku in sku_list:
        hist = history_by_sku.get(sku, [])
        meta = item_meta.get(sku, {})
        rows.append(build_matrix_row(
            sku, hist, dates,
            item_id=meta.get("item_id"),
            item_name=meta.get("item_name"),
            today=today,
            from_iso=from_iso,
            to_iso=to_iso,
            include_fallback=include_fallback,
            rlp_map=rlp_map,
        ))

    warnings: list[str] = []
    if gran == "daily" and len(dates) > MAX_DAILY_COLUMNS_UI:
        warnings.append(
            f"Daily view has {len(dates)} date columns. "
            "Use weekly or monthly granularity for long ranges."
        )

    return {
        "read_only": True,
        "from": from_iso,
        "to": to_iso,
        "granularity": gran,
        "dates": dates,
        "date_count": len(dates),
        "total_skus": total,
        "limit": safe_limit,
        "offset": safe_offset,
        "count": len(rows),
        "include_fallback": include_fallback,
        "warnings": warnings,
        "rows": rows,
    }


def get_price_history_detail_list(
    *,
    q: str = "",
    from_date: str | None = None,
    to_date: str | None = None,
    sku_limit: int = DETAIL_PREVIEW_LIMIT,
    sku_offset: int = 0,
    for_export: bool = False,
) -> dict[str, Any]:
    today = _today_iso()
    to_iso = (to_date or today).strip()
    from_iso = (from_date or "2026-03-01").strip()
    if for_export:
        sku_cap = DETAIL_EXPORT_LIMIT
        sku_off = 0
    else:
        sku_cap = max(1, min(int(sku_limit), DETAIL_PREVIEW_LIMIT))
        sku_off = max(0, int(sku_offset))

    conn = get_connection()
    try:
        sku_list, total_skus = _fetch_sku_page(conn, q=q, limit=sku_cap, offset=sku_off)
        if not sku_list:
            return {
                "read_only": True,
                "from": from_iso,
                "to": to_iso,
                "total_skus": total_skus,
                "total_rows": 0,
                "limit": sku_cap,
                "offset": sku_off,
                "count": 0,
                "rows": [],
            }
        placeholders = ",".join("?" for _ in sku_list)
        rows_raw = conn.execute(
            f"SELECT sku, item_id, map_price, effective_from, effective_to, source, snapshot_month, captured_at "
            f"FROM price_history WHERE UPPER(sku) IN ({placeholders}) "
            f"AND effective_from <= ? AND (effective_to >= ? OR effective_to >= ?) "
            f"ORDER BY sku, effective_from ASC, id ASC",
            tuple(sku_list + [to_iso, from_iso, FAR_FUTURE]),
        ).fetchall()
        item_meta = _fetch_item_meta(conn, sku_list)
    finally:
        conn.close()

    detail_rows: list[dict[str, Any]] = []
    for r in rows_raw:
        sku_u = str(r["sku"]).upper()
        meta = item_meta.get(sku_u, {})
        base = {
            "sku": str(r["sku"] or ""),
            "item_id": r["item_id"] or meta.get("item_id"),
            "item_name": meta.get("item_name"),
            "map_price": float(r["map_price"]),
            "effective_from": str(r["effective_from"] or ""),
            "effective_to": str(r["effective_to"] or ""),
            "source": str(r["source"] or ""),
            "snapshot_month": str(r["snapshot_month"] or ""),
            "captured_at": str(r["captured_at"] or ""),
        }
        enriched = _enrich_price_row(base, today)
        enriched["source_type"] = enriched["source_kind"]
        enriched["warning_caution"] = enriched.get("source_caution") or ""
        detail_rows.append(enriched)

    return {
        "read_only": True,
        "from": from_iso,
        "to": to_iso,
        "total_skus": total_skus,
        "total_rows": len(detail_rows),
        "limit": sku_cap,
        "offset": sku_off,
        "count": len(detail_rows),
        "rows": detail_rows,
    }


def _matrix_csv(payload: dict[str, Any]) -> bytes:
    buf = io.StringIO()
    dates = payload["dates"]
    header = ["SKU", "Item ID", "Item Name", "Current MAP", "Price changed", *dates]
    writer = csv.writer(buf)
    writer.writerow(header)
    for row in payload["rows"]:
        cells = [
            row["sku"],
            row.get("item_id") or "",
            row.get("item_name") or "",
            row.get("current_map") if row.get("current_map") is not None else "",
            row.get("price_changed_display") or row.get("price_change_label") or "",
        ]
        for d in dates:
            cell = row["prices"].get(d)
            cells.append(cell["map_price"] if cell else "")
        writer.writerow(cells)
    return buf.getvalue().encode("utf-8-sig")


def _detail_csv(payload: dict[str, Any]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "SKU", "Item ID", "Item Name", "MAP Price", "Effective From", "Effective To",
        "Source", "Source Type", "Snapshot Month", "Active Today", "Captured At", "Warning / Caution",
    ])
    for row in payload["rows"]:
        writer.writerow([
            row["sku"],
            row.get("item_id") or "",
            row.get("item_name") or "",
            row["map_price"],
            row["effective_from"],
            row.get("effective_to_display") or row["effective_to"],
            row["source"],
            row.get("source_type") or row.get("source_kind") or "",
            row.get("snapshot_month") or "",
            "Yes" if row.get("is_active_for_today") else "",
            row.get("captured_at") or "",
            row.get("warning_caution") or row.get("source_caution") or "",
        ])
    return buf.getvalue().encode("utf-8-sig")


def _append_matrix_cell_sources(ws, payload: dict[str, Any]) -> None:
    dates = payload["dates"]
    headers = ["SKU", "Date", "MAP Price", "Source", "Source Type", "Is Fallback"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in payload["rows"]:
        for d in dates:
            hit = row["prices"].get(d)
            if not hit:
                continue
            is_fb = bool(
                hit.get("is_fallback")
                or hit.get("source_type") == RLP_TEMPLATE_SOURCE_TYPE
                or hit.get("source") == "R_LP_template"
            )
            ws.append([
                row["sku"],
                d,
                hit["map_price"],
                hit.get("source") or "",
                hit.get("source_type") or "",
                "Yes" if is_fb else "",
            ])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 3).number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(22, max(10, len(h) + 2))


def _matrix_xlsx(payload: dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Timeline"
    dates = payload["dates"]
    headers = ["SKU", "Item ID", "Item Name", "Current MAP", "Price change", *dates]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in payload["rows"]:
        line = [
            row["sku"],
            row.get("item_id") or "",
            row.get("item_name") or "",
            row.get("current_map"),
            row.get("price_changed_display") or row.get("price_change_label") or "",
        ]
        for d in dates:
            hit = row["prices"].get(d)
            line.append(hit["map_price"] if hit else None)
        ws.append(line)
    for col_idx in range(4, 5 + len(dates)):
        for r in range(2, ws.max_row + 1):
            c = ws.cell(r, col_idx)
            if c.value is not None:
                c.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(18, max(10, len(str(headers[i - 1])) + 2))

    sources = wb.create_sheet("Cell Sources")
    _append_matrix_cell_sources(sources, payload)

    legend = wb.create_sheet("Legend")
    legend.append(["Source pattern", "Meaning"])
    for cell in legend[1]:
        cell.font = Font(bold=True)
    for pat, meaning in LEGEND_ROWS:
        legend.append([pat, meaning])
    legend.append([])
    legend.append(["Export setting", "Value"])
    legend.append(["Include fallback (R_LP template)", "Yes" if payload.get("include_fallback") else "No"])
    legend.append(["Date range", f"{payload.get('from')} → {payload.get('to')}"])
    legend.append(["Granularity", str(payload.get("granularity") or "")])
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 60

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _detail_xlsx(payload: dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price History Detail"
    headers = [
        "SKU", "Item ID", "Item Name", "MAP Price", "Effective From", "Effective To",
        "Source", "Source Type", "Snapshot Month", "Active Today", "Captured At", "Warning / Caution",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in payload["rows"]:
        ws.append([
            row["sku"],
            row.get("item_id") or "",
            row.get("item_name") or "",
            row["map_price"],
            row["effective_from"],
            row.get("effective_to_display") or row["effective_to"],
            row["source"],
            row.get("source_type") or row.get("source_kind") or "",
            row.get("snapshot_month") or "",
            "Yes" if row.get("is_active_for_today") else "",
            row.get("captured_at") or "",
            row.get("warning_caution") or row.get("source_caution") or "",
        ])
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 4).number_format = '"$"#,##0.00'
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(22, max(10, len(h) + 2))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_price_history_file(
    *,
    mode: ExportMode,
    fmt: ExportFormat,
    q: str = "",
    from_date: str | None = None,
    to_date: str | None = None,
    granularity: str | None = None,
    include_fallback: bool = False,
    template_path: Path | None = None,
) -> tuple[bytes, str, str]:
    """Return (content_bytes, filename, media_type)."""
    stamp = _today_iso().replace("-", "")
    if mode == "matrix":
        payload = get_price_history_matrix(
            q=q,
            from_date=from_date,
            to_date=to_date,
            granularity=granularity,
            include_fallback=include_fallback,
            limit=MATRIX_EXPORT_LIMIT,
            offset=0,
            template_path=template_path,
        )
        if fmt == "csv":
            return _matrix_csv(payload), f"price_timeline_matrix_{stamp}.csv", "text/csv"
        return _matrix_xlsx(payload), f"price_timeline_matrix_{stamp}.xlsx", (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    payload = get_price_history_detail_list(
        q=q,
        from_date=from_date,
        to_date=to_date,
        for_export=True,
    )
    if fmt == "csv":
        return _detail_csv(payload), f"price_history_detail_{stamp}.csv", "text/csv"
    return _detail_xlsx(payload), f"price_history_detail_{stamp}.xlsx", (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
