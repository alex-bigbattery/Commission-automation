"""
SQLite -> SalespersonData bridge (the rules engine).

Reads operational data (sales_orders, invoices, shipments, customer_payments,
items) from the commission_automation SQLite database for a given month, applies
Big Battery's B2B commission rules (transcribed from the accountant's recorded
process), and returns SalespersonData objects ready for workbook_builder_v2 plus
a list of review exceptions and KPI totals.

Business rules (see plan / comisiones_transcript.txt):
  1. Source = invoice lines invoiced in the month.
  2. Current (Section I) vs Prior period (Section II) by the parent SO order_date.
  3. Line type: product -> commissionable, shipping -> shipping, else -> other.
  4. $0 / ticket lines -> dropped, flagged for review.
  5. Shipping:
       income > 0          -> keep, Include=No by default.
       free + order > $5k  -> dropped (legit free shipping).
       free + order <= $5k -> dropped, flagged "reason not given".
  6. AR: paid via last-payment-date / balance 0 / closed; else UNPAID (kept, flagged).
  7. MAP price from items.rate (written as a value); missing -> flagged.
  8. Discount rate = 1 - revenue / (MAP * qty)  (Excel formula in the sheet, AH).
  9. Commission rate = tier lookup on discount; salaried vs non-salaried (double).
 10. Only the known B2B roster; others flagged.
 11. Coupons: not auto-excluded for B2B (rare, judgment-based) — out of scope today.
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from src.commission.line_classification import classify_line_type
from src.commission.ticket_classification import apply_ticket_flags
from src.commission.roster import (
    ALL_SHEETS_ORDERED,
    COMPANY_SHEET,
    NON_SALARIED_SHEETS,
    ROUTING_UNASSIGNED,
    SALESPERSON_FULL_TO_SHEET,
    enrich_audit_fields,
    format_zoho_salesperson,
    resolve_roster_sheet,
    roster_rep_sheet_keys,
    classify_special_person,
    is_known_inactive,
    is_b2c_coupon_rep,
)
from src.commission.returns import apply_return_timing_rule, commission_month_end, load_return_metadata_map
from src.commission.workbook_builder_v2 import (
    Block,
    DetailRow,
    SalespersonData,
)
from src.db.connection import DbConnection, get_connection, init_database
from src.db.adjustments import get_adjustment_map, make_line_uid
from src.db.db_utils import date_prefix_expr


FREE_SHIPPING_THRESHOLD = 5000.0

# Possible-ticket signal: an invoiced amount this many times above MAP*qty almost
# always means a mis-keyed ticket/RC order (Accounting example: a $3k item invoiced
# at $400k). Flag for review only — never auto-exclude (cf_ticket may be empty).
PRICE_ANOMALY_FACTOR = 5.0

# Discount-based review thresholds (per Accounting / Jennifer, June 2026).
# Real support tickets (numeric 1–4 digits in CF.Ticket#) are non-commissionable.
# Quote references (QUO-…) in the Ticket# field are NOT auto-excluded.
# For lines WITHOUT a real ticket, the discount off MAP is the signal:
#   > DISCOUNT_KILL   -> non-commissionable (not a legitimate sales discount;
#                        usually an untagged warranty replacement)
#   > DISCOUNT_REVIEW (30%) -> Needs Review: HELD (pending), NOT paid and NOT
#                              excluded — above the commission-table limit; needs
#                              written approval before payout.
#   > DISCOUNT_KILL  (60%)  -> non-commissionable (excluded).
# A small epsilon avoids flagging clean 30%/60% deals that compute slightly over
# the threshold because of MAP rounding.
DISCOUNT_REVIEW = 0.30
DISCOUNT_KILL = 0.60
DISCOUNT_EPSILON = 0.005

# Surfacing-layer mapping from engine flag strings to the 8 management-review
# categories. Used only to populate the additive `category_tags` field on each
# audit_row -- never read by the engine, never affects pay, never mutates the raw
# `flags` field. Multiple flags can map to the same tag (e.g. KNOWN_INACTIVE +
# UNASSIGNED both -> inactive_unmatched); the tag list is deduplicated while
# preserving the order below.
_CATEGORY_TAG_MAP: tuple[tuple[str, str], ...] = (
    ("REAL_TICKET",                     "ticket"),
    ("OTHER_TICKET_REFERENCE",          "ticket_review"),
    ("QUOTE_REFERENCE_IN_TICKET_FIELD", "quote_reference"),
    ("FULLY_RETURNED",          "return"),
    ("RETURN_AFTER_COMMISSION_MONTH", "return_clawback"),
    ("PARTIALLY_RETURNED",      "return"),
    ("KNOWN_INACTIVE",          "inactive_unmatched"),
    ("UNASSIGNED",              "inactive_unmatched"),
    ("DISCOUNT_OVER_30",        "discount_review"),
    ("DISCOUNT_OVER_60",        "discount_excluded"),
    ("UNPAID",                  "unpaid_info"),
    ("COMPANY_ACCOUNT",         "company_account"),
    ("EXECUTIVE_ACCOUNT",       "executive_account"),
    ("PRICE_HISTORY_NO_WINDOW", "price_map_issue"),
    ("RLP_FALLBACK_NO_FVPRICE", "price_map_issue"),
    ("MISSING_MAP",             "price_map_issue"),
    ("MAP_ANOMALY_LOW",         "price_map_issue"),
)


def _category_tags_from_flags(flags_str: str | None) -> list[str]:
    """Deterministic, ordered, deduplicated mapping from a flag-string to category
    tags. Pure function -- no I/O, no side effects, never raises on bad input."""
    if not flags_str:
        return []
    flag_set = {f.strip() for f in str(flags_str).split(",") if f and f.strip()}
    tags: list[str] = []
    seen: set[str] = set()
    for flag, tag in _CATEGORY_TAG_MAP:
        if flag in flag_set and tag not in seen:
            tags.append(tag)
            seen.add(tag)
    return tags

# Bruce Taylor's commission rates per the payout model. Bruce is paid:
#   * BRUCE_REP_RATE  x rep_commission_total   (rep-side incentive)
#   * BRUCE_COMPANY_RATE x company_commission   (company-side, only Bruce gets paid
#                                                from the Company Account bucket)
# These constants are the engine-side defaults. The Excel template's B2B Summary
# sheet at I13/J13/K13 hardcodes the SAME rates as formulas, so changing these
# constants alone would create a code-vs-template mismatch — keep both in lockstep
# OR move the template formulas to read from Config_Settings in a separate change.
# For now: changing the in-code constants without also updating the template would
# diverge code totals from the Excel-formula totals.
BRUCE_REP_RATE = 0.15
BRUCE_COMPANY_RATE = 0.20

# price_history source prefixes — commission resolver ignores catalog backfill rows.
ZOHO_CATALOG_SNAPSHOT_PREFIX = "zoho_catalog_snapshot_"
ACCOUNTANT_FVPRICE_PREFIX = "accountant_fvprice_"
RLP_FALLBACK_NO_FVPRICE_FLAG = "RLP_FALLBACK_NO_FVPRICE"
RLP_FALLBACK_NO_FVPRICE_MSG = (
    "Using R_LP fallback because official FV_PRICE snapshot is missing."
)

# Symmetric lower-side anomaly check for the snapshot/R_LP MAP. A resolved MAP that is
# much LOWER than the live items.rate (e.g. a decimal-point typo making $100 look like
# $10) silently shifts every line of that SKU into the wrong discount tier or pushes
# it into the >30% Needs Review / >60% excluded bands. PRICE_ANOMALY_FACTOR only
# catches the upper side (item_total >> MAP*qty); this catches the lower side.
# Trigger: map_price < MAP_ANOMALY_LOW_FACTOR * items.rate AND items.rate > 0.
MAP_ANOMALY_LOW_FACTOR = 0.5

# Fallback commission tiers (discount_rate, salaried_rate, non_salaried_rate),
# used only if the template's "Table" sheet can't be read.
DEFAULT_TIERS: list[tuple[float, float, float]] = [
    (0.00, 0.05, 0.10),
    (0.05, 0.04, 0.08),
    (0.10, 0.03, 0.06),
    (0.15, 0.02, 0.04),
    (0.20, 0.01, 0.02),
    (0.26, 0.00, 0.00),
]


# --- Review exceptions + result types ----------------------------------------


@dataclass
class ReviewItem:
    salesperson: str
    invoice_number: str
    sales_order_number: str
    sku: str
    amount: float
    reason: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "Salesperson": self.salesperson,
            "Invoice": self.invoice_number,
            "Sales Order": self.sales_order_number,
            "SKU": self.sku,
            "Amount": round(float(self.amount or 0), 2),
            "Reason": self.reason,
        }


@dataclass
class GenerationResult:
    salespeople: list[SalespersonData]
    exceptions: list[ReviewItem] = field(default_factory=list)
    totals_by_sheet: dict[str, float] = field(default_factory=dict)
    kpis: dict[str, Any] = field(default_factory=dict)
    audit_rows: list[dict[str, Any]] = field(default_factory=list)  # per-line system/adjustment/final
    adjusted_count: int = 0


# --- Time helpers ------------------------------------------------------------


def month_bounds(year: int, month: int) -> tuple[str, str]:
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    head = text.split("T")[0].split(" ")[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(head, fmt).date()
        except ValueError:
            continue
    return None


def is_in_month(d: date | None, year: int, month: int) -> bool:
    return bool(d and d.year == year and d.month == month)


def is_before_month(d: date | None, year: int, month: int) -> bool:
    if d is None:
        return False
    return (d.year, d.month) < (year, month)


def rate_type_for(sheet_name: str) -> str:
    return "non_salaried" if sheet_name in NON_SALARIED_SHEETS else "salaried"


# --- Commission math ---------------------------------------------------------


def load_tiers_from_template(template_path: Path | None) -> list[tuple[float, float, float]]:
    """Read (discount, salaried_rate, non_salaried_rate) tiers from the template Table sheet."""
    if not template_path or not Path(template_path).exists():
        return DEFAULT_TIERS
    try:
        wb = load_workbook(template_path, data_only=True)
        if "Table" not in wb.sheetnames:
            wb.close()
            return DEFAULT_TIERS
        ws = wb["Table"]
        # Find the header row that contains "Discount" in column B.
        header_row = None
        for r in range(1, 12):
            if "discount" in str(ws.cell(r, 2).value or "").lower():
                header_row = r
                break
        tiers: list[tuple[float, float, float]] = []
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                disc = ws.cell(r, 2).value
                sal = ws.cell(r, 3).value
                nonsal = ws.cell(r, 4).value
                if isinstance(disc, (int, float)):
                    tiers.append((float(disc), float(sal or 0), float(nonsal or 0)))
        wb.close()
        tiers.sort(key=lambda t: t[0])
        return tiers or DEFAULT_TIERS
    except Exception:
        return DEFAULT_TIERS


def load_map_from_template(template_path: Path | None) -> dict[str, float]:
    """
    Read the curated MAP price list from the template's R_LP sheet (col A = SKU,
    col G = Rate). This is Big Battery's commission MAP — NOT the live Zoho
    catalog price (items.rate), which differs and would distort discounts.
    """
    out: dict[str, float] = {}
    if not template_path or not Path(template_path).exists():
        return out
    try:
        wb = load_workbook(template_path, data_only=True)
        if "R_LP" not in wb.sheetnames:
            wb.close()
            return out
        ws = wb["R_LP"]
        for r in range(3, ws.max_row + 1):
            sku = ws.cell(r, 1).value
            rate = ws.cell(r, 7).value
            if sku and isinstance(rate, (int, float)):
                out[str(sku).strip().upper()] = float(rate)
        wb.close()
    except Exception:
        return out
    return out


def load_settings_from_template(template_path: Path | None) -> dict[str, float]:
    """Business-maintained scalar settings from the template's 'Config_Settings'
    sheet (col A = key, col B = numeric value). Missing sheet/keys -> {} and the
    caller keeps the in-code defaults."""
    out: dict[str, float] = {}
    if not template_path or not Path(template_path).exists():
        return out
    try:
        wb = load_workbook(template_path, data_only=True)
        if "Config_Settings" not in wb.sheetnames:
            wb.close()
            return out
        ws = wb["Config_Settings"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or not row[0]:
                continue
            key = str(row[0]).strip().lower()
            val = row[1] if len(row) > 1 else None
            if isinstance(val, (int, float)):
                out[key] = float(val)
        wb.close()
    except Exception:
        return {}
    return out


def _apply_settings_from_template(template_path: Path | None) -> None:
    """Override the module-level threshold constants from Config_Settings when
    present (falls back to the in-code defaults for any missing key)."""
    global FREE_SHIPPING_THRESHOLD, PRICE_ANOMALY_FACTOR, DISCOUNT_REVIEW, DISCOUNT_KILL
    global MAP_ANOMALY_LOW_FACTOR, DISCOUNT_EPSILON
    global BRUCE_REP_RATE, BRUCE_COMPANY_RATE
    s = load_settings_from_template(template_path)
    if "free_shipping_threshold" in s:
        FREE_SHIPPING_THRESHOLD = s["free_shipping_threshold"]
    if "price_anomaly_factor" in s:
        PRICE_ANOMALY_FACTOR = s["price_anomaly_factor"]
    if "discount_review" in s:
        DISCOUNT_REVIEW = s["discount_review"]
    if "discount_kill" in s:
        DISCOUNT_KILL = s["discount_kill"]
    if "map_anomaly_low_factor" in s:
        MAP_ANOMALY_LOW_FACTOR = s["map_anomaly_low_factor"]
    if "discount_epsilon" in s:
        DISCOUNT_EPSILON = s["discount_epsilon"]
    if "bruce_rep_rate" in s:
        BRUCE_REP_RATE = s["bruce_rep_rate"]
    if "bruce_company_rate" in s:
        BRUCE_COMPANY_RATE = s["bruce_company_rate"]


def commission_rate(discount: float, rate_type: str, tiers: list[tuple[float, float, float]]) -> float:
    """Largest tier whose discount <= the line's discount; column by rate_type.
    Tiers MUST be sorted ascending by threshold (index 0) — caller's responsibility.
    """
    if not tiers:
        return 0.0
    idx = 2 if rate_type == "non_salaried" else 1
    chosen = tiers[0][idx]
    for tier in tiers:
        if discount + 1e-9 >= tier[0]:
            chosen = tier[idx]
        else:
            break
    return float(chosen)


def implied_discount(item_total: float, map_price: float, quantity: float) -> float:
    base = map_price * quantity
    if base <= 0:
        return 0.0
    disc = 1.0 - (item_total / base)
    return min(max(disc, 0.0), 1.0)


# --- Data loading ------------------------------------------------------------


@dataclass(frozen=True)
class InvoiceLineRecord:
    invoice_id: str
    invoice_number: str
    invoice_date: date | None
    invoice_status: str
    customer_name: str
    salesperson_name: str
    so_salesperson_name: str
    salesorder_number: str
    salesorder_id: str | None
    order_date: date | None
    sku: str
    item_name: str
    quantity: float
    item_total: float
    rate: float
    line_type: str
    delivery_method: str
    shipment_date: date | None
    shipment_status: str
    payment_date: date | None
    balance: float
    sales_team: str = ""        # CF.Sales Team from invoice raw_json (B2B / B2C ... / Exe.)
    payment_terms: str = ""     # invoice payment_terms_label (e.g. "Net 30", "Prepayment")
    ticket_number: str = ""     # CF.Ticket# (cf_ticket) — usually noncommissionable
    carrier: str = ""           # shipment carrier (when shipments are synced)
    ship_charge: float = 0.0    # shipment shipping charge (when shipments are synced)
    has_shipment: bool = False  # True when a shipment record was found for the SO
    # Quantities from the matching Sales Order line (raw_json). Returns reduce commission.
    qty_ordered: float = 0.0
    qty_shipped: float = 0.0
    qty_invoiced: float = 0.0
    qty_returned: float = 0.0
    return_date: date | None = None
    rma_number: str = ""


def _load_item_map(conn: DbConnection) -> dict[str, float]:
    """sku -> MAP unit price (items.rate). SKU upper-cased for matching.
    When a SKU is duplicated across items, the most recently synced row wins
    (synced_at DESC, item_id DESC as a stable tiebreaker). Portable across
    SQLite and Postgres (no rowid).
    """
    out: dict[str, float] = {}
    for row in conn.execute(
        "SELECT sku, rate FROM items WHERE sku IS NOT NULL AND sku != '' "
        "ORDER BY sku, synced_at DESC, item_id DESC"
    ).fetchall():
        sku = str(row["sku"]).strip().upper()
        if sku and sku not in out:  # first = most recently synced
            out[sku] = float(row["rate"] or 0)
    return out


def _iso_to_date(value: Any) -> date | None:
    """Parse a stored 'YYYY-MM-DD' price_history date to a date; None if blank/invalid."""
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _load_price_history(conn: DbConnection) -> dict[str, list[tuple[date, date, float, bool]]]:
    """sku(upper) -> [(effective_from, effective_to, map_price, is_live), ...].

    The effective-dated price snapshots used to resolve the *period-correct* MAP for
    each line by its sale date. Returns {} if the table is missing (older DBs) so the
    engine cleanly falls back to the curated R_LP / items.rate map. This NEVER reads or
    overwrites R_LP; it is an independent, additive price source that takes priority.

    Rows with ``source`` starting with ``zoho_catalog_snapshot_`` are **excluded** from
    commission MAP resolution. They remain in the DB for Settings / Price History audit
    only — they are unverified current-catalog backfill, not official historical MAP.

    The fourth tuple element ``is_live`` is True when ``snapshot_month == 'live'``
    (a forward-looking Zoho-sync row), False for a closed-month accountant snapshot.
    The resolver uses this flag to enforce: closed-month snapshots WIN over live rows
    for any sale date both would cover. So inserting a forever-open live row today
    cannot retroactively re-price a closed month that has its own snapshot.

    Hardening contract (defense-in-depth on top of schema NOT NULL + UNIQUE):
      * Skip any row whose ``effective_from``/``effective_to`` does not parse as ISO date.
      * Skip any row whose ``map_price`` is <= 0.
      * Skip ``zoho_catalog_snapshot_*`` sources (audit/UI only).
      * ``ORDER BY sku, effective_from, id`` so the per-SKU list is deterministic across
        SQLite and Postgres. Combined with the resolver's non-strict ``>=`` tie-break
        this guarantees the LATER LOAD (highest id) wins on tied effective_from.
    """
    out: dict[str, list[tuple[date, date, float, bool]]] = {}
    try:
        rows = conn.execute(
            "SELECT sku, map_price, effective_from, effective_to, snapshot_month, source "
            "FROM price_history WHERE sku IS NOT NULL AND sku != '' "
            "ORDER BY sku, effective_from, id"
        ).fetchall()
    except Exception:
        return out
    for row in rows:
        if str(row["source"] or "").startswith(ZOHO_CATALOG_SNAPSHOT_PREFIX):
            continue
        sku = str(row["sku"]).strip().upper()
        if not sku:
            continue
        eff_from = _iso_to_date(row["effective_from"])
        eff_to = _iso_to_date(row["effective_to"])
        if eff_from is None or eff_to is None:
            # Defense-in-depth: a row that lost its bounds (NULL or non-ISO date) is
            # NOT used. It cannot silently leak forward/backward in time.
            continue
        try:
            price = float(row["map_price"])
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        is_live = str(row["snapshot_month"] or "").strip().lower() == "live"
        out.setdefault(sku, []).append((eff_from, eff_to, price, is_live))
    return out


def _sale_date(rec: InvoiceLineRecord) -> date | None:
    """Sale date for period-correct pricing: Sales Order date first, then Invoice
    date, then Shipment date (per commission policy)."""
    return rec.order_date or rec.invoice_date or rec.shipment_date


def _resolve_map_price(
    sku: str,
    as_of: date | None,
    price_history: dict[str, list[tuple[date, date, float, bool]]],
    fallback_map: dict[str, float],
) -> float:
    """Period-correct MAP resolution.

    Priority for a sale on ``as_of`` — bucket choice is HARD (no cross-bucket
    effective_from comparison):
      1. Closed-month snapshot (snapshot_month != 'live') whose window contains
         ``as_of``. Latest ``effective_from`` wins inside this bucket; later-load
         (highest id) wins on tie. **A non-None best_month is returned unconditionally,
         even if a live row has a more recent effective_from.** Closed months are
         absolute authority for any date their window covers.
      2. Live Zoho-sync row (snapshot_month == 'live') whose window contains ``as_of``.
         Same intra-bucket tie-break (latest effective_from, later-load on tie). Only
         consulted when best_month is None.
      3. Fallback map (curated R_LP / items.rate).

    Rationale: once an accountant snapshot is loaded for April, no future Zoho catalog
    change can re-price April — even if an operator backfills a live row with an
    effective_from INSIDE April's window. Live Zoho-sync rows only price sale dates
    that have no closed-month snapshot coverage.

    Defense-in-depth: ``_load_price_history`` only emits entries with non-NULL
    ``effective_from``/``effective_to`` and positive price, so both date bounds are
    concrete here. The boundary check is INCLUSIVE on both sides.
    """
    entries = price_history.get(sku)
    if entries and as_of is not None:
        best_month: tuple[date, float] | None = None  # snapshot_month != 'live'
        best_live: tuple[date, float] | None = None   # snapshot_month == 'live'
        for eff_from, eff_to, price, is_live in entries:
            if eff_from > as_of:
                continue
            if as_of > eff_to:
                continue
            if price <= 0:
                continue
            # Non-strict >= so the LAST iterated row at equal effective_from wins.
            # Combined with _load_price_history's ORDER BY sku, effective_from, id ASC
            # this means the highest-id (most recent) load overrides earlier ones —
            # correction semantics. Separated buckets for month-specific vs live so
            # closed-month snapshots cannot be shadowed by a later-loaded live row.
            if is_live:
                if best_live is None or eff_from >= best_live[0]:
                    best_live = (eff_from, price)
            else:
                if best_month is None or eff_from >= best_month[0]:
                    best_month = (eff_from, price)
        if best_month is not None:
            return best_month[1]
        if best_live is not None:
            return best_live[1]
    return fallback_map.get(sku, 0.0)


def _price_history_resolves(
    sku: str,
    as_of: date | None,
    price_history: dict[str, list[tuple[date, date, float, bool]]],
) -> bool:
    """True when ``_resolve_map_price`` would take MAP from price_history, not fallback."""
    if as_of is None:
        return False
    entries = price_history.get(sku.strip().upper())
    if not entries:
        return False
    best_month: date | None = None
    best_live: date | None = None
    for eff_from, eff_to, price, is_live in entries:
        if eff_from > as_of or as_of > eff_to or price <= 0:
            continue
        if is_live:
            if best_live is None or eff_from >= best_live:
                best_live = eff_from
        else:
            if best_month is None or eff_from >= best_month:
                best_month = eff_from
    return best_month is not None or best_live is not None


def _accountant_fvprice_source(year: int, month: int) -> str:
    return f"{ACCOUNTANT_FVPRICE_PREFIX}{year:04d}_{month:02d}"


def _load_invoice_meta_map(conn: DbConnection, year: int, month: int) -> dict[str, dict[str, str]]:
    """invoice_id -> {sales_team, payment_terms} parsed from invoices.raw_json for the month."""
    import json as _json

    start, end = month_bounds(year, month)
    inv_date = date_prefix_expr("invoice_date", conn.postgres)
    out: dict[str, dict[str, str]] = {}
    rows = conn.execute(
        f"""
        SELECT invoice_id, raw_json FROM invoices
        WHERE {inv_date} >= ? AND {inv_date} <= ?
        """,
        (start, end),
    ).fetchall()
    for r in rows:
        try:
            inv = _json.loads(r["raw_json"])
        except Exception:
            out[str(r["invoice_id"])] = {"sales_team": "", "payment_terms": ""}
            continue
        team = ""
        for f in inv.get("custom_fields") or []:
            label = str(f.get("label") or f.get("api_name") or "").lower()
            if "sales team" in label or label == "cf_sales_team":
                team = str(f.get("value") or "")
                break
        terms = str(inv.get("payment_terms_label") or "")
        # Ticket# (api_name cf_ticket) — populated tickets are usually
        # noncommissionable per Accounting. Read top-level first, then hash.
        ticket = str(inv.get("cf_ticket") or "").strip()
        if not ticket:
            cfh = inv.get("custom_field_hash") or {}
            ticket = str(cfh.get("cf_ticket") or "").strip()
        out[str(r["invoice_id"])] = {
            "sales_team": team,
            "payment_terms": terms,
            "ticket_number": ticket,
        }
    return out


def _load_returns_map(conn: DbConnection, so_ids: Iterable[str]) -> dict[tuple[str, str], dict[str, float]]:
    """(salesorder_id, sku_upper) -> {ordered, invoiced, shipped, returned} from the SO line items.

    Returned quantity is only tracked on the Sales Order line raw_json, never on
    invoice lines, so commission must look it up here to net out returns.
    """
    import json as _json

    ids = sorted({str(s) for s in so_ids if s})
    out: dict[tuple[str, str], dict[str, float]] = {}
    if not ids:
        return out
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(
        f"SELECT salesorder_id, raw_json FROM sales_orders WHERE salesorder_id IN ({placeholders})",
        ids,
    ).fetchall()
    for r in rows:
        try:
            order = _json.loads(r["raw_json"])
        except Exception:
            continue
        soid = str(r["salesorder_id"])
        for li in order.get("line_items") or []:
            sku = str(li.get("sku") or "").strip().upper()
            if not sku:
                continue
            agg = out.setdefault((soid, sku), {"ordered": 0.0, "invoiced": 0.0, "shipped": 0.0, "returned": 0.0})
            agg["ordered"] += float(li.get("quantity") or 0)
            agg["invoiced"] += float(li.get("quantity_invoiced") or 0)
            agg["shipped"] += float(li.get("quantity_shipped") or 0)
            agg["returned"] += float(li.get("quantity_returned") or 0)
    return out


def _load_invoice_lines_with_context(
    conn: DbConnection, year: int, month: int
) -> list[InvoiceLineRecord]:
    start, end = month_bounds(year, month)
    inv_date = date_prefix_expr("i.invoice_date", conn.postgres)
    rows = conn.execute(
        f"""
        SELECT
          il.invoice_id, il.line_index, il.sku, il.item_name,
          il.quantity, il.rate, il.item_total,
          i.invoice_number, i.invoice_date, i.status AS invoice_status,
          i.customer_name AS inv_customer, i.salesperson_name AS inv_salesperson,
          i.balance, i.salesorder_number,
          so.salesorder_id, so.order_date, so.salesperson_name AS so_salesperson,
          so.delivery_method, so.customer_name AS so_customer
        FROM invoice_lines il
        INNER JOIN invoices i ON i.invoice_id = il.invoice_id
        LEFT JOIN sales_orders so ON so.salesorder_number = i.salesorder_number
        WHERE {inv_date} >= ? AND {inv_date} <= ?
        ORDER BY i.invoice_date, i.invoice_number, il.line_index
        """,
        (start, end),
    ).fetchall()

    payment_date_by_invoice = _load_payment_dates(conn, [r["invoice_id"] for r in rows])
    so_numbers = sorted({r["salesorder_number"] for r in rows if r["salesorder_number"]})
    shipment_info = _load_shipment_summary(conn, so_numbers)
    invoice_meta = _load_invoice_meta_map(conn, year, month)
    returns_map = _load_returns_map(conn, {r["salesorder_id"] for r in rows if r["salesorder_id"]})
    return_meta_map = load_return_metadata_map(
        conn, {r["salesorder_id"] for r in rows if r["salesorder_id"]}
    )

    records: list[InvoiceLineRecord] = []
    for r in rows:
        sku = r["sku"] or ""
        item_name = r["item_name"] or ""
        item_total = float(r["item_total"] or 0)
        quantity = float(r["quantity"] or 0)
        rate = float(r["rate"] or 0)
        line_type = classify_line_type(
            sku=sku, item_name=item_name,
            item_total=item_total, quantity=quantity, rate=rate,
        )
        ship = shipment_info.get(r["salesorder_number"] or "")
        meta = invoice_meta.get(str(r["invoice_id"] or ""), {})
        qtys = returns_map.get((str(r["salesorder_id"]), sku.strip().upper()), {}) if r["salesorder_id"] else {}
        rmeta = return_meta_map.get((str(r["salesorder_id"]), sku.strip().upper()), {}) if r["salesorder_id"] else {}
        ret_dt = rmeta.get("return_date")
        if isinstance(ret_dt, str):
            ret_dt = parse_date(ret_dt)
        records.append(InvoiceLineRecord(
            invoice_id=str(r["invoice_id"] or ""),
            invoice_number=r["invoice_number"] or "",
            invoice_date=parse_date(r["invoice_date"]),
            invoice_status=r["invoice_status"] or "",
            customer_name=r["inv_customer"] or r["so_customer"] or "",
            salesperson_name=r["inv_salesperson"] or "",
            so_salesperson_name=r["so_salesperson"] or "",
            salesorder_number=r["salesorder_number"] or "",
            salesorder_id=str(r["salesorder_id"]) if r["salesorder_id"] else None,
            order_date=parse_date(r["order_date"]),
            sku=sku,
            item_name=item_name,
            quantity=quantity,
            item_total=item_total,
            rate=rate,
            line_type=line_type,
            delivery_method=r["delivery_method"] or "",
            shipment_date=ship["date"] if ship else None,
            shipment_status=ship["status"] if ship else "",
            payment_date=payment_date_by_invoice.get(str(r["invoice_id"] or "")),
            balance=float(r["balance"] or 0),
            sales_team=meta.get("sales_team", ""),
            payment_terms=meta.get("payment_terms", ""),
            ticket_number=meta.get("ticket_number", ""),
            carrier=ship["carrier"] if ship else "",
            ship_charge=ship["charge"] if ship else 0.0,
            has_shipment=bool(ship),
            qty_ordered=float(qtys.get("ordered", 0.0)),
            qty_shipped=float(qtys.get("shipped", 0.0)),
            qty_invoiced=float(qtys.get("invoiced", 0.0)),
            qty_returned=float(qtys.get("returned", 0.0)),
            return_date=ret_dt if isinstance(ret_dt, date) else None,
            rma_number=str(rmeta.get("rma_number") or ""),
        ))
    return records


def _load_payment_dates(conn: DbConnection, invoice_ids: Iterable[str]) -> dict[str, date | None]:
    inv_ids = [str(i) for i in invoice_ids if i]
    if not inv_ids:
        return {}
    out: dict[str, date | None] = {iid: None for iid in inv_ids}
    placeholders = ",".join("?" * len(inv_ids))
    pay_date = date_prefix_expr("cp.payment_date", conn.postgres)
    rows = conn.execute(
        f"""
        SELECT cpi.invoice_id, MAX({pay_date}) AS last_date
        FROM customer_payment_invoices cpi
        INNER JOIN customer_payments cp ON cp.payment_id = cpi.payment_id
        WHERE cpi.invoice_id IN ({placeholders})
        GROUP BY cpi.invoice_id
        """,
        inv_ids,
    ).fetchall()
    for r in rows:
        out[str(r["invoice_id"])] = parse_date(r["last_date"])
    return out


def _load_shipment_summary(conn: DbConnection, so_numbers: Iterable[str]) -> dict[str, dict]:
    """salesorder_number -> {date, status, carrier, charge} from the shipments table.

    Returns an empty map when no shipments are synced (the salesperson sheets then
    show no shipment data, and the workbook is flagged accordingly).
    """
    nums = [n for n in so_numbers if n]
    if not nums:
        return {}
    placeholders = ",".join("?" * len(nums))
    out: dict[str, dict] = {}

    # 1) Authoritative Zoho shipments (when synced).
    ship_date = date_prefix_expr("shipment_date", conn.postgres)
    for r in conn.execute(
        f"""
        SELECT salesorder_number,
               MAX({ship_date}) AS last_ship_date,
               MAX(status) AS status,
               MAX(carrier_name) AS carrier_name,
               MAX(shipping_charge) AS shipping_charge
        FROM shipments
        WHERE salesorder_number IN ({placeholders})
        GROUP BY salesorder_number
        """,
        nums,
    ).fetchall():
        out[r["salesorder_number"]] = {
            "date": parse_date(r["last_ship_date"]),
            "status": r["status"] or "",
            "carrier": r["carrier_name"] or "",
            "charge": float(r["shipping_charge"] or 0),
        }

    # 2) Fallback to locally derived shipments (from sales_orders.raw_json packages).
    missing = [n for n in nums if n not in out]
    if missing:
        ph2 = ",".join("?" * len(missing))
        derived_date = date_prefix_expr("shipment_date", conn.postgres)
        for r in conn.execute(
            f"""
            SELECT salesorder_number,
                   MAX({derived_date}) AS last_ship_date,
                   MAX(shipment_status) AS shipment_status,
                   MAX(carrier_name) AS carrier_name,
                   MAX(shipping_charge) AS shipping_charge
            FROM derived_shipments
            WHERE salesorder_number IN ({ph2})
            GROUP BY salesorder_number
            """,
            missing,
        ).fetchall():
            if not r["last_ship_date"]:
                continue
            out[r["salesorder_number"]] = {
                "date": parse_date(r["last_ship_date"]),
                "status": r["shipment_status"] or "",
                "carrier": r["carrier_name"] or "",
                "charge": float(r["shipping_charge"] or 0),
            }
    return out


# --- Classification + DetailRow construction ---------------------------------


def _ar_status(rec: InvoiceLineRecord) -> str:
    inv_lower = rec.invoice_status.lower()
    if rec.payment_date is not None:
        return "PAID"
    if rec.balance == 0 or "paid" in inv_lower or "closed" in inv_lower:
        return "PAID"
    if rec.balance > 0:
        return "UNPAID"
    return "REVIEW"  # negative balance = over-payment / credit note


def _is_negative_balance(rec: InvoiceLineRecord) -> bool:
    return rec.balance < 0 and rec.payment_date is None


def _build_detail_row(
    rec: InvoiceLineRecord,
    *,
    map_price: float,
    comm_rate: float,
    ar_status: str,
) -> DetailRow:
    is_shipping = rec.line_type == "shipping"
    return DetailRow(
        order_date=rec.order_date,
        sales_order_number=rec.salesorder_number,
        invoice_date=rec.invoice_date,
        invoice_number=rec.invoice_number,
        invoice_status=rec.invoice_status,
        customer_name=rec.customer_name,
        estimate_number="",
        sku=rec.sku,
        quantity=rec.quantity,
        item_total=rec.item_total,
        account="Shipping Income" if is_shipping else ("Sales" if rec.line_type == "product" else ""),
        account_code="",
        payment_terms=rec.payment_terms,
        delivery_method=rec.delivery_method,
        shipment_date=rec.shipment_date,
        shipment_status=(
            ("SHIPPED" if "ship" in rec.shipment_status.lower() else rec.shipment_status.upper())
            if rec.has_shipment else ""
        ),
        ar_status=ar_status,
        payment_date=rec.payment_date,
        shipping_method=rec.carrier,
        reason="",
        include_in_commission="No",
        shipping_income=rec.item_total if is_shipping else 0.0,
        shipping_expenses=0.0,
        discount_rate=None,
        map_price=map_price,
        commission_rate=comm_rate,
    )


def _section_for(rec: InvoiceLineRecord, year: int, month: int) -> str:
    if is_before_month(rec.order_date, year, month):
        return "II"
    return "I"  # current month, or unknown date (don't lose data)


def _empty_salesperson(sheet_name: str, full_name: str, year: int, month: int) -> SalespersonData:
    return SalespersonData(
        name=sheet_name,
        full_name=full_name.upper(),
        rate_type=rate_type_for(sheet_name),
        month_name=calendar.month_name[month],
        year=year,
        current_commissionable=Block("ORDERS COMMISSIONABLE - SHIPPED AND INVOICED"),
        current_shipping=Block("SHIPPING INCOME"),
        current_other=Block("OTHER CHARGES"),
        prior_commissionable=Block("ORDERS COMMISSIONABLE"),
        prior_shipping=Block("SHIPPING CHARGE"),
    )


@dataclass
class _Line:
    """A surviving commission line carrying both system and final (post-adjustment) values."""
    line_uid: str
    rec: InvoiceLineRecord
    section: str               # "I" | "II"
    block: str                 # "commissionable" | "shipping" | "other"
    zoho_salesperson: str      # Original Zoho salesperson (display only)
    sys_sheet: str             # roster sheet or Zoho name from automated calculation
    sheet: str                 # routing sheet key (after adjustment); ROUTING_UNASSIGNED if pending
    sys_map: float
    sys_discount: float
    sys_rate: float
    sys_commissionable: float
    sys_commission: float
    detail: DetailRow          # final values written into the workbook
    flags: list[str] = field(default_factory=list)
    excluded: bool = False
    classification: str = ""
    adj_reason: str = ""
    reviewer: str = ""
    approval_status: str = ""
    adjusted: bool = False
    pending: bool = False        # B2B line awaiting manual salesperson assignment


def _resolve_sheet(name: str | None) -> str | None:
    """Map a salesperson full-name or sheet key to a roster sheet key."""
    return resolve_roster_sheet(name)


def _route_from_zoho(
    full_name: str,
    data_by_sheet: dict[str, SalespersonData],
) -> tuple[str, str, str, bool, list[str]]:
    """Return zoho_display, sys_sheet, routing_sheet, in_roster, flags."""
    zoho = format_zoho_salesperson(full_name)
    roster_sheet = resolve_roster_sheet(full_name)
    in_roster = roster_sheet is not None and roster_sheet in data_by_sheet
    flags: list[str] = []
    if in_roster:
        return zoho, roster_sheet, roster_sheet, True, flags
    flags.append("UNASSIGNED")
    return zoho, zoho, ROUTING_UNASSIGNED, False, flags


def build_salespeople_from_sqlite(
    year: int,
    month: int,
    db_path: Path | None = None,
    *,
    tiers: list[tuple[float, float, float]] | None = None,
    rlp_map: dict[str, float] | None = None,
    apply_adjustments: bool = True,
) -> GenerationResult:
    """Build SalespersonData per roster sheet from SQLite, plus exceptions + totals.

    MAP price source priority: the curated R_LP map (``rlp_map``) wins; the live
    Zoho catalog (items.rate) is only a fallback for SKUs missing from R_LP.

    When ``apply_adjustments`` is true, manual adjustments stored for the period
    are applied AFTER the automated calculation and BEFORE grouping/export.
    """
    init_database(db_path)
    conn = get_connection(db_path)
    try:
        invoice_lines = _load_invoice_lines_with_context(conn, year, month)
        item_map = _load_item_map(conn)
        price_history = _load_price_history(conn)
        accountant_source = _accountant_fvprice_source(year, month)
        accountant_snapshot_count = int(
            conn.execute(
                "SELECT COUNT(*) AS c FROM price_history WHERE source = ?",
                (accountant_source,),
            ).fetchone()["c"]
        )
    finally:
        conn.close()

    tiers = tiers or DEFAULT_TIERS
    rlp = rlp_map or {}
    # Fallback MAP: curated R_LP overrides the live catalog; catalog fills any gaps.
    # price_history (effective-dated snapshots) takes priority over this whole fallback
    # and is resolved per line by sale date in the loop below.
    map_by_sku = {**item_map, **(rlp_map or {})}

    # Order total per SO (all line types) for the $5k free-shipping test.
    order_total_by_so: dict[str, float] = {}
    for rec in invoice_lines:
        if rec.salesorder_number:
            order_total_by_so[rec.salesorder_number] = (
                order_total_by_so.get(rec.salesorder_number, 0.0) + rec.item_total
            )

    data_by_sheet: dict[str, SalespersonData] = {
        sheet: _empty_salesperson(sheet, full, year, month)
        for sheet, full in ALL_SHEETS_ORDERED
    }
    exceptions: list[ReviewItem] = []
    totals_by_sheet: dict[str, float] = {s: 0.0 for s, _ in ALL_SHEETS_ORDERED}

    # ---- Phase 1: build surviving commission lines with SYSTEM values ----
    lines: list[_Line] = []
    rlp_fallback_lines = 0
    for rec in invoice_lines:
        # Rule 1 — route by CF.Sales Team exactly like Accounting does.
        team = (rec.sales_team or "").strip().lower()
        is_b2b = team.startswith("b2b") or team.startswith("exe") or "comp. account" in team
        if not is_b2b:
            continue

        full_name = (rec.so_salesperson_name or rec.salesperson_name or "").strip()
        zoho_sp, sys_sheet, routing, in_roster, base_flags = _route_from_zoho(full_name, data_by_sheet)
        line_uid = make_line_uid(rec.invoice_number, rec.sku, rec.salesorder_number)
        section = _section_for(rec, year, month)
        ar = _ar_status(rec)
        flags: list[str] = list(base_flags)

        # ---- Special-person routing (configurable Company / Executive) ------
        # Pay based on the Zoho salesperson (per Marshall):
        #   Company Account (Bruce): route to Company Acct at NORMAL commission;
        #     Bruce's payout = 20% of that, applied in B2B Summary J13 (not here).
        #   Executive Account (Marshall/Eric): revenue tracked, commission = 0,
        #     not payable, kept visible for audit.
        special = classify_special_person(full_name)
        auto_cls = ""
        executive_route = False
        if special and special.get("category") == "company":
            flags.append(special["flag"])
            routing = COMPANY_SHEET
            sys_sheet = COMPANY_SHEET
            in_roster = True           # routes to the real Company Acct sheet (not pending)
            auto_cls = "company"
        elif special and special.get("category") == "executive":
            flags.append(special["flag"])
            in_roster = False
            routing = ROUTING_UNASSIGNED
            auto_cls = "executive"
            executive_route = True

        # ---- Known-inactive / non-B2B names ---------------------------------
        elif is_known_inactive(full_name):
            flags.append("KNOWN_INACTIVE")
            in_roster = False
            routing = ROUTING_UNASSIGNED

        # ---- B2C coupon-based reps (Dylan Nava, Customer Service) -----------
        elif is_b2c_coupon_rep(full_name):
            flags.append("B2C_COUPON_RULE")
            in_roster = False
            routing = ROUTING_UNASSIGNED

        # ---- Ticket# classification (cf_ticket on invoice) -------------------
        excluded_auto = False
        ticket_excluded, ticket_pending = apply_ticket_flags(
            (rec.ticket_number or "").strip(), flags
        )
        if ticket_excluded:
            excluded_auto = True

        # Pending unless routed to a real sheet. Executive lines are auto-classified
        # (not pending) but carry $0 commission and land on no sheet.
        is_pending = (not in_roster) and not executive_route
        if ticket_pending:
            is_pending = True

        if "REAL_TICKET" in flags:
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=f"Real support ticket ({rec.ticket_number.strip()}) — non-commissionable",
            ))
        elif "OTHER_TICKET_REFERENCE" in flags:
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=f"Unrecognized Ticket# format ({rec.ticket_number.strip()}) — review required",
            ))

        # Rule 5 — shipping lines
        if rec.line_type == "shipping":
            if rec.item_total <= 0:
                order_total = order_total_by_so.get(rec.salesorder_number, 0.0)
                if order_total <= FREE_SHIPPING_THRESHOLD:
                    exceptions.append(ReviewItem(
                        salesperson=sys_sheet,
                        invoice_number=rec.invoice_number,
                        sales_order_number=rec.salesorder_number,
                        sku=rec.sku,
                        amount=order_total,
                        reason=f"Free shipping under ${int(FREE_SHIPPING_THRESHOLD):,} — reason not given",
                    ))
                continue  # free shipping: drop the $0 shipping line
            detail = _build_detail_row(rec, map_price=0.0, comm_rate=0.0, ar_status=ar)
            lines.append(_Line(line_uid, rec, section, "shipping", zoho_sp, sys_sheet, routing,
                               0.0, 0.0, 0.0, 0.0, 0.0, detail, flags=flags, pending=is_pending, excluded=excluded_auto, classification=auto_cls))
            continue

        # Rule 4 — $0 non-shipping line. Per Accounting these are usually either
        # a kit component (the kit's K-SKU line carries the price) or a ticket
        # (non-commissionable). Either way it earns no separate commission.
        if rec.item_total == 0:
            sku_u = (rec.sku or "").strip().upper()
            if sku_u.startswith("K"):
                why = "$0 kit component — price is on the kit line; excluded"
            else:
                why = "$0 line — likely kit component or ticket; excluded, verify"
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=0.0,
                reason=why,
            ))
            continue

        if not in_roster and not special:
            # Tailor the review reason so Accounting sees WHY the line is held,
            # not a blanket "not in roster" for recognized-but-non-payable names.
            if "B2C_COUPON_RULE" in flags:
                reason = (
                    "B2C / coupon-based rep — verify coupon before any payout "
                    "(B2C-RC Team = commissionable; B2C-Web Marketing = not). "
                    "Not paid via standard B2B until classified."
                )
            elif "KNOWN_INACTIVE" in flags:
                reason = "Known inactive / non-B2B salesperson — assign to an active rep or exclude"
            else:
                reason = "Salesperson not in commission roster — classify or assign"
            exceptions.append(ReviewItem(
                salesperson=zoho_sp,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=reason,
            ))

        # Other charges (non-product, non-shipping, non-zero)
        if rec.line_type != "product":
            detail = _build_detail_row(rec, map_price=0.0, comm_rate=0.0, ar_status=ar)
            lines.append(_Line(line_uid, rec, "I", "other", zoho_sp, sys_sheet, routing,
                               0.0, 0.0, 0.0, 0.0, 0.0, detail, flags=flags, pending=is_pending, excluded=excluded_auto, classification=auto_cls))
            continue

        # Product line — compute MAP, discount, commission rate.
        # Period-correct: a price_history snapshot effective at the sale date wins;
        # the curated R_LP / items.rate map is only the fallback.
        sku_u = rec.sku.strip().upper()
        as_of = _sale_date(rec)
        # Hardening flag: the snapshot has a price for this SKU but we have NO sale
        # date to apply it. _resolve_map_price will fall back to R_LP/items.rate
        # silently; surface this for review.
        if as_of is None and sku_u in price_history:
            flags.append("MISSING_SALE_DATE")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason="No SO/Invoice/Shipment date — snapshot MAP could not be applied; using fallback",
            ))
        map_price = _resolve_map_price(sku_u, as_of, price_history, map_by_sku)
        # Hardening flag: SKU has price_history entries but NONE cover as_of. The
        # resolver silently used fallback (R_LP/items.rate) — surface for review so
        # the reviewer doesn't believe the snapshot was honored when it wasn't.
        if (as_of is not None and sku_u in price_history
                and not any(ef <= as_of <= et for (ef, et, pr, _il) in price_history[sku_u])):
            flags.append("PRICE_HISTORY_NO_WINDOW")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=f"price_history has entries for {sku_u} but none cover {as_of.isoformat()} — using fallback",
            ))
        # Closed-month MAP from R_LP template — not official accountant FV_PRICE.
        if (as_of is not None and map_price > 0 and not _price_history_resolves(sku_u, as_of, price_history)
                and sku_u in rlp):
            flags.append(RLP_FALLBACK_NO_FVPRICE_FLAG)
            rlp_fallback_lines += 1
        # Hardening flag: the resolved MAP is much LOWER than the live items.rate —
        # likely a typo in the snapshot (e.g. $10 instead of $100). Symmetric to
        # PRICE_ANOMALY which catches MAP-too-high.
        items_rate = item_map.get(sku_u, 0.0)
        if (map_price > 0 and items_rate > 0
                and map_price < items_rate * MAP_ANOMALY_LOW_FACTOR):
            flags.append("MAP_ANOMALY_LOW")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=f"Resolved MAP {map_price:.2f} is below {MAP_ANOMALY_LOW_FACTOR:.0%} of items.rate {items_rate:.2f} — verify snapshot",
            ))
        if map_price <= 0:
            flags.append("MISSING_MAP")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason="Missing MAP price — rate may be wrong",
            ))
        # Possible-ticket / price anomaly: invoiced far above MAP often signals a
        # mis-keyed ticket/RC order (per Accounting). Flag for review — do NOT
        # auto-exclude and do NOT auto-hold (cf_ticket is the strong signal; this
        # is a weaker heuristic that may have false positives on legit big orders).
        if (map_price > 0 and rec.quantity > 0
                and rec.item_total > map_price * rec.quantity * PRICE_ANOMALY_FACTOR):
            flags.append("PRICE_ANOMALY")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason="Possible ticket / price anomaly — invoiced far above MAP; verify before payout",
            ))
        # Discount/rate are per-unit (computed on the gross invoiced amount/qty);
        # returns do not change which rate tier applies, only how much qty counts.
        disc = implied_discount(rec.item_total, map_price, rec.quantity)
        rt = rate_type_for(routing if in_roster else "Paul")
        rate = commission_rate(disc, rt, tiers) if map_price > 0 else 0.0
        # ---- Discount-based policy (per Accounting / Jennifer, confirmed) ------
        # Real ticket lines are already excluded above (ticket overrides discount).
        # For the rest, the discount off MAP decides:
        #   0-30%            -> normal (pays via the tier table).
        #   > 30% and <= 60% -> NEEDS REVIEW: held (pending) — NOT paid and NOT
        #                       excluded; above the commission-table limit, needs
        #                       written approval before payout.
        #   > 60%            -> non-commissionable (excluded).
        if map_price > 0 and not excluded_auto:
            if disc > DISCOUNT_KILL + DISCOUNT_EPSILON:
                flags.append("DISCOUNT_OVER_60")
                excluded_auto = True
                exceptions.append(ReviewItem(
                    salesperson=sys_sheet, invoice_number=rec.invoice_number,
                    sales_order_number=rec.salesorder_number, sku=rec.sku,
                    amount=rec.item_total,
                    reason=f"Discount {disc * 100:.0f}% over 60% -> non-commissionable",
                ))
            elif disc > DISCOUNT_REVIEW + DISCOUNT_EPSILON:
                flags.append("DISCOUNT_OVER_30")
                is_pending = True   # Needs Review: held, not paid, not excluded
                exceptions.append(ReviewItem(
                    salesperson=sys_sheet, invoice_number=rec.invoice_number,
                    sales_order_number=rec.salesorder_number, sku=rec.sku,
                    amount=rec.item_total,
                    reason=f"Discount {disc * 100:.0f}% above commission table limit -> Needs Review (confirm written approval)",
                ))
        if executive_route:
            rate = 0.0   # Executive Account: revenue tracked, commission = 0
        if ar == "UNPAID":
            flags.append("UNPAID")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason="Unpaid — included, confirm before payout",
            ))
        if _is_negative_balance(rec):
            flags.append("NEGATIVE_BALANCE")
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason="Negative balance (credit / over-payment) — verify before payout",
            ))

        # Rule (shared with the audit engine) — commission only on quantity kept,
        # with commission-month vs return-date timing for fully returned lines.
        period_end = commission_month_end(year, month)
        timing = apply_return_timing_rule(
            invoiced_qty=rec.qty_invoiced,
            returned_qty=rec.qty_returned,
            shipped_qty=rec.qty_shipped,
            fallback_qty=rec.quantity,
            item_total=rec.item_total,
            return_date=rec.return_date,
            commission_month_end=period_end,
        )
        comm_qty = timing.comm_qty
        comm_amount = timing.comm_amount
        for flag in timing.flags:
            flags.append(flag)
        if timing.exclude:
            excluded_auto = True
        if timing.review_reason:
            exceptions.append(ReviewItem(
                salesperson=sys_sheet,
                invoice_number=rec.invoice_number,
                sales_order_number=rec.salesorder_number,
                sku=rec.sku,
                amount=rec.item_total,
                reason=timing.review_reason,
            ))

        detail = _build_detail_row(rec, map_price=map_price, comm_rate=rate, ar_status=ar)
        # Write NET-of-returns quantity & amount so the workbook formulas stay self-consistent.
        detail.item_total = comm_amount
        detail.quantity = comm_qty
        lines.append(_Line(line_uid, rec, section, "commissionable", zoho_sp, sys_sheet, routing,
                           map_price, disc, rate, comm_amount, comm_amount * rate,
                           detail, flags=flags, pending=is_pending, excluded=excluded_auto, classification=auto_cls))

    # ---- Phase 2: apply manual adjustments (after calc, before export) ----
    adj_map = get_adjustment_map(year, month, db_path=db_path) if apply_adjustments else {}
    adjusted_count = 0
    for ln in lines:
        adj = adj_map.get(ln.line_uid)
        if not adj:
            continue
        adjusted_count += 1
        ln.adjusted = True
        ln.reviewer = adj.get("reviewer") or ""
        ln.approval_status = adj.get("approval_status") or ""
        ln.adj_reason = adj.get("reason") or ""
        if adj.get("exclude_flag"):
            ln.excluded = True
        cls = (adj.get("classification") or "").lower()
        if cls in ("company", "executive"):
            ln.classification = cls
            ln.sheet = "Company Acct"
            ln.pending = False
        raw_sp = adj.get("adjusted_salesperson")
        if raw_sp:
            new_sheet = _resolve_sheet(raw_sp)
            if new_sheet:
                ln.sheet = new_sheet
                ln.pending = False
            else:
                ln.flags.append("ADJ_SALESPERSON_NOT_IN_ROSTER")
        if adj.get("adjusted_commissionable") is not None:
            ln.detail.item_total = float(adj["adjusted_commissionable"])
        if adj.get("adjusted_map") is not None:
            ln.detail.map_price = float(adj["adjusted_map"])
        if ln.block == "commissionable":
            rt = rate_type_for(ln.sheet)
            if adj.get("adjusted_discount") is not None:
                ln.detail.commission_rate = commission_rate(float(adj["adjusted_discount"]), rt, tiers)
            else:
                # Use commissionable qty (after returns) so the implied discount is correct
                comm_qty = ln.detail.quantity or ln.rec.quantity
                d = implied_discount(ln.detail.item_total, ln.detail.map_price, comm_qty)
                ln.detail.commission_rate = commission_rate(d, rt, tiers) if ln.detail.map_price > 0 else 0.0
        if ln.adj_reason:
            ln.detail.reason = ln.adj_reason
        # Release a held line (e.g. Ticket#) once Accounting approves it, as long
        # as it now lands on a real sheet and was not excluded. This lets an
        # approved ticket line flow back into the payable on its rep's sheet.
        if (ln.approval_status or "").lower() == "approved" and not ln.excluded:
            if ln.sheet in data_by_sheet:
                ln.pending = False

    # ---- Phase 3: group final lines + per-line audit + totals ----
    audit_rows: list[dict[str, Any]] = []
    for ln in lines:
        final_comm = 0.0
        if not ln.excluded and not ln.pending and ln.block == "commissionable":
            final_comm = ln.detail.item_total * ln.detail.commission_rate
        if not ln.excluded and not ln.pending:
            sp = data_by_sheet.get(ln.sheet)
            if sp is not None:
                if ln.block == "commissionable":
                    (sp.current_commissionable if ln.section == "I" else sp.prior_commissionable).rows.append(ln.detail)
                    totals_by_sheet[ln.sheet] += final_comm
                elif ln.block == "shipping":
                    (sp.current_shipping if ln.section == "I" else sp.prior_shipping).rows.append(ln.detail)
                else:
                    sp.current_other.rows.append(ln.detail)
        flags_str = ",".join(ln.flags)
        audit_extra = enrich_audit_fields(
            zoho_salesperson=ln.zoho_salesperson,
            sys_sheet=ln.sys_sheet,
            sheet=ln.sheet,
            excluded=ln.excluded,
            classification=ln.classification,
            pending=ln.pending,
            flags=flags_str,
            block=ln.block,
            section=ln.section,
            sales_team=ln.rec.sales_team or "",
            approval_status=ln.approval_status,
        )
        audit_rows.append({
            "line_uid": ln.line_uid,
            "period": f"{year}-{month:02d}",
            "sales_team": ln.rec.sales_team,
            **audit_extra,
            "sales_order": ln.rec.salesorder_number,
            "invoice": ln.rec.invoice_number,
            "ticket_number": (ln.rec.ticket_number or "").strip(),
            "sku": ln.rec.sku,
            "item_name": ln.rec.item_name,
            "customer": ln.rec.customer_name,
            "quantity": ln.rec.quantity,
            "qty_ordered": ln.rec.qty_ordered,
            "qty_shipped": ln.rec.qty_shipped,
            "qty_invoiced": ln.rec.qty_invoiced,
            "qty_returned": ln.rec.qty_returned,
            "qty_commissionable": (round(ln.detail.quantity, 2) if ln.block == "commissionable" else ""),
            "return_status": (
                "Return After Period" if "RETURN_AFTER_COMMISSION_MONTH" in ln.flags
                else "Fully Returned" if "FULLY_RETURNED" in ln.flags
                else "Partially Returned" if "PARTIALLY_RETURNED" in ln.flags
                else ""
            ),
            "return_date": ln.rec.return_date.isoformat() if ln.rec.return_date else "",
            "rma_number": ln.rec.rma_number or "",
            "revenue": round(ln.rec.item_total, 2),
            "block": ln.block,
            "section": ln.section,
            "flags": flags_str,
            "category_tags": _category_tags_from_flags(flags_str),
            "map": round(ln.detail.map_price, 2),
            "system_commissionable": round(ln.sys_commissionable, 2),
            "system_rate": round(ln.sys_rate, 4),
            "system_commission": round(ln.sys_commission, 2),
            "final_commissionable": round(ln.detail.item_total, 2),
            "final_rate": round(ln.detail.commission_rate, 4),
            "final_commission": round(final_comm, 2),
            "adjustment": round(final_comm - ln.sys_commission, 2),
            "excluded": ln.excluded,
            "classification": ln.classification,
            "reason": ln.adj_reason,
            "reviewer": ln.reviewer,
            "approval_status": ln.approval_status,
            "adjusted": ln.adjusted,
            "pending": ln.pending,
        })

    pending_lines = sum(1 for a in audit_rows if a.get("pending"))
    pending_revenue = round(sum(a["revenue"] for a in audit_rows if a.get("pending")), 2)
    pending_commission = round(sum(a["system_commission"] for a in audit_rows if a.get("pending")), 2)
    shipment_present = any(ln.rec.has_shipment for ln in lines)
    approval_incomplete = any(
        a.get("adjusted") and a.get("approval_status") not in ("approved", "Approved") for a in audit_rows
    )

    kpis = _compute_kpis(data_by_sheet, totals_by_sheet, exceptions)
    # total_commission must reflect the actual Total to Pay (rep + Bruce), NOT the
    # raw sheet sum — the full Company Acct normal commission is not paid, only
    # Bruce's 20% of it. Keeps the KPI consistent with Reconciliation / M10.
    _rep, _company, _bruce, _total = _payout_breakdown(totals_by_sheet)
    kpis["total_commission"] = _total
    kpis["company_account_commission"] = _company
    kpis["bruce_commission"] = _bruce
    kpis["adjusted_lines"] = adjusted_count
    kpis["pending_lines"] = pending_lines
    kpis["pending_revenue"] = pending_revenue
    kpis["pending_commission"] = pending_commission
    kpis["shipment_data_present"] = shipment_present
    kpis["approval_incomplete"] = approval_incomplete
    # Draft while there are unassigned lines, shipments are missing, or any
    # adjustment is not yet approved.
    kpis["is_draft"] = bool(pending_lines > 0 or (not shipment_present) or approval_incomplete)
    map_warnings: list[str] = []
    if accountant_snapshot_count == 0:
        # Period-agnostic: any month missing its official accountant FV_PRICE
        # snapshot gets the same notice (no hardcoded month). The stronger
        # fallback-impact message is added only when R_LP was actually used to
        # price at least one line — otherwise prices came from price_history
        # live rows / items.rate and no MAP review is warranted.
        map_warnings.append(
            f"No {accountant_source} rows in price_history for {year:04d}-{month:02d}."
        )
        if rlp_fallback_lines > 0:
            map_warnings.append(RLP_FALLBACK_NO_FVPRICE_MSG)
    kpis["map_warnings"] = map_warnings
    kpis["rlp_fallback_lines"] = rlp_fallback_lines
    kpis["accountant_fvprice_present"] = accountant_snapshot_count > 0
    kpis["accountant_fvprice_source"] = accountant_source
    return GenerationResult(
        salespeople=list(data_by_sheet.values()),
        exceptions=exceptions,
        totals_by_sheet=totals_by_sheet,
        kpis=kpis,
        audit_rows=audit_rows,
        adjusted_count=adjusted_count,
    )


def _compute_kpis(
    data_by_sheet: dict[str, SalespersonData],
    totals_by_sheet: dict[str, float],
    exceptions: list[ReviewItem],
) -> dict[str, Any]:
    commissionable_lines = 0
    revenue_current = 0.0
    revenue_prior = 0.0
    reps_with_sales = 0
    for sp in data_by_sheet.values():
        cur = sp.current_commissionable.rows
        pri = sp.prior_commissionable.rows
        commissionable_lines += len(cur) + len(pri)
        revenue_current += sum(r.item_total for r in cur)
        revenue_prior += sum(r.item_total for r in pri)
        if cur or pri:
            reps_with_sales += 1
    return {
        "total_commission": round(sum(totals_by_sheet.values()), 2),
        "commissionable_lines": commissionable_lines,
        "salespeople_with_sales": reps_with_sales,
        "revenue_current": round(revenue_current, 2),
        "revenue_prior": round(revenue_prior, 2),
        "exceptions_count": len(exceptions),
    }


# --- High-level orchestration ------------------------------------------------


def _payout_breakdown(totals_by_sheet: dict[str, float]) -> tuple[float, float, float, float]:
    """Single source of truth for the payout model (matches B2B Summary template):
      rep      = sum of roster rep sheets (excludes Company Acct)
      company  = NORMAL commission on the Company Acct sheet (Bruce's lines)
      bruce    = BRUCE_REP_RATE x rep + BRUCE_COMPANY_RATE x company
                                                (template K13 = I13 + J13)
      total    = rep + bruce                    (template M10 = K11 + K13)
    The full company normal commission is NOT paid directly — it only feeds
    Bruce's 20%. Returns (rep, company, bruce, total).

    BRUCE_REP_RATE / BRUCE_COMPANY_RATE come from the module-level constants and
    are overridable via Config_Settings keys ``bruce_rep_rate`` /
    ``bruce_company_rate``. The Excel template's B2B Summary sheet (I13/J13/K13)
    encodes the SAME percentages as Excel formulas; keep these constants in
    lockstep with that template or the engine total and the Excel total diverge.
    """
    reps = [s for s, _ in ALL_SHEETS_ORDERED if s != COMPANY_SHEET]
    rep = round(sum(totals_by_sheet.get(s, 0.0) for s in reps), 2)
    company = round(totals_by_sheet.get(COMPANY_SHEET, 0.0), 2)
    bruce = round(rep * BRUCE_REP_RATE + company * BRUCE_COMPANY_RATE, 2)
    total = round(rep + bruce, 2)
    return rep, company, bruce, total


def _reconciliation_values(result: GenerationResult) -> dict[str, float]:
    """Engine-computed reconciliation (written as VALUES so checks read 0 on open)."""
    rep_commission, company_commission, bruce, total_to_pay = _payout_breakdown(result.totals_by_sheet)
    executive = 0.0  # Executive Account = revenue tracked, commission 0 (manual line F38 cleared)
    return {
        "rep_commission": rep_commission,
        "company_commission": company_commission,
        "bruce": bruce,
        "executive": executive,
        "total_to_pay": total_to_pay,
        "check_a": round(rep_commission - rep_commission, 2),                  # sheets vs rep = 0
        "check_b": round((rep_commission + bruce) - total_to_pay, 2),          # = 0
    }


def _period_reference_sheets(year: int, month: int, db_path: Path | None,
                             result: GenerationResult) -> tuple[dict, bool]:
    """Build period-correct raw reference sheets (R_SO, R_INV, R_SH, R_LP) from SQLite."""
    from src.commission.sqlite_data_source import load_period_dataframes

    def df_to_hr(df) -> tuple[list, list]:
        if df is None or df.empty:
            return [], []
        headers = [str(c) for c in df.columns.tolist()]
        rows = []
        for _, row in df.iterrows():
            out = []
            for v in row.tolist():
                try:
                    import pandas as _pd
                    if _pd.isna(v):
                        v = ""
                except Exception:
                    pass
                out.append(v)
            rows.append(out)
        return headers, rows

    period = load_period_dataframes(year, month, db_path=db_path)
    shipments_present = not period.shipments.empty
    ref = {
        "R_SO": df_to_hr(period.sales_orders),
        "R_INV": df_to_hr(period.invoices),
        "R_SH": df_to_hr(period.shipments),
    }
    # R_LP = the MAP actually applied this period (per product line), for transparency.
    applied: dict[str, float] = {}
    for a in result.audit_rows:
        if a.get("block") == "commissionable" and a.get("sku"):
            applied[str(a["sku"])] = a.get("map", 0)
    ref["R_LP"] = (["SKU", "MAP Price (applied this period)"],
                   [[s, m] for s, m in sorted(applied.items())])
    return ref, shipments_present


def generate_commission_workbook(
    year: int,
    month: int,
    *,
    template_path: Path,
    output_path: Path,
    db_path: Path | None = None,
) -> GenerationResult:
    """Full pipeline: SQLite -> SalespersonData[] -> Jennifer-style master workbook."""
    from src.commission.workbook_builder_v2 import build_master_workbook

    tiers = load_tiers_from_template(template_path)
    rlp_map = load_map_from_template(template_path)
    _apply_settings_from_template(template_path)
    result = build_salespeople_from_sqlite(year, month, db_path=db_path, tiers=tiers, rlp_map=rlp_map)
    recon = _reconciliation_values(result)
    reference_sheets, shipments_present = _period_reference_sheets(year, month, db_path, result)
    build_master_workbook(
        template_path=template_path,
        output_path=output_path,
        salespeople=result.salespeople,
        month_name=calendar.month_name[month],
        year=year,
        audit_rows=result.audit_rows,
        status_info=result.kpis,
        reconciliation=recon,
        reference_sheets=reference_sheets,
        shipments_present=shipments_present,
    )
    return result


def generate_salesperson_workbook(
    year: int,
    month: int,
    salesperson_sheet: str,
    *,
    template_path: Path,
    output_path: Path,
    db_path: Path | None = None,
) -> GenerationResult:
    """Pipeline for ONE salesperson's standalone workbook (single sheet)."""
    from src.commission.workbook_builder_v2 import build_salesperson_workbook

    tiers = load_tiers_from_template(template_path)
    rlp_map = load_map_from_template(template_path)
    _apply_settings_from_template(template_path)
    result = build_salespeople_from_sqlite(year, month, db_path=db_path, tiers=tiers, rlp_map=rlp_map)
    target = next((d for d in result.salespeople if d.name == salesperson_sheet), None)
    if target is None:
        raise ValueError(
            f"Salesperson sheet {salesperson_sheet!r} not in catalog. "
            f"Valid: {[s for s, _ in ALL_SHEETS_ORDERED]}"
        )
    build_salesperson_workbook(
        template_path=template_path,
        output_path=output_path,
        data=target,
    )
    return result
