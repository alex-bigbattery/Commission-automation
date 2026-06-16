"""
B2C (RC Team) commission engine.

Parallel to the B2B engine (``sqlite_to_workbook``) but for the B2C RC Team,
transcribed from the accountant's ``2026-04_Commissions_B2C.xlsx``:

Rules (confirmed against April 2026 actuals):
  1. Eligibility is decided by ``CF.Sales Team``, not the coupon code itself:
       * ``B2C Web - RC Team`` and ``B2C - RC Team``  -> commissionable.
       * ``B2C - RC Team (No Commissionable)``        -> excluded.
       * ``B2C Web - Marketing`` / ``B2C Web - Affiliate`` / organic / B2B / Exe.
                                                       -> not part of B2C commission.
     (The RC Team rides the ``volt5`` coupon, but Sales Team is the authority.)
  2. Commission rate is a FLAT 2% — it does NOT use the tiered B2B %Discount table.
  3. Commissionable amount = net invoiced item total (returns netted out, same
     timing rule as B2B). MAP/discount are computed for display/audit only.
  4. Only product lines earn B2C commission (shipping / $0 lines are ignored).
  5. Pool = sum(commissionable) * rate. The split of that pool between the RC
     Team members (e.g. manager / Dylan / River) CHANGES month to month, so this
     engine does NOT assign who gets paid — it computes the pool and the per-rep
     subtotals and leaves the final split to Accounting review.

This reuses the B2B engine's data loading and period-correct MAP resolution so
the two stay consistent; it never mutates B2B output.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from src.commission.returns import apply_return_timing_rule, commission_month_end
from src.commission.sqlite_to_workbook import (
    _load_invoice_lines_with_context,
    _load_item_map,
    _load_price_history,
    _resolve_map_price,
    _sale_date,
    implied_discount,
)
from src.db.connection import get_connection, init_database


# Flat B2C commission rate (overridable via the ``rate`` argument / settings).
B2C_COMMISSION_RATE = 0.02

# CF.Sales Team values (normalized: lower-cased, stripped) that earn B2C
# commission. The "(no commissionable)" RC variant is deliberately NOT here.
B2C_RC_TEAMS: frozenset[str] = frozenset({
    "b2c web - rc team",
    "b2c - rc team",
})

# Placeholder / catch-all SKUs that carry no MAP and are not real products. The
# accountant excludes these from B2C commission (verified on April 2026:
# INV-05580/05581/05739 "MISCELLANEOUS" lines are absent from the payout). They
# classify as "product" by line type (they have a SKU + amount), so they must be
# filtered explicitly here.
NON_COMMISSIONABLE_SKUS: frozenset[str] = frozenset({
    "MISCELLANEOUS",
})


@dataclass
class B2CResult:
    rows: list[dict[str, Any]] = field(default_factory=list)
    by_salesperson: dict[str, float] = field(default_factory=dict)
    pool_commissionable: float = 0.0
    pool_commission: float = 0.0
    rate: float = B2C_COMMISSION_RATE
    kpis: dict[str, Any] = field(default_factory=dict)


def _is_b2c_rc_team(sales_team: str | None) -> bool:
    return (sales_team or "").strip().lower() in B2C_RC_TEAMS


def build_b2c_commission(
    year: int,
    month: int,
    db_path: Path | None = None,
    *,
    rate: float | None = None,
    rlp_map: dict[str, float] | None = None,
) -> B2CResult:
    """Compute B2C RC-Team commission for the period.

    Returns line-level rows, per-salesperson subtotals, and the pool. The pool
    split between RC members is intentionally left to manual review.
    """
    init_database(db_path)
    conn = get_connection(db_path)
    try:
        invoice_lines = _load_invoice_lines_with_context(conn, year, month)
        item_map = _load_item_map(conn)
        price_history = _load_price_history(conn)
    finally:
        conn.close()

    rate = B2C_COMMISSION_RATE if rate is None else float(rate)
    # MAP fallback: curated R_LP overrides the live catalog; price_history
    # (effective-dated) still wins per line via _resolve_map_price.
    map_by_sku = {**item_map, **(rlp_map or {})}
    period_end = commission_month_end(year, month)

    rows: list[dict[str, Any]] = []
    by_sp: dict[str, float] = defaultdict(float)
    pool_commissionable = 0.0

    for rec in invoice_lines:
        if not _is_b2c_rc_team(rec.sales_team):
            continue
        if rec.line_type != "product":
            continue  # B2C commission is product lines only

        sku_u = rec.sku.strip().upper()
        if sku_u in NON_COMMISSIONABLE_SKUS:
            continue  # placeholder SKU (no MAP) — not commissionable
        as_of = _sale_date(rec)
        map_price = _resolve_map_price(sku_u, as_of, price_history, map_by_sku)

        timing = apply_return_timing_rule(
            invoiced_qty=rec.qty_invoiced,
            returned_qty=rec.qty_returned,
            shipped_qty=rec.qty_shipped,
            fallback_qty=rec.quantity,
            item_total=rec.item_total,
            return_date=rec.return_date,
            commission_month_end=period_end,
        )
        # Fully returned within the month -> not commissionable.
        commissionable = 0.0 if timing.exclude else timing.comm_amount
        comm_qty = 0.0 if timing.exclude else timing.comm_qty
        discount = implied_discount(commissionable, map_price, comm_qty) if map_price > 0 else 0.0
        commission = round(commissionable * rate, 2)

        pool_commissionable += commissionable
        sp = (rec.so_salesperson_name or rec.salesperson_name or "").strip()
        by_sp[sp] += commission

        rows.append({
            "order_date": rec.order_date.isoformat() if rec.order_date else "",
            "salesperson": sp,
            "sales_order": rec.salesorder_number,
            "invoice_date": rec.invoice_date.isoformat() if rec.invoice_date else "",
            "invoice_number": rec.invoice_number,
            "invoice_status": rec.invoice_status,
            "customer": rec.customer_name,
            "sales_team": rec.sales_team,
            "coupon": "",  # CF.Coupon(s) not currently parsed into the record
            "sku": rec.sku,
            "quantity": round(comm_qty, 2),
            "item_total": round(commissionable, 2),
            "map_price": round(map_price, 2),
            "total_map_price": round(map_price * comm_qty, 2),
            "discount_rate": round(discount, 6),
            "commission_rate": rate,
            "commission_amount": commission,
            "return_status": timing.return_status,
            "flags": ",".join(timing.flags),
        })

    pool_commissionable = round(pool_commissionable, 2)
    pool_commission = round(pool_commissionable * rate, 2)
    by_salesperson = {k: round(v, 2) for k, v in sorted(by_sp.items())}

    kpis = {
        "rate": rate,
        "commissionable_lines": len(rows),
        "pool_commissionable": pool_commissionable,
        "pool_commission": pool_commission,
        "rc_members": sorted({r["salesperson"] for r in rows if r["salesperson"]}),
        # The pool split varies monthly and is decided by Accounting; this engine
        # does not assign payouts. Surfaced so the UI can show "split is manual".
        "split_is_manual": True,
    }

    return B2CResult(
        rows=rows,
        by_salesperson=by_salesperson,
        pool_commissionable=pool_commissionable,
        pool_commission=pool_commission,
        rate=rate,
        kpis=kpis,
    )


# --- Excel export ------------------------------------------------------------

# (engine row key, sheet header, number format) for the B2C_Commission detail.
_DETAIL_COLUMNS: tuple[tuple[str, str, str], ...] = (
    ("order_date",       "Order Date",          ""),
    ("salesperson",      "Sales person",        ""),
    ("sales_order",      "Sales Order Number",  ""),
    ("invoice_date",     "Invoice Date",        ""),
    ("invoice_number",   "Invoice Number",      ""),
    ("invoice_status",   "Invoice Status",      ""),
    ("customer",         "Customer Name",       ""),
    ("sku",              "SKU",                 ""),
    ("quantity",         "Quantity",            "#,##0.00"),
    ("item_total",       "Item Total",          "#,##0.00"),
    ("map_price",        "Map Price",           "#,##0.00"),
    ("total_map_price",  "Total Map Price",     "#,##0.00"),
    ("discount_rate",    "Discount Rate",       "0.00%"),
    ("commission_rate",  "Commission Rate",     "0.00%"),
    ("commission_amount","Commission Amount",   "#,##0.00"),
    ("return_status",    "Return Status",       ""),
    ("flags",            "Flags",               ""),
)


def write_b2c_workbook(
    result: B2CResult,
    output_path: Path,
    *,
    year: int,
    month: int,
) -> Path:
    """Write the B2C RC-Team commission report to an .xlsx (Summary + detail).

    Mirrors the accountant's ``…_Commissions_B2C.xlsx`` layout: a Summary sheet
    with the pool and per-rep subtotals (the RC member split is left blank for
    manual entry, since it changes monthly), and a B2C_Commission detail sheet.
    """
    import calendar

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    money = "#,##0.00"

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws["B2"] = f"B2C RC Team - Commission {calendar.month_name[month]} {year}"
    ws["B2"].font = title_font

    ws["B4"] = "Pool"
    ws["B4"].font = bold
    pool_rows = [
        ("Commissionable Amount", result.pool_commissionable, money),
        ("Commission Rate", result.rate, "0.00%"),
        ("Commission Amount (pool)", result.pool_commission, money),
    ]
    r = 5
    for label, value, fmt in pool_rows:
        ws.cell(r, 2, label)
        c = ws.cell(r, 3, value)
        c.number_format = fmt
        r += 1

    r += 1
    ws.cell(r, 2, "Per salesperson (informational)").font = bold
    r += 1
    ws.cell(r, 2, "Salesperson").font = bold
    ws.cell(r, 3, "Commission").font = bold
    r += 1
    for sp, amount in result.by_salesperson.items():
        ws.cell(r, 2, sp)
        ws.cell(r, 3, amount).number_format = money
        r += 1

    r += 1
    note = ws.cell(
        r, 2,
        "NOTE: the split of the pool between RC Team members changes each month "
        "and is entered manually by Accounting — it is not auto-calculated here.",
    )
    note.font = Font(italic=True)
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22

    # ---- B2C_Commission detail sheet ----
    ws2 = wb.create_sheet("B2C_Commission")
    for ci, (_key, header, _fmt) in enumerate(_DETAIL_COLUMNS, start=1):
        cell = ws2.cell(1, ci, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ri, row in enumerate(result.rows, start=2):
        for ci, (key, _header, fmt) in enumerate(_DETAIL_COLUMNS, start=1):
            cell = ws2.cell(ri, ci, row.get(key))
            if fmt:
                cell.number_format = fmt

    # Total row at the bottom (Item Total + Commission Amount).
    total_row = len(result.rows) + 2
    item_col = next(i for i, (k, _h, _f) in enumerate(_DETAIL_COLUMNS, 1) if k == "item_total")
    comm_col = next(i for i, (k, _h, _f) in enumerate(_DETAIL_COLUMNS, 1) if k == "commission_amount")
    ws2.cell(total_row, item_col - 1, "TOTAL").font = bold
    tc = ws2.cell(total_row, item_col, round(result.pool_commissionable, 2))
    tc.font = bold
    tc.number_format = money
    cc = ws2.cell(total_row, comm_col, round(result.pool_commission, 2))
    cc.font = bold
    cc.number_format = money

    # Reasonable column widths.
    for ci, (_key, header, _fmt) in enumerate(_DETAIL_COLUMNS, start=1):
        from openpyxl.utils import get_column_letter
        ws2.column_dimensions[get_column_letter(ci)].width = max(12, min(len(header) + 2, 24))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
