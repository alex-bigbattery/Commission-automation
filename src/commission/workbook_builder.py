"""
Commission workbook builder.

DEPRECATED / LEGACY. The production pipeline uses ``workbook_builder_v2``
(``build_master_workbook`` / ``build_salesperson_workbook``), which fills the real
template — including a fully formula-driven B2B Summary. This module is the original
from-scratch builder and is retained only for ``scripts/test_workbook_builder.py``.
Its ``build_b2b_summary_placeholder`` is an intentional stub, NOT unfinished
production code. Do not extend this for new work; use v2.

Generates the Jennifer-style commission workbook (B2B Summary + one sheet per salesperson)
from operational data, with dynamic row counts per block.

The output workbook keeps the layout Accounting is used to:
  - per-salesperson sheet with sections I (current period) and II (prior periods)
  - each section split into sub-blocks (Orders Commissionable, Shipping, Other Charges)
  - subtotal row anchored at the bottom of each block
  - top-of-sheet summary cells with formulas pointing to those anchors
  - B2B Summary sheet aggregating all salespeople

Because anchor row numbers depend on how many rows each block has,
they are computed at build time and injected into formulas dynamically.
"""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- Layout constants --------------------------------------------------------

# Detail block column layout (cols A..AL = 1..38).
DETAIL_HEADERS: dict[str, str] = {
    "D": "Order Date",
    "E": "SalesOrder Number",
    "F": "Invoice Date",
    "G": "Invoice Number",
    "H": "Invoice Status",
    "I": "Customer Name",
    "J": "Estimate Number",
    "K": "SKU",
    "L": "Quantity\nOrdered",
    "M": "Item Total",
    "N": "Account",
    "O": "Account Code",
    "P": "Payment Terms Label",
    "Q": "Delivery Method",
    "S": "Shipment Date",
    "T": "Shipment Status",
    "U": "AR Status",
    "V": "Payment\nDate",
    "X": "Shipping Method",
    "Y": "Reason",
    "Z": "Include in Commission",
    "AA": "Shipping Income",
    "AB": "Shipping Expenses",
    "AC": "Shipping Net",
    "AE": "Map Price",
    "AF": "Total Map Price",
    "AG": "Revenue",
    "AH": "Discount Rate",
    "AJ": "Commissionable\nAmount",
    "AK": "Commission\nRate",
    "AL": "Commission\nAmount",
}

# Columns that receive per-row formulas (formula = f-string with `{r}` placeholder).
ROW_FORMULAS: dict[str, str] = {
    "AC": "=AA{r}-AB{r}",
    "AE": "=IFERROR(VLOOKUP(K{r},R_LP!$A:$G,7,FALSE),0)",
    "AF": "=AE{r}*L{r}",
    "AG": "=M{r}",
    # AH (discount rate) is set per-row from data (manual or derived).
    "AJ": '=IF(Z{r}="Yes",AG{r}+AC{r},AG{r})',
    "AK": "=IFERROR(VLOOKUP(AH{r},Table!$B:$D,3,FALSE),0)",
    "AL": "=AJ{r}*AK{r}",
}

# Columns that get a sum at the subtotal row (SUM over the block range).
SUBTOTAL_SUM_COLS = ("M", "AA", "AB", "AC", "AF", "AG", "AJ", "AL")

# Columns that get weighted average at subtotal: weight by another column.
SUBTOTAL_WEIGHTED_COLS: dict[str, str] = {
    # value_col -> weight_col
    "AH": "AG",   # weighted-avg Discount Rate, weighted by Revenue
    "AK": "AJ",   # weighted-avg Commission Rate, weighted by Commissionable Amount
}

# Salesperson rate type: "salaried" tier table column (Table sheet col index in VLOOKUP),
# vs "non_salaried". Brett uses col 3 (non-salaried), most others col 2 or 3.
SALARIED_TIER_COL = 2
NON_SALARIED_TIER_COL = 3
NON_SALARIED_SALESPEOPLE = {
    "Brett",
    "Brett Bern",
    "Leslie",
    "Leslie Neipert",
    "Carmen",
    "Carmen Daetz",
    "Garrett",
    "Garrett Lockhart",
}


# --- Style helpers -----------------------------------------------------------

NAVY = "1F4E78"
WHITE = "FFFFFF"
YELLOW = "FFFF00"
GREEN = "D9EAD3"
LIGHT_BLUE = "E9F0FE"

_thin = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HEADER_FONT = Font(bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBTOTAL_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor=YELLOW)
SECTION_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)


def _apply_header_style(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER_ALL


def _apply_subtotal_style(cell) -> None:
    cell.font = SUBTOTAL_FONT
    cell.fill = SUBTOTAL_FILL
    cell.border = BORDER_ALL


# --- Data models -------------------------------------------------------------


@dataclass
class DetailRow:
    """One line of detail in a block. Most fields map to columns D..V."""

    order_date: Any = None
    sales_order_number: str = ""
    invoice_date: Any = None
    invoice_number: str = ""
    invoice_status: str = ""
    customer_name: str = ""
    estimate_number: str = ""
    sku: str = ""
    quantity: float = 0.0
    item_total: float = 0.0
    account: str = ""
    account_code: str = ""
    payment_terms: str = ""
    delivery_method: str = ""
    shipment_date: Any = None
    shipment_status: str = ""
    ar_status: str = ""
    payment_date: Any = None
    shipping_method: str = ""
    reason: str = ""
    include_in_commission: str = "No"   # Yes/No (controls AJ formula behavior)
    shipping_income: float = 0.0
    shipping_expenses: float = 0.0
    discount_rate: float | None = None  # if None, will be left blank (no value)


@dataclass
class Block:
    """A detail block (header + rows + subtotal). Position is computed when written."""

    label: str                   # e.g. "ORDERS COMMISSIONABLE - SHIPPED AND INVOICED"
    rows: list[DetailRow] = field(default_factory=list)
    use_subtotal: bool = True


@dataclass
class SalespersonData:
    """All data needed to write one salesperson sheet."""

    name: str                    # e.g. "Brett"
    full_name: str               # e.g. "BRETT BERN"
    rate_type: str               # "salaried" or "non_salaried"
    current_commissionable: Block
    current_shipping: Block
    current_other: Block
    prior_commissionable: Block
    prior_shipping: Block
    month_name: str              # e.g. "March"
    year: int                    # e.g. 2026


@dataclass
class SheetAnchors:
    """Row numbers of subtotal anchors after the sheet is written.

    Summary cells at the top reference these rows by absolute address.
    """

    cur_comm_subtotal: int       # I.1 subtotal row
    cur_ship_subtotal: int       # I.2 subtotal row
    cur_other_subtotal: int      # I.3 subtotal row
    prior_comm_subtotal: int     # II.1 subtotal row
    prior_ship_subtotal: int     # II.2 subtotal row


# --- Sheet writer ------------------------------------------------------------


def _tier_col_for(rate_type: str) -> int:
    return NON_SALARIED_TIER_COL if rate_type == "non_salaried" else SALARIED_TIER_COL


def _write_detail_block(
    ws: Worksheet,
    start_row: int,
    block: Block,
    tier_col: int,
    *,
    min_rows: int = 1,
) -> tuple[int, int, int]:
    """
    Write a detail block: section label row, header row, data rows, subtotal row.

    Returns (header_row, first_data_row, subtotal_row).
    The block always has at least `min_rows` data rows reserved (empty if needed)
    so Excel users can insert/edit without breaking formulas.
    """
    # Section label row
    label_row = start_row
    ws.cell(label_row, 4, block.label).font = SECTION_FONT

    # Header row
    header_row = label_row + 1
    for col_letter, header in DETAIL_HEADERS.items():
        col_idx = _col_letter_to_index(col_letter)
        cell = ws.cell(header_row, col_idx, header)
        _apply_header_style(cell)

    # Data rows
    first_data_row = header_row + 1
    n_rows = max(len(block.rows), min_rows)
    last_data_row = first_data_row + n_rows - 1

    for i, drow in enumerate(block.rows):
        r = first_data_row + i
        _write_detail_row(ws, r, drow, tier_col)

    # Fill any reserved empty rows with row-formulas so Excel insertions inherit them.
    for i in range(len(block.rows), n_rows):
        r = first_data_row + i
        _write_empty_detail_row(ws, r, tier_col)

    # Subtotal row
    subtotal_row = last_data_row + 1
    if block.use_subtotal:
        _write_subtotal_row(ws, subtotal_row, first_data_row, last_data_row)

    return header_row, first_data_row, subtotal_row


def _write_detail_row(ws: Worksheet, r: int, d: DetailRow, tier_col: int) -> None:
    """Write a single populated data row (cols D..AL)."""
    ws.cell(r, _col("D"), d.order_date)
    ws.cell(r, _col("E"), d.sales_order_number)
    ws.cell(r, _col("F"), d.invoice_date)
    ws.cell(r, _col("G"), d.invoice_number)
    ws.cell(r, _col("H"), d.invoice_status)
    ws.cell(r, _col("I"), d.customer_name)
    ws.cell(r, _col("J"), d.estimate_number)
    ws.cell(r, _col("K"), d.sku)
    ws.cell(r, _col("L"), d.quantity)
    ws.cell(r, _col("M"), d.item_total)
    ws.cell(r, _col("N"), d.account)
    ws.cell(r, _col("O"), d.account_code)
    ws.cell(r, _col("P"), d.payment_terms)
    ws.cell(r, _col("Q"), d.delivery_method)
    ws.cell(r, _col("S"), d.shipment_date)
    ws.cell(r, _col("T"), d.shipment_status)
    ws.cell(r, _col("U"), d.ar_status)
    ws.cell(r, _col("V"), d.payment_date)
    ws.cell(r, _col("X"), d.shipping_method)
    ws.cell(r, _col("Y"), d.reason)
    ws.cell(r, _col("Z"), d.include_in_commission or "No")
    ws.cell(r, _col("AA"), d.shipping_income)
    ws.cell(r, _col("AB"), d.shipping_expenses)
    # AH (discount rate) is a manual value
    if d.discount_rate is not None:
        ws.cell(r, _col("AH"), d.discount_rate)
    _inject_row_formulas(ws, r, tier_col)


def _write_empty_detail_row(ws: Worksheet, r: int, tier_col: int) -> None:
    """Empty reserved row that still has formulas so Excel insertions inherit them."""
    ws.cell(r, _col("L"), 0)
    ws.cell(r, _col("M"), 0)
    ws.cell(r, _col("Z"), "No")
    ws.cell(r, _col("AA"), 0)
    ws.cell(r, _col("AB"), 0)
    _inject_row_formulas(ws, r, tier_col)


def _inject_row_formulas(ws: Worksheet, r: int, tier_col: int) -> None:
    """Write the per-row formulas (AC, AE, AF, AG, AJ, AK, AL)."""
    for col_letter, template in ROW_FORMULAS.items():
        formula = template.format(r=r)
        if col_letter == "AK":
            # Use the right tier column for the salesperson.
            formula = f"=IFERROR(VLOOKUP(AH{r},Table!$B:$D,{tier_col},FALSE),0)"
        ws.cell(r, _col(col_letter), formula)


def _write_subtotal_row(ws: Worksheet, sub_row: int, first: int, last: int) -> None:
    """Write the Subtotal row at `sub_row` summing the [first..last] range."""
    ws.cell(sub_row, _col("K"), "Subtotal").font = SUBTOTAL_FONT
    for col_letter in SUBTOTAL_SUM_COLS:
        formula = f"=SUM({col_letter}{first}:{col_letter}{last})"
        cell = ws.cell(sub_row, _col(col_letter), formula)
        _apply_subtotal_style(cell)
    for value_col, weight_col in SUBTOTAL_WEIGHTED_COLS.items():
        formula = (
            f"=IFERROR(SUMPRODUCT(${weight_col}${first}:${weight_col}${last},"
            f"{value_col}{first}:{value_col}{last})/SUM(${weight_col}${first}:${weight_col}${last}),0)"
        )
        cell = ws.cell(sub_row, _col(value_col), formula)
        _apply_subtotal_style(cell)


def _write_summary_top(
    ws: Worksheet,
    data: SalespersonData,
    anchors: SheetAnchors,
) -> None:
    """Write the top-of-sheet summary block (rows 2..18) with formulas pointing to anchors."""
    # Title rows
    ws.cell(2, _col("D"), data.full_name).font = TITLE_FONT
    title_text = f"COMMISSION REPORT - {data.month_name.upper()} {data.year}"
    ws.cell(3, _col("D"), title_text).font = TITLE_FONT
    rate_label = "Non-Salary Commission Rate" if data.rate_type == "non_salaried" else "Salary Commission Rate"
    ws.cell(4, _col("D"), rate_label).font = Font(bold=True, italic=True)

    # Header row 5
    for col_letter, label in [
        ("E", "MAP Price"),
        ("F", "Revenue"),
        ("G", "Discount"),
        ("H", "Commissionable\nAmount"),
        ("I", "Commission\nRate"),
        ("J", "Commission\nAmount"),
    ]:
        cell = ws.cell(5, _col(col_letter), label)
        _apply_header_style(cell)

    # --- CURRENT period block (rows 6-11) ---
    ws.cell(6, _col("D"), "Orders of the period").font = Font(bold=True)

    sub = anchors.cur_comm_subtotal
    ws.cell(7, _col("D"), "Shipped and Invoiced")
    ws.cell(7, _col("E"), f"=AF{sub}")
    ws.cell(7, _col("F"), f"=AG{sub}")
    ws.cell(7, _col("G"), f"=AH{sub}")
    ws.cell(7, _col("H"), f"=AJ{sub}")
    ws.cell(7, _col("I"), f"=AK{sub}")
    ws.cell(7, _col("J"), f"=AL{sub}")

    ws.cell(8, _col("D"), "Shipped and Terms")
    ws.cell(8, _col("F"), 0)

    ws.cell(9, _col("D"), "Pending Shipment")
    ws.cell(9, _col("F"), 0)

    other_refs = []
    if anchors.cur_ship_subtotal:
        other_refs.append(f"M{anchors.cur_ship_subtotal}")
    if anchors.cur_other_subtotal:
        other_refs.append(f"M{anchors.cur_other_subtotal}")
    other_formula = "=" + "+".join(other_refs) if other_refs else 0
    ws.cell(10, _col("D"), "Other Income")
    ws.cell(10, _col("F"), other_formula)

    # Totals row 11
    for col_letter in ("E", "F", "G", "H", "I", "J"):
        cell = ws.cell(11, _col(col_letter), f"=SUM({col_letter}7:{col_letter}10)")
        _apply_subtotal_style(cell)

    # Total Commission box (L10..N12)
    ws.cell(10, _col("L"), "Total Commission").font = Font(bold=True)
    ws.cell(10, _col("N"), "=J7+J14")
    ws.cell(11, _col("L"), "Credits/Advances")
    ws.cell(11, _col("N"), 0)
    ws.cell(12, _col("L"), "Net to Pay").font = Font(bold=True)
    ws.cell(12, _col("N"), "=SUM(N10:N11)").font = Font(bold=True)

    # --- PRIOR period block (rows 13-18) ---
    ws.cell(13, _col("D"), "Orders of previous periods").font = Font(bold=True)

    psub = anchors.prior_comm_subtotal
    ws.cell(14, _col("D"), "Shipped and Invoiced")
    ws.cell(14, _col("E"), f"=AF{psub}")
    ws.cell(14, _col("F"), f"=AG{psub}")
    ws.cell(14, _col("G"), f"=AH{psub}")
    ws.cell(14, _col("H"), f"=AJ{psub}")
    ws.cell(14, _col("I"), f"=AK{psub}")
    ws.cell(14, _col("J"), f"=AL{psub}")

    ws.cell(15, _col("D"), "Shipped and Terms")
    ws.cell(16, _col("D"), "Pending Shipment")
    ws.cell(17, _col("D"), "Other Income")
    if anchors.prior_ship_subtotal:
        ws.cell(17, _col("F"), f"=M{anchors.prior_ship_subtotal}")
    else:
        ws.cell(17, _col("F"), 0)

    for col_letter in ("E", "F", "G", "H", "I", "J"):
        cell = ws.cell(18, _col(col_letter), f"=SUM({col_letter}14:{col_letter}17)")
        _apply_subtotal_style(cell)


def build_salesperson_sheet(
    ws: Worksheet,
    data: SalespersonData,
    *,
    min_rows_per_block: int = 5,
) -> SheetAnchors:
    """
    Build one salesperson sheet. Returns the anchor rows used by summary formulas.

    Layout (positions calculated dynamically):
       R1-R18: Title + summary block (placeholders, fixed positions)
       R20:    Section I header
       R22+:   Block I.1 (label, header, rows..., subtotal)
       blank
       Block I.2 Shipping Income
       blank
       Block I.3 Other Charges
       blank
       Section II header
       Block II.1 Prior Orders Commissionable
       blank
       Block II.2 Prior Shipping
    """
    tier_col = _tier_col_for(data.rate_type)

    # Reserve summary block rows 1..18 for top-of-sheet (formulas filled at the end).
    # Write section I starting at row 20.
    section_i_row = 20
    ws.cell(section_i_row, _col("B"), "I.").font = SECTION_FONT
    section_i_label = f"SALES ORDERS INVOICED IN {data.month_name.upper()} {data.year}"
    ws.cell(section_i_row, _col("C"), section_i_label).font = SECTION_FONT

    # Block I.1
    cursor = section_i_row + 2
    _, _, cur_comm_sub = _write_detail_block(
        ws, cursor, data.current_commissionable, tier_col, min_rows=min_rows_per_block
    )

    cursor = cur_comm_sub + 2
    _, _, cur_ship_sub = _write_detail_block(
        ws, cursor, data.current_shipping, tier_col, min_rows=min_rows_per_block
    )

    cursor = cur_ship_sub + 2
    _, _, cur_other_sub = _write_detail_block(
        ws, cursor, data.current_other, tier_col, min_rows=min_rows_per_block
    )

    # Section II
    section_ii_row = cur_other_sub + 3
    ws.cell(section_ii_row, _col("A"), "II.").font = SECTION_FONT
    ws.cell(section_ii_row, _col("B"), "SALES ORDERS - PRIOR PERIODS").font = SECTION_FONT

    cursor = section_ii_row + 2
    _, _, prior_comm_sub = _write_detail_block(
        ws, cursor, data.prior_commissionable, tier_col, min_rows=min_rows_per_block
    )

    cursor = prior_comm_sub + 2
    _, _, prior_ship_sub = _write_detail_block(
        ws, cursor, data.prior_shipping, tier_col, min_rows=min_rows_per_block
    )

    anchors = SheetAnchors(
        cur_comm_subtotal=cur_comm_sub,
        cur_ship_subtotal=cur_ship_sub,
        cur_other_subtotal=cur_other_sub,
        prior_comm_subtotal=prior_comm_sub,
        prior_ship_subtotal=prior_ship_sub,
    )

    _write_summary_top(ws, data, anchors)

    # Cosmetic widths
    _apply_column_widths(ws)

    return anchors


def _apply_column_widths(ws: Worksheet) -> None:
    widths = {
        "A": 4, "B": 4, "C": 4, "D": 12, "E": 16, "F": 12, "G": 14,
        "H": 12, "I": 28, "J": 14, "K": 16, "L": 8, "M": 12, "N": 14,
        "O": 10, "P": 14, "Q": 14, "S": 12, "T": 14, "U": 10, "V": 12,
        "X": 12, "Y": 18, "Z": 10, "AA": 12, "AB": 12, "AC": 12,
        "AE": 12, "AF": 14, "AG": 12, "AH": 12, "AJ": 14, "AK": 12, "AL": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# --- Reference sheets --------------------------------------------------------


def write_reference_sheet(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Write a flat reference sheet (R_LP, R_SO, R_INV, R_SH)."""
    # Row 2: headers; row 3..: data (matches Jennifer's convention with a blank R1).
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(2, col_idx, header)
        _apply_header_style(cell)
    for r_offset, row in enumerate(rows, start=3):
        for c_offset, value in enumerate(row, start=1):
            ws.cell(r_offset, c_offset, value)


def write_table_sheet(ws: Worksheet, tiers: list[tuple[float, float, float]]) -> None:
    """
    Write the commission tier table.
    `tiers` is a list of (discount_rate, salaried_rate, non_salaried_rate).
    """
    ws.cell(1, 2, "BigBattery - Commission Table").font = Font(bold=True, size=12)
    ws.cell(4, 2, "Discount Rate")
    ws.cell(4, 3, "Salaried")
    ws.cell(4, 4, "Non-Salaried")
    for c in (2, 3, 4):
        _apply_header_style(ws.cell(4, c))
    for i, (disc, sal, nonsal) in enumerate(tiers, start=5):
        ws.cell(i, 2, disc)
        ws.cell(i, 3, sal)
        ws.cell(i, 4, nonsal)


# --- Master workbook orchestrator (skeleton) ---------------------------------


def build_workbook(
    output_path: Path,
    salespeople: list[SalespersonData],
    reference: dict[str, dict[str, Any]] | None = None,
    tier_table: list[tuple[float, float, float]] | None = None,
    include_b2b_summary: bool = True,
) -> Path:
    """Build a complete commission workbook and save to `output_path`."""
    wb = Workbook()
    # Default sheet "Sheet" — repurpose as B2B Summary placeholder later.
    ws_default = wb.active
    wb.remove(ws_default)

    if include_b2b_summary:
        ws_summary = wb.create_sheet("B2B Summary")

    sheet_anchors: dict[str, SheetAnchors] = {}
    for sp in salespeople:
        ws = wb.create_sheet(sp.name)
        anchors = build_salesperson_sheet(ws, sp)
        sheet_anchors[sp.name] = anchors

    # Table + reference sheets
    if tier_table:
        ws_table = wb.create_sheet("Table")
        write_table_sheet(ws_table, tier_table)

    if reference:
        for sheet_name, payload in reference.items():
            ws_ref = wb.create_sheet(sheet_name)
            write_reference_sheet(ws_ref, payload["headers"], payload["rows"])

    if include_b2b_summary:
        # Legacy builder: a minimal summary stub by design. The production
        # formula-driven B2B Summary lives in workbook_builder_v2.
        build_b2b_summary_placeholder(ws_summary, salespeople)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_b2b_summary_placeholder(ws: Worksheet, salespeople: list[SalespersonData]) -> None:
    """Minimal B2B Summary — to be expanded in next iteration."""
    ws.cell(2, 4, "SUMMARY").font = TITLE_FONT
    ws.cell(7, 6, "SUMMARY - B2B Commissions").font = TITLE_FONT
    headers = ["Salesperson", "Revenue", "Commissionable", "Commission Rate", "Commission Payable"]
    for i, h in enumerate(headers, start=5):
        cell = ws.cell(9, i, h)
        _apply_header_style(cell)
    for i, sp in enumerate(salespeople, start=10):
        ws.cell(i, 5, sp.full_name)
        ws.cell(i, 6, f"={sp.name}!F11")
        ws.cell(i, 7, f"={sp.name}!H7")
        ws.cell(i, 8, f"={sp.name}!I7")
        ws.cell(i, 9, f"={sp.name}!J7")


# --- Utilities ---------------------------------------------------------------


def _col(letter: str) -> int:
    return _col_letter_to_index(letter)


def _col_letter_to_index(letter: str) -> int:
    """Convert 'A' -> 1, 'AA' -> 27, 'AL' -> 38."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def month_name(month: int) -> str:
    return calendar.month_name[month]
