"""Read-only commission configuration for the Settings UI.

Aggregates template, database, roster, and code-default sources without mutating
engine globals or writing to any table.
"""
from __future__ import annotations

import os
from datetime import date as _date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from src.commission import roster as roster_mod
from src.commission.sqlite_to_workbook import (
    BRUCE_COMPANY_RATE,
    BRUCE_REP_RATE,
    DEFAULT_TIERS,
    DISCOUNT_EPSILON,
    DISCOUNT_KILL,
    DISCOUNT_REVIEW,
    FREE_SHIPPING_THRESHOLD,
    MAP_ANOMALY_LOW_FACTOR,
    PRICE_ANOMALY_FACTOR,
    load_map_from_template,
    load_settings_from_template,
    load_tiers_from_template,
)
from src.commission.ticket_classification import classify_ticket_number
from src.db.connection import DbConnection, get_connection, using_postgres


# Same value as backend/app.py OVER_5000_THRESHOLD — virtual review annotation only.
OVER_5000_REVIEW_THRESHOLD = 5000.0

_TEMPLATE_SETTING_KEYS: tuple[tuple[str, str, float], ...] = (
    ("free_shipping_threshold", "free_shipping_threshold", FREE_SHIPPING_THRESHOLD),
    ("price_anomaly_factor", "price_anomaly_factor", PRICE_ANOMALY_FACTOR),
    ("discount_review_threshold", "discount_review", DISCOUNT_REVIEW),
    ("discount_exclude_threshold", "discount_kill", DISCOUNT_KILL),
    ("discount_epsilon", "discount_epsilon", DISCOUNT_EPSILON),
    ("map_anomaly_low_factor", "map_anomaly_low_factor", MAP_ANOMALY_LOW_FACTOR),
    ("bruce_rep_rate", "bruce_rep_rate", BRUCE_REP_RATE),
    ("bruce_company_rate", "bruce_company_rate", BRUCE_COMPANY_RATE),
)


def _template_meta(template_path: Path | None) -> dict[str, Any]:
    if not template_path or not template_path.exists():
        return {
            "path": str(template_path) if template_path else None,
            "exists": False,
            "modified_at": None,
        }
    stat = template_path.stat()
    return {
        "path": str(template_path.resolve()),
        "exists": True,
        "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
    }


def _setting_entry(
    display_key: str,
    template_key: str,
    code_default: float,
    template_settings: dict[str, float],
) -> dict[str, Any]:
    if template_key in template_settings:
        return {
            "key": display_key,
            "value": template_settings[template_key],
            "source": "Config_Settings",
            "template_key": template_key,
            "code_default": code_default,
        }
    return {
        "key": display_key,
        "value": code_default,
        "source": "code_default",
        "template_key": template_key,
        "code_default": code_default,
    }


def _rate_table_source(template_path: Path | None) -> str:
    if not template_path or not template_path.exists():
        return "code_default / DEFAULT_TIERS"
    try:
        from openpyxl import load_workbook

        wb = load_workbook(template_path, read_only=True, data_only=True)
        has_table = "Table" in wb.sheetnames
        wb.close()
        if has_table:
            return f"{Path(template_path).name} / Table sheet"
    except Exception:
        pass
    return "code_default / DEFAULT_TIERS"


def get_commission_settings(template_path: Path | None) -> dict[str, Any]:
    """Rate table, policy thresholds, Bruce rates, and ticket policy (read-only)."""
    tpl_meta = _template_meta(template_path)
    template_settings = load_settings_from_template(template_path)

    tiers_raw = load_tiers_from_template(template_path)
    rate_source = _rate_table_source(template_path)

    rate_table = [
        {
            "discount_pct": round(t[0] * 100, 2),
            "salaried_commission_pct": round(t[1] * 100, 2),
            "non_salaried_commission_pct": round(t[2] * 100, 2),
            "effective_from": None,
            "effective_to": None,
            "source": rate_source,
            "version_note": (
                "Current template Table sheet is not effective-dated; "
                "these tiers apply to all periods until versioning is added."
            ),
        }
        for t in tiers_raw
    ]

    policy_thresholds = [
        _setting_entry(display, tpl_key, default, template_settings)
        for display, tpl_key, default in _TEMPLATE_SETTING_KEYS
        if display not in ("bruce_rep_rate", "bruce_company_rate")
    ]
    policy_thresholds.append({
        "key": "over_5000_review_threshold",
        "value": OVER_5000_REVIEW_THRESHOLD,
        "source": "code_default",
        "template_key": None,
        "code_default": OVER_5000_REVIEW_THRESHOLD,
        "note": "Virtual Adjustments/Audit annotation only; does not affect commission math.",
    })

    bruce_rates = {
        "bruce_rep_rate": _setting_entry(
            "bruce_rep_rate", "bruce_rep_rate", BRUCE_REP_RATE, template_settings
        ),
        "bruce_company_rate": _setting_entry(
            "bruce_company_rate", "bruce_company_rate", BRUCE_COMPANY_RATE, template_settings
        ),
        "note": (
            "B2B Summary sheet formulas (I13/J13/K13) must stay in lockstep with these rates."
        ),
    }

    ticket_policy = get_ticket_policy()

    return {
        "read_only": True,
        "template": tpl_meta,
        "rate_table": rate_table,
        "policy_thresholds": policy_thresholds,
        "bruce_rates": bruce_rates,
        "ticket_policy": ticket_policy,
        "hardcoded_constants": {
            "note": "Values shown above merge Config_Settings (when present) with code defaults.",
            "engine_module": "src/commission/sqlite_to_workbook.py",
            "ticket_module": "src/commission/ticket_classification.py",
        },
    }


def get_ticket_policy() -> dict[str, Any]:
    """Static ticket classification rules (mirrors ticket_classification.py)."""
    return {
        "read_only": True,
        "field": "CF.Ticket# (cf_ticket on invoice)",
        "rules": [
            {
                "id": "real_ticket",
                "label": "Real support ticket",
                "match": "Numeric only, length 1–4 digits (e.g. 650, 87, 1234)",
                "flags": ["REAL_TICKET", "TICKET_NUMBER"],
                "auto_exclude": True,
                "force_pending": False,
                "examples": ["650", "87", "1234"],
            },
            {
                "id": "quote_reference",
                "label": "Quote reference",
                "match": "Starts with QUO or QUO- (case-insensitive)",
                "flags": ["QUOTE_REFERENCE_IN_TICKET_FIELD", "TICKET_NUMBER"],
                "auto_exclude": False,
                "force_pending": False,
                "examples": ["QUO-04421", "QUO04421"],
            },
            {
                "id": "other_ticket_reference",
                "label": "Other ticket reference",
                "match": "Any other non-empty value",
                "flags": ["OTHER_TICKET_REFERENCE", "TICKET_NUMBER"],
                "auto_exclude": False,
                "force_pending": True,
                "examples": ["12345", "WARRANTY-99"],
            },
            {
                "id": "none",
                "label": "Blank",
                "match": "Empty or whitespace",
                "flags": [],
                "auto_exclude": False,
                "force_pending": False,
                "examples": [""],
            },
        ],
        "classifier": "classify_ticket_number()",
        "verify_examples": {
            "650": classify_ticket_number("650"),
            "QUO-04421": classify_ticket_number("QUO-04421"),
            "12345": classify_ticket_number("12345"),
            "": classify_ticket_number(""),
        },
    }


def get_roster_settings() -> dict[str, Any]:
    """Roster and people routing config (read-only snapshot at request time)."""
    reps = roster_mod.roster_rep_entries()
    non_salaried = set(roster_mod.NON_SALARIED_SHEETS)
    company_names = set(roster_mod.COMPANY_ACCOUNT_NAMES)
    executive_names = set(roster_mod.EXECUTIVE_ACCOUNT_NAMES)
    inactive_names = set(roster_mod.KNOWN_INACTIVE_NAMES)

    rows: list[dict[str, Any]] = []

    for sheet_key, full_name in reps:
        rows.append({
            "salesperson": full_name,
            "sheet_key": sheet_key,
            "status": "active",
            "pay_type": "non_salaried" if sheet_key in non_salaried else "salaried",
            "company_account": False,
            "executive_account": False,
            "sales_team": None,
            "role": "rep",
        })

    for name in sorted(company_names):
        rows.append({
            "salesperson": name,
            "sheet_key": roster_mod.COMPANY_SHEET,
            "status": "special",
            "pay_type": None,
            "company_account": True,
            "executive_account": False,
            "sales_team": None,
            "role": "company",
        })

    for name in sorted(executive_names):
        rows.append({
            "salesperson": name,
            "sheet_key": None,
            "status": "special",
            "pay_type": None,
            "company_account": False,
            "executive_account": True,
            "sales_team": None,
            "role": "executive",
        })

    for name in sorted(inactive_names, key=str.lower):
        rows.append({
            "salesperson": name,
            "sheet_key": None,
            "status": "inactive",
            "pay_type": None,
            "company_account": False,
            "executive_account": False,
            "sales_team": None,
            "role": "inactive",
        })

    config_source = "Config_People (template)"
    if getattr(roster_mod, "_TPL_PEOPLE", None) is None:
        if roster_mod._parse_env_roster(os.environ.get("COMMISSION_ROSTER", "")):
            config_source = "COMMISSION_ROSTER env"
        else:
            config_source = "code_default"

    return {
        "read_only": True,
        "config_source": config_source,
        "b2c_coupon_reps": sorted(roster_mod.B2C_COUPON_REPS),
        "sales_team_note": (
            "Sales team is per-invoice (Zoho CF.Sales Team), not stored in roster config."
        ),
        "rows": rows,
        "active_rep_count": len(reps),
    }


# Matches zoho_price_history_sync.FAR_FUTURE — open-ended live rows.
FAR_FUTURE = "9999-12-31"
ZOHO_SYNC_PREFIX = "zoho_sync_"
ZOHO_CATALOG_SNAPSHOT_PREFIX = "zoho_catalog_snapshot_"
ACCOUNTANT_FVPRICE_PREFIX = "accountant_fvprice_"
IMPORTED_RLP_PREFIX = "imported_rlp_"
IMPORTED_RLP_CAUTION = "R_LP fallback source — not confirmed FV_PRICE."
ZOHO_CATALOG_CAUTION = (
    "Unverified catalog backfill — not confirmed historical MAP."
)


def _today_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def _is_open_end(effective_to: str) -> bool:
    return str(effective_to or "") >= FAR_FUTURE


def _format_effective_to(effective_to: str) -> str:
    return "Current" if _is_open_end(effective_to) else str(effective_to or "")


def _source_kind(source: str, snapshot_month: str) -> str:
    """Classify price_history provenance for Settings UI badges."""
    src = str(source or "")
    if src.startswith(ZOHO_SYNC_PREFIX):
        return "zoho_live_sync"
    if src.startswith(ZOHO_CATALOG_SNAPSHOT_PREFIX):
        return "zoho_catalog_snapshot"
    if src.startswith(IMPORTED_RLP_PREFIX):
        return "imported_rlp"
    if src.startswith(ACCOUNTANT_FVPRICE_PREFIX):
        return "accountant_fvprice"
    if src.startswith("manual"):
        return "manual"
    snap = str(snapshot_month or "")
    if snap and snap != "live":
        return "other_snapshot"
    return "other"


def _row_active_on(iso_date: str, effective_from: str, effective_to: str) -> bool:
    ef = str(effective_from or "")
    et = str(effective_to or "")
    if not ef:
        return False
    if iso_date < ef:
        return False
    return iso_date <= et or _is_open_end(et)


def _enrich_price_row(row: dict[str, Any], today: str) -> dict[str, Any]:
    source = str(row.get("source") or "")
    snapshot_month = str(row.get("snapshot_month") or "")
    effective_to = str(row.get("effective_to") or "")
    kind = _source_kind(source, snapshot_month)
    is_snapshot = snapshot_month != "live"
    is_open = _is_open_end(effective_to)
    is_current_live = (
        snapshot_month == "live"
        and is_open
        and source.startswith(ZOHO_SYNC_PREFIX)
    )
    is_active = _row_active_on(today, str(row.get("effective_from") or ""), effective_to)
    caution = None
    if kind == "imported_rlp":
        caution = IMPORTED_RLP_CAUTION
    elif kind == "zoho_catalog_snapshot":
        caution = ZOHO_CATALOG_CAUTION
    return {
        **row,
        "effective_to_display": _format_effective_to(effective_to),
        "source_kind": kind,
        "source_caution": caution,
        "is_snapshot": is_snapshot,
        "is_current_live": is_current_live,
        "is_active_for_today": is_active,
    }


def _windows_overlap(a_from: str, a_to: str, b_from: str, b_to: str) -> bool:
    """True when two [from, to] windows share at least one day."""
    if not a_from or not b_from:
        return False
    a_end = FAR_FUTURE if _is_open_end(a_to) else a_to
    b_end = FAR_FUTURE if _is_open_end(b_to) else b_to
    return a_from <= b_end and b_from <= a_end


def _parse_iso_date(value: str) -> _date | None:
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _detect_coverage_gaps(rows: list[dict[str, Any]]) -> list[tuple[str, str]]:
    """Days between consecutive windows with no price_history row."""
    if len(rows) < 2:
        return []
    ordered = sorted(rows, key=lambda r: str(r.get("effective_from") or ""))
    gaps: list[tuple[str, str]] = []
    for i in range(len(ordered) - 1):
        a, b = ordered[i], ordered[i + 1]
        a_to = str(a.get("effective_to") or "")
        b_from = str(b.get("effective_from") or "")
        if _is_open_end(a_to):
            continue
        end_d = _parse_iso_date(a_to)
        start_d = _parse_iso_date(b_from)
        if not end_d or not start_d:
            continue
        gap_start = end_d + timedelta(days=1)
        if gap_start < start_d:
            gaps.append((gap_start.isoformat(), (start_d - timedelta(days=1)).isoformat()))
    return gaps


def _detect_price_history_warnings(rows: list[dict[str, Any]], today: str) -> list[str]:
    warnings: list[str] = []
    if not rows:
        return warnings

    imported_rlp = [r for r in rows if r.get("source_kind") == "imported_rlp"]
    if imported_rlp:
        warnings.append(
            f"{IMPORTED_RLP_CAUTION} ({len(imported_rlp)} row(s) on this SKU.)"
        )

    catalog_backfill = [r for r in rows if r.get("source_kind") == "zoho_catalog_snapshot"]
    if catalog_backfill:
        warnings.append(
            f"{ZOHO_CATALOG_CAUTION} ({len(catalog_backfill)} row(s) on this SKU.)"
        )

    for gap_from, gap_to in _detect_coverage_gaps(rows):
        msg = (
            f"Missing price coverage: {gap_from} to {gap_to} — "
            "commission may use R_LP / items.rate fallback."
        )
        if gap_from <= "2026-06-04" and gap_to >= "2026-06-01":
            msg += " Includes 2026-06-01 to 2026-06-04 before live Zoho sync."
        warnings.append(msg)

    active_today = [r for r in rows if r.get("is_active_for_today")]
    if not active_today:
        warnings.append("No price row is active for today — commission may use R_LP / items.rate fallback.")

    live_current = [r for r in rows if r.get("is_current_live") and r.get("is_active_for_today")]
    if len(live_current) > 1:
        warnings.append(
            f"Multiple current live Zoho rows active for today ({len(live_current)}) — review overlapping windows."
        )

    overlap_pairs: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for i, a in enumerate(rows):
        for b in rows[i + 1:]:
            if _windows_overlap(
                str(a.get("effective_from") or ""),
                str(a.get("effective_to") or ""),
                str(b.get("effective_from") or ""),
                str(b.get("effective_to") or ""),
            ):
                overlap_pairs.append((a, b))

    if overlap_pairs:
        warnings.append(
            f"{len(overlap_pairs)} overlapping effective window(s) detected — shown for review, not auto-corrected."
        )
        for a, b in overlap_pairs[:3]:
            warnings.append(
                f"Overlap: {a.get('effective_from')}–{a.get('effective_to_display')} ({a.get('source')}) "
                f"vs {b.get('effective_from')}–{b.get('effective_to_display')} ({b.get('source')})."
            )

    # Deduplicate while preserving order
    seen: set[str] = set()
    out: list[str] = []
    for w in warnings:
        if w not in seen:
            seen.add(w)
            out.append(w)
    return out


def _summarize_sku_history(hist: list[Any], today: str) -> dict[str, Any] | None:
    if not hist:
        return None
    latest = hist[0]
    active_rows = [
        r for r in hist
        if _row_active_on(today, str(r["effective_from"]), str(r["effective_to"]))
    ]
    current = active_rows[0] if active_rows else latest
    return {
        "sku": str(latest["sku"]),
        "item_id": latest["item_id"],
        "current_price": float(current["map_price"]),
        "latest_effective_from": str(latest["effective_from"]),
        "latest_source": str(latest["source"]),
        "latest_snapshot_month": str(latest["snapshot_month"]),
        "row_count": len(hist),
    }


def _fetch_hist_for_skus(conn: DbConnection, sku_list: list[str]) -> dict[str, list[Any]]:
    if not sku_list:
        return {}
    placeholders = ",".join("?" for _ in sku_list)
    rows = conn.execute(
        f"SELECT sku, item_id, map_price, effective_from, effective_to, source, snapshot_month "
        f"FROM price_history WHERE UPPER(sku) IN ({placeholders}) "
        f"ORDER BY sku, effective_from DESC, id DESC",
        tuple(sku_list),
    ).fetchall()
    grouped: dict[str, list[Any]] = {}
    for row in rows:
        key = str(row["sku"]).upper()
        grouped.setdefault(key, []).append(row)
    return grouped


def _price_history_diagnostics(conn: DbConnection) -> dict[str, Any]:
    """Read-only counts for Settings UI troubleshooting."""
    try:
        row_count = int(conn.execute("SELECT COUNT(*) AS c FROM price_history").fetchone()["c"])
        sku_count = int(
            conn.execute(
                "SELECT COUNT(DISTINCT UPPER(sku)) AS c FROM price_history "
                "WHERE sku IS NOT NULL AND sku != ''"
            ).fetchone()["c"]
        )
    except Exception as exc:
        return {
            "database_backend": "postgres" if using_postgres() else "sqlite",
            "price_history_row_count": None,
            "price_history_sku_count": None,
            "diagnostics_error": str(exc),
        }
    return {
        "database_backend": "postgres" if using_postgres() else "sqlite",
        "price_history_row_count": row_count,
        "price_history_sku_count": sku_count,
        "diagnostics_error": None,
    }


def list_price_history_catalog(
    *,
    q: str = "",
    limit: int = 100,
    offset: int = 0,
) -> dict[str, Any]:
    """Paginated catalog of all SKUs in price_history (for dropdown + browse table)."""
    conn = get_connection()
    today = _today_iso()
    needle = (q or "").strip().upper()
    safe_limit = max(1, min(int(limit), 500))
    safe_offset = max(0, int(offset))

    clauses = ["sku IS NOT NULL", "sku != ''"]
    params: list[Any] = []
    if needle:
        like = f"%{needle}%"
        clauses.append("(UPPER(sku) LIKE ? OR UPPER(COALESCE(item_id, '')) LIKE ?)")
        params.extend([like, like])
    where = " AND ".join(clauses)

    count_row = conn.execute(
        f"SELECT COUNT(DISTINCT UPPER(sku)) AS c FROM price_history WHERE {where}",
        tuple(params),
    ).fetchone()
    total = int(count_row["c"]) if count_row else 0

    sku_rows = conn.execute(
        f"SELECT DISTINCT UPPER(sku) AS sku_u FROM price_history WHERE {where} "
        f"ORDER BY sku_u LIMIT ? OFFSET ?",
        tuple(params + [safe_limit, safe_offset]),
    ).fetchall()
    sku_list = [str(r["sku_u"]) for r in sku_rows]

    grouped = _fetch_hist_for_skus(conn, sku_list)
    results: list[dict[str, Any]] = []
    for sku_u in sku_list:
        summary = _summarize_sku_history(grouped.get(sku_u, []), today)
        if summary:
            results.append(summary)

    return {
        "read_only": True,
        "query": q,
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
        "count": len(results),
        "results": results,
        **_price_history_diagnostics(conn),
    }


def search_price_history(q: str, *, limit: int = 25) -> dict[str, Any]:
    """Autocomplete search across SKU and item_id."""
    if not (q or "").strip():
        return {"read_only": True, "query": q, "results": [], "count": 0}
    safe_limit = max(1, min(int(limit), 100))
    payload = list_price_history_catalog(q=q, limit=safe_limit, offset=0)
    return {
        "read_only": True,
        "query": q,
        "count": payload["count"],
        "results": payload["results"],
    }


def get_price_history_for_sku(
    sku: str,
    *,
    template_path: Path | None = None,
    source: str | None = None,
    snapshot_month: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> dict[str, Any]:
    """Full read-only trajectory for one SKU with enrichment and warnings."""
    conn = get_connection()
    today = _today_iso()
    sku_u = sku.strip().upper()
    if not sku_u:
        return {
            "read_only": True,
            "sku": sku,
            "rows": [],
            "warnings": ["SKU is required."],
            "current_price": None,
            "r_lp_fallback": None,
        }

    raw_rows = conn.execute(
        "SELECT sku, item_id, map_price, effective_from, effective_to, source, "
        "snapshot_month, captured_at "
        "FROM price_history WHERE UPPER(sku) = ? "
        "ORDER BY effective_from ASC, id ASC",
        (sku_u,),
    ).fetchall()

    rows: list[dict[str, Any]] = []
    for r in raw_rows:
        row = {
            "sku": str(r["sku"] or ""),
            "item_id": r["item_id"],
            "map_price": float(r["map_price"]),
            "effective_from": str(r["effective_from"] or ""),
            "effective_to": str(r["effective_to"] or ""),
            "source": str(r["source"] or ""),
            "snapshot_month": str(r["snapshot_month"] or ""),
            "captured_at": str(r["captured_at"] or ""),
        }
        if source and source.strip().lower() not in row["source"].lower():
            continue
        if snapshot_month and snapshot_month.strip() and row["snapshot_month"] != snapshot_month.strip():
            continue
        if date_from and row["effective_to"] and row["effective_to"] < date_from:
            continue
        if date_to and row["effective_from"] and row["effective_from"] > date_to:
            continue
        rows.append(_enrich_price_row(row, today))

    r_lp_map = load_map_from_template(template_path) if template_path else {}
    r_lp_fallback = r_lp_map.get(sku_u)

    warnings = _detect_price_history_warnings(rows, today)
    if not rows:
        warnings.insert(0, f"No price_history rows found for SKU {sku_u}.")
        if r_lp_fallback is not None:
            warnings.append(f"R_LP template fallback MAP available: ${r_lp_fallback:.2f} (not from price_history).")

    active_today = [r for r in rows if r.get("is_active_for_today")]
    live_active = [r for r in active_today if r.get("is_current_live")]
    current_row = live_active[0] if live_active else (active_today[0] if active_today else None)
    current_price = float(current_row["map_price"]) if current_row else None

    sources = sorted({r["source"] for r in rows})
    snapshot_months = sorted({r["snapshot_month"] for r in rows if r["snapshot_month"]})

    return {
        "read_only": True,
        "sku": sku_u,
        "item_id": rows[0]["item_id"] if rows else None,
        "current_price": current_price,
        "r_lp_fallback": r_lp_fallback,
        "row_count": len(rows),
        "sources": sources,
        "snapshot_months": snapshot_months,
        "warnings": warnings,
        "rows": rows,
        "filters_applied": {
            "source": (source or "").strip() or None,
            "snapshot_month": (snapshot_month or "").strip() or None,
            "date_from": (date_from or "").strip() or None,
            "date_to": (date_to or "").strip() or None,
        },
        **_price_history_diagnostics(conn),
    }


def query_price_history(
    *,
    sku: str | None = None,
    snapshot_month: str | None = None,
    limit: int = 500,
    offset: int = 0,
) -> dict[str, Any]:
    """Read-only price_history browse with optional filters."""
    conn = get_connection()

    clauses: list[str] = ["sku IS NOT NULL", "sku != ''"]
    params: list[Any] = []

    if sku and sku.strip():
        clauses.append("UPPER(sku) LIKE ?")
        params.append(f"%{sku.strip().upper()}%")
    if snapshot_month and snapshot_month.strip():
        clauses.append("snapshot_month = ?")
        params.append(snapshot_month.strip())

    where = " AND ".join(clauses)
    count_row = conn.execute(
        f"SELECT COUNT(*) AS c FROM price_history WHERE {where}",
        tuple(params),
    ).fetchone()
    total = int(count_row["c"]) if count_row else 0

    safe_limit = max(1, min(int(limit), 2000))
    safe_offset = max(0, int(offset))

    rows = conn.execute(
        f"SELECT sku, map_price, effective_from, effective_to, source, "
        f"snapshot_month, captured_at, item_id "
        f"FROM price_history WHERE {where} "
        f"ORDER BY sku, effective_from, id "
        f"LIMIT ? OFFSET ?",
        tuple(params + [safe_limit, safe_offset]),
    ).fetchall()

    months_row = conn.execute(
        "SELECT DISTINCT snapshot_month FROM price_history ORDER BY snapshot_month"
    ).fetchall()
    snapshot_months = [str(r["snapshot_month"]) for r in months_row if r["snapshot_month"]]

    return {
        "read_only": True,
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
        "filters": {
            "sku": (sku or "").strip() or None,
            "snapshot_month": (snapshot_month or "").strip() or None,
        },
        "snapshot_months": snapshot_months,
        "rows": [
            _enrich_price_row(
                {
                    "sku": str(r["sku"] or ""),
                    "map_price": float(r["map_price"]),
                    "effective_from": str(r["effective_from"] or ""),
                    "effective_to": str(r["effective_to"] or ""),
                    "source": str(r["source"] or ""),
                    "snapshot_month": str(r["snapshot_month"] or ""),
                    "captured_at": str(r["captured_at"] or ""),
                    "item_id": r["item_id"],
                },
                _today_iso(),
            )
            for r in rows
        ],
    }
